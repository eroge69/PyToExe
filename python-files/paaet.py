import os
import requests
import tkinter as tk
from tkinter import ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import threading
from langdetect import detect
import json


class CustomCombobox(ttk.Combobox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.bind('<1>', self.show_list)

    def show_list(self, event):
        self.event_generate('<Down>')


def submit(event=None):
    selected_location_name = location_combobox.get()
    selected_location_id = locations[selected_location_name]

    location_combobox.config(state='disabled')
    entry_accession_number_start.config(state='disabled', highlightbackground='red')
    entry_accession_number_end.config(state='disabled', highlightbackground='red')
    language_filter.config(state='disabled')
    submit_button.config(text='Fetching Data...', bg='#303030', fg='white', relief='flat')

    lang_filter = language_filter.get()

    threading.Thread(target=fetch_data, args=(
        selected_location_id,
        selected_location_name,
        entry_accession_number_start.get(),
        entry_accession_number_end.get(),
        lang_filter
    )).start()


def fetch_data(selected_location_id, selected_location_name, accession_number_start, accession_number_end, lang_filter):
    headers = {
        'x-okapi-tenant': 'paaet',
        'x-okapi-token': 'eyJhbGciOiJIUzI1NiJ9.eyJzdWIiOiJoZWxwZGVza19hZG1pbiIsInR5cGUiOiJsZWdhY3ktYWNjZXNzIiwidXNlcl9pZCI6Ijk4M2E1OWQyLTNiMTctNGRlOC1hNTJjLTFhZTllMmI0ZjIzYSIsImlhdCI6MTc0NDYxNTk5NCwidGVuYW50IjoicGFhZXQifQ.0q1_3sl0cUcyxyElM6GgVq8KXKzc37OwYGR9qPhYPZA'
        
    }
    extracted_data = []

    for i in range(int(accession_number_start), int(accession_number_end) + 1):
        url = f'https://api02-v1.ils.medad.com/item-storage/items?query=("effectiveLocationId"=="{selected_location_id}" and "accessionNumber"=={i})'
        response = requests.get(url, headers=headers)
        data = response.json()
        items = data['items']

        for item in items:
            accession_number = item.get('accessionNumber')
            holdings_record_id = item.get('holdingsRecordId')
            item_level_call_number = item.get('itemLevelCallNumber')

            url_holdings = f'https://api02-v1.ils.medad.com/holdings-storage/holdings?query=(id=={holdings_record_id})'
            response_holdings = requests.get(url_holdings, headers=headers)
            data_holdings = response_holdings.json()
            holdings_records = data_holdings.get('holdingsRecords', [])

            instance_id = holdings_records[0].get('instanceId') if len(holdings_records) > 0 else None

            if instance_id:
                url_instances = f'https://api02-v1.ils.medad.com/instance-storage/instances?query=(id=={instance_id})'
                response_instances = requests.get(url_instances, headers=headers)
                data_instances = response_instances.json()
                instances = data_instances.get('instances', [])

                if len(instances) > 0:
                    title = instances[0].get('title')
                    try:
                        lang = detect(title)
                    except:
                        title = None
                        lang = None

                    extracted_data.append({
                        'accessionNumber': accession_number,
                        'itemLevelCallNumber': item_level_call_number,
                        'title': title,
                        'language': lang
                    })

    if lang_filter == 'English':
        extracted_data = [item for item in extracted_data if item.get('language', '').startswith('en')]
    elif lang_filter == 'Arabic':
        extracted_data = [item for item in extracted_data if item.get('language', '').startswith('ar')]

    for item in extracted_data:
        if 'language' in item:
            del item['language']

    location_combobox.config(state='readonly')
    entry_accession_number_start.config(state='normal', highlightbackground='white')
    entry_accession_number_end.config(state='normal', highlightbackground='white')
    language_filter.config(state='readonly')
    submit_button.config(text='Submit', bg='SystemButtonFace', fg='black', relief='raised')

    df = pd.DataFrame(extracted_data)
    df.rename(columns={
        'accessionNumber': 'Accession Number',
        'itemLevelCallNumber': 'Call Number',
        'title': 'Title'
    }, inplace=True)

    df = df[['Title', 'Call Number', 'Accession Number']]
    df.index.name = '#'
    df.reset_index(inplace=True)

    def sanitize_filename(filename):
        invalid_chars = ['/', '\\', ':', '*', '?', '<', '>', '|', '"']
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        return filename

    filename = os.path.expanduser(
        f'~/Desktop/{sanitize_filename(selected_location_name)}({accession_number_start} - {accession_number_end}).xlsx')
    df.to_excel(filename, index=False, engine='openpyxl')

    wb = load_workbook(filename)
    ws = wb.active
    ws.column_dimensions['A'].width = 5

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        col_idx = get_column_letter(column_cells[0].column)
        if col_idx != 'A':
            ws.column_dimensions[col_idx].width = length
        for cell in column_cells:
            cell.alignment = Alignment(horizontal='center')

    wb.save(filename)
    os.startfile(filename)


def get_locations():
    headers = {
        'x-okapi-tenant': 'paaet',
        'x-okapi-token': 'eyJhbGciOiJIUzI1NiJ9.eyJzdWIiOiJwYWFldF9hZG1pbiIsInR5cGUiOiJsZWdhY3ktYWNjZXNzIiwidXNlcl9pZCI6IjNiYTdlNWQyLWE2MjEtNGNhNS04NTY0LWIxNzNkOTZmZWYzYiIsImlhdCI6MTcxNTc4MDQxMiwidGVuYW50IjoicGFhZXQifQ.IzG8P5IxxQCJ5Q-QPgLs6nzu7IzZvZbpxywiyvIzsXg'
    }
    response = requests.get('https://api02-v1.ils.medad.com/locations?limit=30', headers=headers)
    data = response.json()
    locations = {location['name']: location['id'] for location in data['locations']}
    return locations


def start_move(event):
    global x, y
    x = event.x
    y = event.y


def stop_move(event):
    global x, y
    x = None
    y = None


def do_move(event):
    global x, y
    dx = event.x - x
    dy = event.y - y
    x = root.winfo_x() + dx
    y = root.winfo_y() + dy
    root.geometry(f'+{x}+{y}')


root = tk.Tk()
root.title('Paaet Book Report')
root.configure(bg='#303030')
root.geometry('300x200')
root.resizable(False, False)

drag_area = tk.Frame(root, bg='#303030')
drag_area.place(x=0, y=0, relwidth=1, relheight=1)
drag_area.lower()
drag_area.bind('<ButtonPress-1>', start_move)
drag_area.bind('<ButtonRelease-1>', stop_move)
drag_area.bind('<B1-Motion>', do_move)

label_effective_location_id = tk.Label(root, text='Location', bg='#303030', fg='white')
label_effective_location_id.pack(pady=(12, 1))

locations = get_locations()
location_names = list(locations.keys())

location_combobox = ttk.Combobox(root, values=location_names, width=32, justify='center')
location_combobox.pack(pady=(5, 5))
location_combobox.current(0)
location_combobox.config(state='readonly')

input_frame = tk.Frame(root, bg='#303030')
input_frame.pack(pady=(10, 10), padx=(0, 10))

label_accession_number_start = tk.Label(input_frame, text='From:', bg='#303030', fg='white')
label_accession_number_start.grid(row=0, column=0, padx=(0, 5))

entry_accession_number_start = tk.Entry(input_frame, fg='black', bg='white', width=10, justify='center')
entry_accession_number_start.grid(row=0, column=1)

label_accession_number_end = tk.Label(input_frame, text='To:', bg='#303030', fg='white')
label_accession_number_end.grid(row=0, column=2, padx=(10, 5))

entry_accession_number_end = tk.Entry(input_frame, fg='black', bg='white', width=10, justify='center')
entry_accession_number_end.grid(row=0, column=3)

language_frame = tk.Frame(root, bg='#303030')
language_frame.pack(pady=(5, 10), padx=(0, 9))

label_language_filter = tk.Label(language_frame, text='Language:', bg='#303030', fg='white')
label_language_filter.grid(row=0, column=0, padx=(0, 2))

language_filter_values = ['English & Arabic', 'English', 'Arabic']
language_filter = ttk.Combobox(language_frame, values=language_filter_values, width=20, justify='center')
language_filter.grid(row=0, column=1)
language_filter.current(0)
language_filter.config(state='readonly')

submit_button = tk.Button(root, text='Submit', command=submit, bg='SystemButtonFace', fg='black', relief='raised')
submit_button.pack(pady=10)

root.bind_all('<Return>', submit)
root.mainloop()

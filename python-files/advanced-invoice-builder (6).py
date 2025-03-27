import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import csv
import os
import sqlite3
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import json
import xml.etree.ElementTree as ET
import openpyxl
from datetime import datetime

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Generator Faktur Budowlanych")
        self.root.geometry("900x700")
        self.root.configure(bg='#f0f0f0')

        # Inicjalizacja bazy danych
        self.init_database()

        # Notebook (Tabs)
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(expand=True, fill="both", padx=15, pady=15)

        # Sections
        self.company_frame = self._create_section("Dane Firmy", self._create_company_section)
        self.client_frame = self._create_section("Dane Klienta", self._create_client_section)
        self.invoice_frame = self._create_section("Pozycje Faktury", self._create_invoice_section)

        # Status bar
        self.status_var = tk.StringVar()
        self.status_bar = tk.Label(root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Add Import and Export Menu
        self.create_menus()

    def create_menus(self):
        """Create menus for importing and exporting different file types"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # Import Menu
        import_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Import", menu=import_menu)
        
        import_options = [
            ("Import JSON", self.import_from_json),
            ("Import XML", self.import_from_xml),
            ("Import Excel", self.import_from_excel),
            ("Import CSV", self.import_from_csv),
            ("Import z Bazy Danych", self.import_from_sqlite)
        ]

        for label, command in import_options:
            import_menu.add_command(label=label, command=command)

        # Export Menu
        export_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Eksport", menu=export_menu)
        
        export_options = [
            ("Eksport PDF", self.export_to_pdf),
            ("Eksport Excel", self.export_to_excel),
            ("Eksport XML", self.export_to_xml),
            ("Eksport JSON", self.export_to_json),
            ("Zapis do Bazy Danych", self.export_to_sqlite)
        ]

        for label, command in export_options:
            export_menu.add_command(label=label, command=command)

    def init_database(self):
        """Inicjalizacja bazy danych SQLite dla przechowywania faktur"""
        self.conn = sqlite3.connect('invoices.db')
        cursor = self.conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY,
                company_data TEXT,
                client_data TEXT,
                invoice_items TEXT,
                created_at DATETIME
            )
        ''')
        self.conn.commit()

    def _create_section(self, title, section_method):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text=title)
        section_method(frame)
        return frame

    def _create_company_section(self, frame):
        labels = [
            "Nazwa Firmy", "NIP", "Adres", "Kod Pocztowy", 
            "Miasto", "Telefon", "Email", "Numer Konta Bankowego"
        ]

        self.company_entries = {}
        for i, label in enumerate(labels):
            tk.Label(frame, text=label, font=('Arial', 10)).grid(row=i, column=0, sticky="w", padx=10, pady=5)
            entry = tk.Entry(frame, width=50, font=('Arial', 10))
            entry.grid(row=i, column=1, padx=10, pady=5)
            self.company_entries[label] = entry

    def _create_client_section(self, frame):
        labels = [
            "Nazwa Klienta", "NIP", "Adres", "Kod Pocztowy", 
            "Miasto", "Telefon", "Email"
        ]

        self.client_entries = {}
        for i, label in enumerate(labels):
            tk.Label(frame, text=label, font=('Arial', 10)).grid(row=i, column=0, sticky="w", padx=10, pady=5)
            entry = tk.Entry(frame, width=50, font=('Arial', 10))
            entry.grid(row=i, column=1, padx=10, pady=5)
            self.client_entries[label] = entry

    def _create_invoice_section(self, frame):
        # Invoice items table
        columns = ("Lp.", "Opis", "Ilość", "Jednostka", "Cena Netto", "Wartość Netto")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings")
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor="center")
        
        self.tree.pack(expand=True, fill="both", padx=10, pady=10)

        # Action buttons
        button_frame = tk.Frame(frame)
        button_frame.pack(pady=10)

        buttons = [
            ("Dodaj Pozycję", self.add_invoice_item),
            ("Usuń Pozycję", self.remove_invoice_item),
            ("Wyczyść", self.clear_invoice_items)
        ]

        for text, command in buttons:
            tk.Button(button_frame, text=text, command=command, font=('Arial', 10)).pack(side="left", padx=5)

    def _clear_all_entries(self):
        """Helper method to clear all entries before importing"""
        # Clear company entries
        for entry in self.company_entries.values():
            entry.delete(0, tk.END)
        
        # Clear client entries
        for entry in self.client_entries.values():
            entry.delete(0, tk.END)
        
        # Clear invoice items
        for item in self.tree.get_children():
            self.tree.delete(item)

    # Export Methods
    def export_to_pdf(self):
        """Eksport faktury do PDF"""
        filename = filedialog.asksaveasfilename(defaultextension=".pdf")
        if filename:
            try:
                doc = SimpleDocTemplate(filename, pagesize=letter)
                elements = []
                styles = getSampleStyleSheet()

                # Nagłówki sekcji
                elements.append(Paragraph("Dane Firmy", styles['Heading2']))
                company_data = [[label, entry.get()] for label, entry in self.company_entries.items()]
                company_table = Table(company_data, colWidths=[2*inch, 4*inch])
                company_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (0,-1), colors.grey),
                    ('TEXTCOLOR', (0,0), (0,-1), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('GRID', (0,0), (-1,-1), 1, colors.black)
                ]))
                elements.append(company_table)

                # Dane klienta
                elements.append(Paragraph("Dane Klienta", styles['Heading2']))
                client_data = [[label, entry.get()] for label, entry in self.client_entries.items()]
                client_table = Table(client_data, colWidths=[2*inch, 4*inch])
                client_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (0,-1), colors.lightblue),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('GRID', (0,0), (-1,-1), 1, colors.black)
                ]))
                elements.append(client_table)

                # Pozycje faktury
                elements.append(Paragraph("Pozycje Faktury", styles['Heading2']))
                invoice_items = [["Lp.", "Opis", "Ilość", "Jednostka", "Cena Netto", "Wartość Netto"]]
                invoice_items.extend([self.tree.item(item)['values'] for item in self.tree.get_children()])
                invoice_table = Table(invoice_items, colWidths=[0.5*inch, 2*inch, inch, inch, inch, inch])
                invoice_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.grey),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('GRID', (0,0), (-1,-1), 1, colors.black)
                ]))
                elements.append(invoice_table)

                doc.build(elements)
                self.status_var.set(f"Faktura PDF zapisana w {filename}")
                messagebox.showinfo("Eksport", f"Faktura zapisana w {filename}")
            except Exception as e:
                messagebox.showerror("Błąd", str(e))

    def export_to_sqlite(self):
        """Zapis faktury do bazy SQLite"""
        try:
            company_data = json.dumps({label: entry.get() for label, entry in self.company_entries.items()})
            client_data = json.dumps({label: entry.get() for label, entry in self.client_entries.items()})
            invoice_items = json.dumps([self.tree.item(item)['values'] for item in self.tree.get_children()])

            cursor = self.conn.cursor()
            cursor.execute('''
                INSERT INTO invoices (company_data, client_data, invoice_items, created_at)
                VALUES (?, ?, ?, ?)
            ''', (company_data, client_data, invoice_items, datetime.now()))
            self.conn.commit()
            self.status_var.set("Faktura zapisana w bazie danych")
            messagebox.showinfo("Zapis", "Faktura zapisana w bazie danych")
        except Exception as e:
            messagebox.showerror("Błąd", str(e))

    def export_to_xml(self):
        """Eksport faktury do XML"""
        filename = filedialog.asksaveasfilename(defaultextension=".xml")
        if filename:
            try:
                root = ET.Element("Invoice")
                
                company = ET.SubElement(root, "CompanyData")
                for label, entry in self.company_entries.items():
                    ET.SubElement(company, label.replace(" ", "_")).text = entry.get()
                
                client = ET.SubElement(root, "ClientData")
                for label, entry in self.client_entries.items():
                    ET.SubElement(client, label.replace(" ", "_")).text = entry.get()
                
                invoice_items = ET.SubElement(root, "InvoiceItems")
                for item in self.tree.get_children():
                    item_elem = ET.SubElement(invoice_items, "Item")
                    values = self.tree.item(item)['values']
                    for i, col in enumerate(["Lp", "Opis", "Ilość", "Jednostka", "Cena_Netto", "Wartość_Netto"]):
                        ET.SubElement(item_elem, col).text = str(values[i])
                
                tree = ET.ElementTree(root)
                tree.write(filename, encoding='utf-8', xml_declaration=True)
                self.status_var.set(f"Faktura XML zapisana w {filename}")
                messagebox.showinfo("Eksport", f"Faktura zapisana w {filename}")
            except Exception as e:
                messagebox.showerror("Błąd", str(e))

    def export_to_excel(self):
        """Eksport faktury do Excel"""
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if filename:
            try:
                wb = openpyxl.Workbook()
                
                # Dane firmy
                ws_company = wb.create_sheet(title="Dane Firmy")
                for i, (label, entry) in enumerate(self.company_entries.items(), start=1):
                    ws_company.cell(row=i, column=1, value=label)
                    ws_company.cell(row=i, column=2, value=entry.get())
                
                # Dane klienta
                ws_client = wb.create_sheet(title="Dane Klienta")
                for i, (label, entry) in enumerate(self.client_entries.items(), start=1):
                    ws_client.cell(row=i, column=1, value=label)
                    ws_client.cell(row=i, column=2, value=entry.get())
                
                # Pozycje faktury
                ws_items = wb.create_sheet(title="Pozycje Faktury")
                headers = ["Lp.", "Opis", "Ilość", "Jednostka", "Cena Netto", "Wartość Netto"]
                for j, header in enumerate(headers, start=1):
                    ws_items.cell(row=1, column=j, value=header)
                
                for i, item in enumerate(self.tree.get_children(), start=2):
                    for j, value in enumerate(self.tree.item(item)['values'], start=1):
                        ws_items.cell(row=i, column=j, value=value)
                
                # Usuń domyślny arkusz
                wb.remove(wb['Sheet'])
                
                wb.save(filename)
                self.status_var.set(f"Faktura Excel zapisana w {filename}")
                messagebox.showinfo("Eksport", f"Faktura zapisana w {filename}")
            except Exception as e:
                messagebox.showerror("Błąd", str(e))

    def export_to_json(self):
        """Eksport faktury do JSON"""
        filename = filedialog.asksaveasfilename(defaultextension=".json")
        if filename:
            try:
                # Prepare data
                data = {
                    "CompanyData": {label: entry.get() for label, entry in self.company_entries.items()},
                    "ClientData": {label: entry.get() for label, entry in self.client_entries.items()},
                    "InvoiceItems": [
                        dict(zip(["Lp.", "Opis", "Ilość", "Jednostka", "Cena Netto", "Wartość Netto"], 
                                 self.tree.item(item)['values'])) 
                        for item in self.tree.get_children()
                    ]
                }
                
                # Write to file
                with open(filename, 'w', encoding='utf-8') as file:
                    json.dump(data, file, ensure_ascii=False, indent=4)
                
                self.status_var.set(f"Faktura JSON zapisana w {filename}")
                messagebox.showinfo("Eksport", f"Faktura zapisana w {filename}")
            except Exception as e:
                messagebox.showerror("Błąd", str(e))

    # Import Methods
    def import_from_json(self):
        """Import invoice data from a JSON file"""
        filename = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                
                # Clear existing entries
                self._clear_all_entries()
                
                # Import company data
                if 'CompanyData' in data:
                    for label, entry in self.company_entries.items():
                        entry.insert(0, data['CompanyData'].get(label, ''))
                
                # Import client data
                if 'ClientData' in data:
                    for label, entry in self.client_entries.items():
                        entry.insert(0, data['ClientData'].get(label, ''))
                
                # Import invoice items
                if 'InvoiceItems' in data:
                    for item in data['InvoiceItems']:
                        values = [
                            str(len(self.tree.get_children()) + 1),
                            item.get('Opis', ''),
                            item.get('Ilość', ''),
                            item.get('Jednostka', ''),
                            item.get('Cena Netto', ''),
                            item.get('Wartość Netto', '')
                        ]
                        self.tree.insert("", "end", values=values)
                
                self.status_var.set(f"Zaimportowano dane z {filename}")
                messagebox.showinfo("Import", f"Dane zaimportowane z {filename}")
            except Exception as e:
                messagebox.showerror("Błąd", str(e))

    def import_from_xml(self):
        """Import invoice data from an XML file"""
        filename = filedialog.askopenfilename(filetypes=[("XML files", "*.xml")])
        if filename:
            try:
                tree = ET.parse(filename)
                root = tree.getroot()
                
                # Clear existing entries
                self._clear_all_entries()
                
                # Import company data
                company_data = root.find('CompanyData')
                if company_data is not None:
                    for label, entry in self.company_entries.items():
                        xml_label = label.replace(" ", "_")
                        elem = company_data.find(xml_label)
                        if elem is not None:
                            entry.insert(0, elem.text or '')
                
                # Import client data
                client_data = root.find('ClientData')
                if client_data is not None:
                    for label, entry in self.client_entries.items():
                        xml_label = label.replace(" ", "_")
                        elem = client_data.find(xml_label)
                        if elem is not None:
                            entry.insert(0, elem.text or '')
                
                # Import invoice items
                invoice_items = root.find('InvoiceItems')
                if invoice_items is not None:
                    for item in invoice_items.findall('Item'):
                        values = [
                            str(len(self.tree.get_children()) + 1),
                            item.find('Opis').text or '',
                            item.find('Ilość').text or '',
                            item.find('Jednostka').text or '',
                            item.find('Cena_Netto').text or '',
                            item.find('Wartość_Netto').text or ''
                        ]
                        self.tree.insert("", "end", values=values)
                
                self.status_var.set(f"Zaimportowano dane z {filename}")
                messagebox.showinfo("Import", f"Dane zaimportowane z {filename}")
            except Exception as e:
                messagebox.showerror("Błąd", str(e))

    def import_from_excel(self):
        """Import invoice data from an Excel file"""
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            try:
                wb = openpyxl.load_workbook(filename)
                
                # Clear existing entries
                self._clear_all_entries()
                
                # Import company data
                if 'Dane Firmy' in wb.sheetnames:
                    ws_company = wb['Dane Firmy']
                    for row in ws_company.iter_rows(min_row=1, max_col=2, values_only=True):
                        if row[0] in self.company_entries:
                            self.company_entries[row[0]].insert(0, str(row[1] or ''))
                
                # Import client data
                if 'Dane Klienta' in wb.sheetnames:
                    ws_client = wb['Dane Klienta']
                    for row in ws_client.iter_rows(min_row=1, max_col=2, values_only=True):
                        if row[0] in self.client_entries:
                            self.client_entries[row[0]].insert(0, str(row[1] or ''))
                
                # Import invoice items
                if 'Pozycje Faktury' in wb.sheetnames:
                    ws_items = wb['Pozycje Faktury']
                    # Skip header row
                    rows = list(ws_items.iter_rows(min_row=2, values_only=True))
                    for row in rows:
                        if all(cell is None for cell in row):
                            continue
                        values = [
                            str(len(self.tree.get_children()) + 1),
                            str(row[1] or ''),
                            str(row[2] or ''),
                            str(row[3] or ''),
                            str(row[4] or ''),
                            str(row[5] or '')
                        ]
                        self.tree.insert("", "end", values=values)
                
                self.status_var.set(f"Zaimportowano dane z {filename}")
                messagebox.showinfo("Import", f"Dane zaimportowane z {filename}")
            except Exception as e:
                messagebox.showerror("Błąd", str(e))

    def import_from_csv(self):
        """Import invoice data from a CSV file"""
        filename = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if filename:
            try:
                # Clear existing entries
                self._clear_all_entries()
                
                with open(filename, 'r', encoding='utf-8') as file:
                    # Detect delimiter
                    sample = file.read(1024)
                    file.seek(0)
                    sniffer = csv.Sniffer()
                    dialect = sniffer.sniff(sample)
                    
                    reader = csv.reader(file, dialect)
                    
                    # Try to determine file structure
                    headers = next(reader)
                    
                    # Reset file pointer
                    file.seek(0)
                    reader = csv.reader(file, dialect)
                    
                    # Import based on detected structure
                    if 'Nazwa Firmy' in headers:
                        # Company/Client data format
                        next(reader)  # Skip header
                        data = list(reader)
                        
                        # Import company data
                        for label, entry in self.company_entries.items():
                            index = headers.index(label) if label in headers else -1
                            if index != -1 and data:
                                entry.insert(0, data[0][index])
                        
                        # Import client data if available
                        for label, entry in self.client_entries.items():
                            index = headers.index(label) if label in headers else -1
                            if index != -1 and data:
                                entry.insert(0, data[0][index])
                    
                    elif 'Lp.' in headers or 'Opis' in headers:
                        # Invoice items format
                        next(reader)  # Skip header
                        for row in reader:
                            if not row or all(cell == '' for cell in row):
                                continue
                            
                            values = [
                                str(len(self.tree.get_children()) + 1),
                                row[headers.index('Opis')] if 'Opis' in headers else '',
                                row[headers.index('Ilość')] if 'Ilość' in headers else '',
                                row[headers.index('Jednostka')] if 'Jednostka' in headers else '',
                                row[headers.index('Cena Netto')] if 'Cena Netto' in headers else '',
                                row[headers.index('Wartość Netto')] if 'Wartość Netto' in headers else ''
                            ]
                            self.tree.insert("", "end", values=values)
                
                self.status_var.set(f"Zaimportowano dane z {filename}")
                messagebox.showinfo("Import", f"Dane zaimportowane z {filename}")
            except Exception as e:
                messagebox.showerror("Błąd", str(e))

    def import_from_sqlite(self):
        """Import invoice data from SQLite database"""
        filename = filedialog.askopenfilename(filetypes=[("SQLite Database", "*.db")])
        if filename:
            try:
                # Clear existing entries
                self._clear_all_entries()
                
                # Connect to the selected database
                conn = sqlite3.connect(filename)
                cursor = conn.cursor()
                
                # Check for invoices table
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='invoices'")
                if not cursor.fetchone():
                    messagebox.showerror("Błąd", "Nie znaleziono tabeli faktur.")
                    return
                
                # Fetch the latest invoice
                cursor.execute("SELECT * FROM invoices ORDER BY created_at DESC LIMIT 1")
                invoice = cursor.fetchone()
                
                if invoice:
                    # Restore company data
                    company_data = json.loads(invoice[1])
                    for label, entry in self.company_entries.items():
                        entry.insert(0, company_data.get(label, ''))
                    
                    # Restore client data
                    client_data = json.loads(invoice[2])
                    for label, entry in self.client_entries.items():
                        entry.insert(0, client_data.get(label, ''))
                    
                    # Restore invoice items
                    invoice_items = json.loads(invoice[3])
                    for item in invoice_items:
                        values = [
                            str(len(self.tree.get_children()) + 1),
                            item[1], item[2], item[3], item[4], item[5]
                        ]
                        self.tree.insert("", "end", values=values)
                
                conn.close()
                
                self.status_var.set(f"Zaimportowano dane z {filename}")
                messagebox.showinfo("Import", f"Dane zaimportowane z {filename}")
            except Exception as e:
                messagebox.showerror("Błąd", str(e))

    def add_invoice_item(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("Dodaj Pozycję")
        add_window.geometry("400x300")

        labels = ["Opis", "Ilość", "Jednostka", "Cena Netto"]
        entries = {}

        for i, label in enumerate(labels):
            tk.Label(add_window, text=label, font=('Arial', 10)).pack()
            entry = tk.Entry(add_window, width=50, font=('Arial', 10))
            entry.pack()
            entries[label] = entry

        def save_item():
            values = [str(len(self.tree.get_children())+1)]  # Lp.
            values.extend([entries[label].get() for label in labels])
            
            # Calculate net value
            try:
                quantity = float(entries["Ilość"].get())
                price = float(entries["Cena Netto"].get())
                net_value = quantity * price
                values.append(f"{net_value:.2f}")
            except ValueError:
                net_value = 0
                values.append("0.00")
                messagebox.showwarning("Błąd", "Nieprawidłowe wartości ilości lub ceny.")
                return

            self.tree.insert("", "end", values=values)
            self.status_var.set(f"Dodano pozycję: {entries['Opis'].get()}")
            add_window.destroy()

        tk.Button(add_window, text="Zapisz", command=save_item, font=('Arial', 10)).pack(pady=10)

    def remove_invoice_item(self):
        selected_item = self.tree.selection()
        if selected_item:
            item_details = self.tree.item(selected_item)['values']
            self.tree.delete(selected_item)
            self.status_var.set(f"Usunięto pozycję: {item_details[1]}")
        else:
            messagebox.showwarning("Błąd", "Wybierz pozycję do usunięcia.")

    def clear_invoice_items(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.status_var.set("Usunięto wszystkie pozycje faktury.")

def main():
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()

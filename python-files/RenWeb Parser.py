import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import re

# Configuration
CURRENT_YEAR = 2025
MAX_GRADE = 12
DEFAULT_FONT_NAME = "Calibri"
COMMON_SUFFIXES_CAPTURING_GROUP = r'(Jr\.|Sr\.|II|III|IV|V)'

# Load HTML
html_file_path = "directory.cfm"
try:
    with open(html_file_path, "r", encoding="utf-8") as f:
        html_content = f.read()
except FileNotFoundError:
    print(f"Error: HTML file '{html_file_path}' not found. Please check the filename.")
    exit()

soup = BeautifulSoup(html_content, "html.parser")

students_data = []
parents_data_raw = []

# --- Helper function to parse parent names from student row ---
def parse_student_parent_listing(parent_string):
    raw_parents_str = parent_string.strip()
    if not raw_parents_str: return []
    parsed_names_final = []
    individual_parent_chunks = [p.strip() for p in re.split(r'\s+and\s+', raw_parents_str, flags=re.IGNORECASE)]
    shared_last_name_from_first_chunk = None
    for i, chunk_str in enumerate(individual_parent_chunks):
        if not chunk_str: continue
        match_pattern1 = re.match(fr'^(.*?),\s*{COMMON_SUFFIXES_CAPTURING_GROUP},\s*(.+)$', chunk_str, re.IGNORECASE)
        if match_pattern1:
            last, suffix, first = match_pattern1.groups()
            parsed_names_final.append(f"{last.strip()} {suffix.strip()}, {first.strip()}")
            if i == 0: shared_last_name_from_first_chunk = last.strip()
            continue
        match_pattern2 = re.match(fr'^((?:.*?)\s+{COMMON_SUFFIXES_CAPTURING_GROUP}),\s*(.+)$', chunk_str, re.IGNORECASE)
        if match_pattern2:
            last_with_suffix, actual_suffix_part_for_base_name_logic, first = match_pattern2.groups()
            parsed_names_final.append(f"{last_with_suffix.strip()}, {first.strip()}")
            if i == 0: 
                base_last_name = re.sub(r'\s+' + re.escape(actual_suffix_part_for_base_name_logic) + r'$', '', last_with_suffix.strip(), flags=re.IGNORECASE)
                shared_last_name_from_first_chunk = base_last_name.strip()
            continue
        match_pattern3 = re.match(r'^(.*?),\s*(.+)$', chunk_str)
        if match_pattern3:
            last, first = match_pattern3.groups()
            parsed_names_final.append(f"{last.strip()}, {first.strip()}")
            if i == 0: shared_last_name_from_first_chunk = last.strip()
            continue
        if shared_last_name_from_first_chunk and i > 0:
            match_pattern4a = re.match(fr'^{COMMON_SUFFIXES_CAPTURING_GROUP},\s*(.+)$', chunk_str, re.IGNORECASE)
            if match_pattern4a:
                suffix, first = match_pattern4a.groups()
                parsed_names_final.append(f"{shared_last_name_from_first_chunk} {suffix.strip()}, {first.strip()}")
                continue
            match_pattern4b = re.match(fr'^{COMMON_SUFFIXES_CAPTURING_GROUP}\s+(.+)$', chunk_str, re.IGNORECASE)
            if match_pattern4b:
                suffix, first = match_pattern4b.groups()
                if not re.fullmatch(COMMON_SUFFIXES_CAPTURING_GROUP, first.strip(), re.IGNORECASE):
                    parsed_names_final.append(f"{shared_last_name_from_first_chunk} {suffix.strip()}, {first.strip()}")
                    continue
            if ' ' not in chunk_str and not re.fullmatch(COMMON_SUFFIXES_CAPTURING_GROUP, chunk_str.strip(), re.IGNORECASE):
                parsed_names_final.append(f"{shared_last_name_from_first_chunk}, {chunk_str.strip()}")
                continue
        parsed_names_final.append(chunk_str)
        if i == 0 and ' ' in chunk_str and ',' not in chunk_str:
            name_bits = chunk_str.split(' ')
            if len(name_bits) > 1: shared_last_name_from_first_chunk = name_bits[-1]
    return [name for name in parsed_names_final if name.strip()]

# --- STUDENT SECTION ---
student_table = soup.find("table", id="results1-list")
if student_table:
    student_rows_container = student_table.find("tbody") if student_table.find("tbody") else student_table
    student_rows = student_rows_container.find_all("tr")
    for row_idx, row in enumerate(student_rows):
        if row.find("th"): continue
        student_col = row.find("td", class_="col_student")
        if not student_col: continue
        name_tag = student_col.find("a")
        student_first_name, student_last_name, student_full_name_display = "", "", ""
        student_person_id = f"s_row_{row_idx}" # Fallback ID
        if name_tag:
            student_full_name_text = name_tag.text.strip()
            student_full_name_display = student_full_name_text # e.g., "Casper, Jacob"
            if 'personid=' in name_tag['href']:
                 student_person_id = name_tag['href'].split('personid=')[1].split('&')[0]

            if "," in student_full_name_text:
                parts = [x.strip() for x in student_full_name_text.split(",", 1)]
                student_last_name = parts[0]
                first_part = parts[1]
                student_first_name = first_part.split('/')[0].strip() if '/' in first_part else first_part
            else:
                name_parts = student_full_name_text.split()
                if len(name_parts) > 1: student_first_name, student_last_name = " ".join(name_parts[:-1]), name_parts[-1]
                else: student_first_name = student_full_name_text
        grade_col = row.find("td", class_="col_grade")
        grade_text = grade_col.text.strip() if grade_col else ""
        grade_cleaned_for_display, grade_for_calculation, grad_year, student_email = grade_text, None, "", ""
        try:
            numeric_part_match = re.match(r"(\d+)", grade_text)
            if numeric_part_match:
                grade_val = int(numeric_part_match.group(1))
                if 0 < grade_val <= MAX_GRADE:
                    grade_for_calculation, grad_year = grade_val, CURRENT_YEAR + (MAX_GRADE - grade_val)
        except ValueError: pass
        if student_first_name and student_last_name and grade_for_calculation and grad_year:
            clean_first, clean_last = re.sub(r"[^a-z0-9]", "", student_first_name.lower()), re.sub(r"[^a-z0-9]", "", student_last_name.lower())
            if clean_first and clean_last: student_email = f"{clean_first}.{clean_last}{str(grad_year)[-2:]}@scaschools.org"
        parents_names_col = row.find("td", class_="col_parents")
        raw_parent_listing_from_student_row = parents_names_col.get_text(strip=True) if parents_names_col else ""
        parsed_parents_for_student = parse_student_parent_listing(raw_parent_listing_from_student_row)
        phone_col = row.find("td", class_="col_phone")
        student_phone_text = phone_col.get_text(" ", strip=True).replace("H: ", "") if phone_col else ""
        address_col, address_text = row.find("td", class_="col_address"), "(not listed)"
        if address_col:
            address_link, raw_address_text = address_col.find('a'), address_col.get_text(strip=True)
            if address_link and address_link.get_text(strip=True) and address_link.get_text(strip=True) not in [", ,", ""]: address_text = address_link.get_text(strip=True)
            elif raw_address_text and raw_address_text not in [", ,", "", "(not listed)"]: address_text = raw_address_text
        zip_col, zipcode_text = row.find("td", class_="col_zipcode"), ""
        if zip_col:
            zip_link = zip_col.find('a')
            if zip_link and zip_link.get_text(strip=True): zipcode_text = zip_link.get_text(strip=True)
            elif zip_col.get_text(strip=True): zipcode_text = zip_col.get_text(strip=True)
        
        students_data.append({
            "Student Person ID": student_person_id, # Key for linking
            "Student Full Name Display": student_full_name_display, # For display/lookup
            "First Name": student_first_name, "Last Name": student_last_name, "Grade": grade_cleaned_for_display,
            "Graduation Year": grad_year, "Student Email": student_email, "Raw Parent Listing": raw_parent_listing_from_student_row,
            "Parsed Parents from Student Row": parsed_parents_for_student, "Phone": student_phone_text,
            "Address": address_text, "Zip Code": zipcode_text
        })

# --- PARENT SECTION (Initial Parse) ---
parent_table = soup.find("table", id="results2-list")
if parent_table:
    parent_rows_container = parent_table.find("tbody") if parent_table.find("tbody") else parent_table
    parent_rows = parent_rows_container.find_all("tr")
    for row_idx, row in enumerate(parent_rows):
        if row.find("th"): continue
        parent_name_col, email_col = row.find("td", class_="col_parents"), row.find("td", class_="col_email")
        if not (parent_name_col and email_col): continue
        name_tag = parent_name_col.find("a")
        parent_full_name_from_parent_table = name_tag.text.strip() if name_tag else ""
        parent_person_id = f"p_row_{row_idx}" # Fallback ID
        if name_tag and 'personid=' in name_tag['href']:
            parent_person_id = name_tag['href'].split('personid=')[1].split('&')[0]
        parent_first_name, parent_last_name = "", ""
        if parent_full_name_from_parent_table: 
            if "," in parent_full_name_from_parent_table:
                parts = [p.strip() for p in parent_full_name_from_parent_table.split(",",1)]
                parent_last_name = parts[0] 
                if len(parts) > 1: parent_first_name = parts[1]
            else:
                name_parts = parent_full_name_from_parent_table.split()
                if len(name_parts) > 1: parent_first_name, parent_last_name = " ".join(name_parts[:-1]), name_parts[-1]
                else: parent_last_name = parent_full_name_from_parent_table
        email_tag, parent_email_text = email_col.find("a"), ""
        if email_tag: parent_email_text = email_tag.text.strip()
        phone_col, home_phone, cell_phone = row.find("td", class_="col_phone"), "", ""
        if phone_col:
            current_line_elements_texts = []
            for content_node in phone_col.contents:
                if hasattr(content_node, 'name') and content_node.name == 'br':
                    line_str = "".join(current_line_elements_texts).strip()
                    if line_str:
                        if line_str.startswith("H:"): home_phone = line_str.replace("H:", "", 1).strip()
                        elif line_str.startswith("C:"): cell_phone = line_str.replace("C:", "", 1).strip()
                    current_line_elements_texts = [] 
                else:
                    node_text = content_node.get_text(strip=True)
                    if node_text: current_line_elements_texts.append(node_text)
            if current_line_elements_texts:
                line_str = "".join(current_line_elements_texts).strip()
                if line_str:
                    if line_str.startswith("H:") and not home_phone: home_phone = line_str.replace("H:", "", 1).strip()
                    elif line_str.startswith("C:") and not cell_phone: cell_phone = line_str.replace("C:", "", 1).strip()
                    elif not home_phone and not cell_phone: home_phone = line_str
        address_col, parent_address_text = row.find("td", class_="col_address"), "(not listed)"
        if address_col:
            address_link, raw_address_text = address_col.find('a'), address_col.get_text(strip=True)
            if address_link and address_link.get_text(strip=True) and address_link.get_text(strip=True) not in [", ,", ""]: parent_address_text = address_link.get_text(strip=True)
            elif raw_address_text and raw_address_text not in [", ,", "", "(not listed)"]: parent_address_text = raw_address_text
        zip_col, parent_zipcode_text = row.find("td", class_="col_zipcode"), ""
        if zip_col:
            zip_link = zip_col.find('a')
            if zip_link and zip_link.get_text(strip=True): parent_zipcode_text = zip_link.get_text(strip=True)
            elif zip_col.get_text(strip=True): parent_zipcode_text = zip_col.get_text(strip=True)
        parents_data_raw.append({
            "Parent Person ID": parent_person_id, 
            "Parent Full Name (from Parent Table)": parent_full_name_from_parent_table,
            "Parent First Name": parent_first_name, "Parent Last Name": parent_last_name, "Email": parent_email_text,
            "Home Phone": home_phone, "Cell Phone": cell_phone, "Address": parent_address_text, "Zip Code": parent_zipcode_text
        })

# --- LINK PARENTS TO STUDENTS ---
final_parents_list_with_students = []
processed_parent_ids_for_final_list = set()
for parent_entry_from_table in parents_data_raw:
    parent_id = parent_entry_from_table["Parent Person ID"]
    if parent_id in processed_parent_ids_for_final_list: continue 
    current_parent_associated_student_ids = [] # Store student Person IDs
    name_to_match_from_parent_table = parent_entry_from_table["Parent Full Name (from Parent Table)"]
    for student in students_data:
        if name_to_match_from_parent_table in student["Parsed Parents from Student Row"]:
            if student["Student Person ID"] not in current_parent_associated_student_ids:
                current_parent_associated_student_ids.append(student["Student Person ID"])
    new_parent_entry = parent_entry_from_table.copy()
    new_parent_entry["Associated Student IDs"] = current_parent_associated_student_ids # Store IDs
    final_parents_list_with_students.append(new_parent_entry)
    processed_parent_ids_for_final_list.add(parent_id)

# Create a student lookup by Person ID for easy access to display names and grades
student_lookup_by_id = {s["Student Person ID"]: s for s in students_data}

# Create DataFrames for output
students_df_for_output = pd.DataFrame(students_data)[[
    "First Name", "Last Name", "Grade", "Graduation Year", 
    "Student Email", "Raw Parent Listing", "Phone", "Address", "Zip Code"
]].copy()

parents_df_for_output = pd.DataFrame(final_parents_list_with_students)
if not parents_df_for_output.empty:
    # Convert list of student IDs to a comma-separated string of student display names
    def get_student_display_names(student_ids_list):
        names = []
        for s_id in student_ids_list:
            student_info = student_lookup_by_id.get(s_id)
            if student_info:
                names.append(student_info["Student Full Name Display"]) # Use "Last, First" format here
        return ", ".join(sorted(list(set(names))))

    parents_df_for_output["Associated Student(s)"] = parents_df_for_output["Associated Student IDs"].apply(get_student_display_names)
    
    columns_to_keep_parents = ["Parent First Name", "Parent Last Name", "Associated Student(s)", "Email", 
                               "Home Phone", "Cell Phone", "Address", "Zip Code"]
    parents_df_for_output_final = parents_df_for_output[columns_to_keep_parents].drop_duplicates(
        subset=["Parent First Name", "Parent Last Name", "Email", "Address"], keep='first' # Key for uniqueness
    ).reset_index(drop=True)
else: 
    parents_df_for_output_final = pd.DataFrame(columns=["Parent First Name", "Parent Last Name", "Associated Student(s)", "Email", 
                                             "Home Phone", "Cell Phone", "Address", "Zip Code"])

# --- STUDENT EMAILS BY GRADE SHEET ---
student_email_by_grade_data = []
for student in students_data:
    student_email = student.get("Student Email")
    student_grade = student.get("Grade")
    if student_email and pd.notna(student_email) and str(student_email).strip() and student_grade:
        student_email_by_grade_data.append({
            "Student Grade": student_grade,
            "Student Email": student_email
        })
if student_email_by_grade_data:
    student_emails_by_grade_df = pd.DataFrame(student_email_by_grade_data).drop_duplicates().sort_values(
        by=["Student Grade", "Student Email"]
    ).reset_index(drop=True)
else:
    student_emails_by_grade_df = pd.DataFrame(columns=["Student Grade", "Student Email"])


# --- PARENT EMAILS BY GRADE SHEET ---
parent_email_by_grade_data = []
for parent_row in final_parents_list_with_students: # Use the list that still has student IDs
    parent_email = parent_row.get("Email")
    if parent_email and pd.notna(parent_email) and str(parent_email).strip():
        for student_id in parent_row.get("Associated Student IDs", []):
            student_info = student_lookup_by_id.get(student_id)
            if student_info and student_info.get("Grade"):
                parent_email_by_grade_data.append({
                    "Student Grade": student_info["Grade"],
                    "Parent Email": parent_email
                })
if parent_email_by_grade_data:
    parent_emails_by_grade_df = pd.DataFrame(parent_email_by_grade_data).drop_duplicates().sort_values(
        by=["Student Grade", "Parent Email"]
    ).reset_index(drop=True)
else:
    parent_emails_by_grade_df = pd.DataFrame(columns=["Student Grade", "Parent Email"])


# --- EXPORT TO EXCEL ---
excel_file_name = "sca_directory_enhanced_v10.xlsx" 
with pd.ExcelWriter(excel_file_name, engine="openpyxl") as writer:
    students_df_for_output.to_excel(writer, index=False, sheet_name="Students")
    parents_df_for_output_final.to_excel(writer, index=False, sheet_name="Parents")
    student_emails_by_grade_df.to_excel(writer, index=False, sheet_name="Student Emails by Grade") 
    parent_emails_by_grade_df.to_excel(writer, index=False, sheet_name="Parent Emails by Grade")

    workbook = writer.book
    default_font = Font(name=DEFAULT_FONT_NAME, size=11)
    center_align = Alignment(horizontal='center', vertical='center') 
    
    sheet_dfs = {
        "Students": students_df_for_output, 
        "Parents": parents_df_for_output_final,
        "Student Emails by Grade": student_emails_by_grade_df, 
        "Parent Emails by Grade": parent_emails_by_grade_df
    }

    for sheet_name, df_sheet in sheet_dfs.items():
        worksheet = writer.sheets[sheet_name]
        if not df_sheet.empty: 
            for row_num, row_cells in enumerate(worksheet.iter_rows(), 1):
                for cell in row_cells:
                    cell.font = default_font
                    # Center Grade column DATA (not header)
                    grade_column_name = ""
                    if sheet_name == "Students": grade_column_name = "Grade"
                    elif sheet_name == "Student Emails by Grade": grade_column_name = "Student Grade"
                    elif sheet_name == "Parent Emails by Grade": grade_column_name = "Student Grade"
                    
                    if grade_column_name and row_num > 1:
                        current_col_header = worksheet.cell(row=1, column=cell.column).value
                        if current_col_header == grade_column_name:
                            cell.alignment = center_align
            for i, col_name in enumerate(df_sheet.columns):
                column_letter = get_column_letter(i + 1)
                max_len_data = df_sheet[col_name].astype(str).map(len).max()
                max_len_data = 0 if pd.isna(max_len_data) else max_len_data 
                column_width = max(max_len_data, len(str(col_name))) + 3 # Increased padding slightly
                worksheet.column_dimensions[column_letter].width = column_width
            worksheet.auto_filter.ref = worksheet.dimensions
        elif sheet_name in worksheet.parent.sheetnames: 
            if worksheet.max_row > 0: # Check if headers exist
                header_row = worksheet[1]
                for cell in header_row: cell.font = default_font
            # Define default column names for empty sheets for width calculation
            default_cols = []
            if sheet_name == "Student Emails by Grade": default_cols = ["Student Grade", "Student Email"]
            elif sheet_name == "Parent Emails by Grade": default_cols = ["Student Grade", "Parent Email"]
            elif not df_sheet.columns.empty: default_cols = df_sheet.columns # Fallback to df.columns if not one of the above
            
            for i, col_name_str in enumerate(default_cols):
                column_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[column_letter].width = len(str(col_name_str)) + 3


print(f"✅ Export complete: {excel_file_name}")

import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def clean_filename(name):
    return "".join(c for c in str(name) if c not in r'\/:*?"<>|')

def rename_files():
    folder_path = filedialog.askdirectory(title="选择包含 Excel 文件的文件夹")
    if not folder_path:
        return

    renamed_count = 0
    skipped_count = 0

    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            file_path = os.path.join(folder_path, filename)
            try:
                wb = load_workbook(file_path, data_only=True)
                sheetnames = wb.sheetnames
                if len(sheetnames) < 2:
                    skipped_count += 1
                    continue

                sheet2 = wb[sheetnames[1]]
                b7_value = sheet2['B7'].value

                if not b7_value:
                    skipped_count += 1
                    continue

                new_name = clean_filename(str(b7_value))
                new_filename = f"{new_name}.xlsx"
                new_file_path = os.path.join(folder_path, new_filename)

                count = 1
                while os.path.exists(new_file_path):
                    new_filename = f"{new_name}_{count}.xlsx"
                    new_file_path = os.path.join(folder_path, new_filename)
                    count += 1

                os.rename(file_path, new_file_path)
                renamed_count += 1

            except Exception as e:
                print(f"错误处理文件 {filename}：{e}")
                skipped_count += 1

    messagebox.showinfo("处理完成", f"成功重命名：{renamed_count} 个文件\n跳过：{skipped_count} 个文件")

root = tk.Tk()
root.title("Excel 批量重命名工具")
root.geometry("400x200")

label = tk.Label(root, text="点击按钮选择文件夹并批量重命名 Excel 文件（按 Sheet2 的 B7）", wraplength=380, justify="center")
label.pack(pady=30)

btn = tk.Button(root, text="选择文件夹并开始", command=rename_files, bg="green", fg="white", height=2, width=25)
btn.pack()

root.mainloop()

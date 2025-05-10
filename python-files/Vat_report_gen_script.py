import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import warnings
from pandas.errors import SettingWithCopyWarning

# Suppress UserWarnings from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')
warnings.filterwarnings("ignore", category=SettingWithCopyWarning)

# Function to get the theme from the Global Status sheet
def get_theme_from_global_status(file_path):
    global_status_df = pd.read_excel(file_path, sheet_name='Global Status', engine='openpyxl', header=None)
    return global_status_df.at[1, 3]  # Extract Theme from the 2nd row (index 1) and 4th column (index 3)

def process_excel_file(file_path):  
    # Read the Analysis sheet
    df = pd.read_excel(file_path, sheet_name='Analysis', engine='openpyxl', header=0)
    # Find the first empty row
    for index, row in df.iterrows():
        if row.isnull().all():  # Check if all elements in the row are NaN
            df = df.iloc[:index]  # Keep rows up to the first empty row
            break
            
    header = []

    for col in range(df.shape[1]):  # Loop over columns
        # Check if the cell in the second row is NaN
        if pd.isna(df.iat[1, col]):
            merged_value = df.iat[2, col]  # If NaN in the second row, take from the third row
        else:
            merged_value = df.iat[1, col]  # Otherwise, take from the second row
        header.append(merged_value)  # Append the header value

    # Set the new header to the DataFrame
    df.columns = header

    # Drop the first two rows which contained the header information
    df = df.drop(index=[0, 1]).reset_index(drop=True) 
    # Add the Theme to the Analysis DataFrame
    theme = get_theme_from_global_status(file_path)
    # Suppress the FutureWarning
    with warnings.catch_warnings():
        warnings.simplefilter(action='ignore', category=FutureWarning)
        # Replace NaN values with the string 'N/A'
        df.fillna('N/A', inplace=True)
    # Initialize additional columns if they don't exist
    if 'Theme' not in df.columns:
        df['Theme'] = theme  # Assign the Theme if not already present

    # Initialize additional columns if they don't exist
    if 'Implem verif CR' not in df.columns:
        df['Implem verif CR'] = '' 
    if 'Justification' not in df.columns:
        df['Justification'] = ''  
    if 'Base_Line_worked' not in df.columns:
        df['Base_Line_worked'] = ''  

    return df
    


# Directory containing the Excel files
data_folder = r"C:\2.2\VAT_REPORT_SCRIPT\Before_review" 

# Initialize an empty list to store DataFrames
all_data = []
ctc_data = []
wstc_data = []
wsic_data = []
# Loop through all Excel files in the directory
for filename in os.listdir(data_folder):
    if filename.endswith('.xlsm') or filename.endswith('.xlsx'):  # Check for Excel files
        file_path = os.path.join(data_folder, filename)
        
        try:
            # Process each file and get the DataFrame
            processed_df = process_excel_file(file_path)

            # Check if the required columns exist
            required_columns = ['FS TIS ID', 'Req status (FS/AD)', 'Req status (EMIL code)', 'SIL']
            if all(col in processed_df.columns for col in required_columns):
                if 'Applicable Equipment' in processed_df.columns:
                    # Append data to the respective list based on Applicable Equipment
                    ctc_subset = processed_df[processed_df['Applicable Equipment'].str.contains(r'CTC')]
                    wstc_subset = processed_df[processed_df['Applicable Equipment'].str.contains(r'WSTC')]
                    wsic_subset = processed_df[processed_df['Applicable Equipment'].str.contains(r'WSIC')]

                # Append only if there are any entries
                if not ctc_subset.empty:
                    ctc_data.append(ctc_subset)
                if not wstc_subset.empty:
                    wstc_data.append(wstc_subset)
                if not wsic_subset.empty:
                    wsic_data.append(wsic_subset)
                # Append the DataFrame to the list
                #all_data.append(processed_df)
                print(f"processing file {filename}")

        except Exception as e:
            print(f"Error processing file {filename}: {e}")
            
# Define output file paths
output_ctc_path = r"C:\2.2\VAT_REPORT_SCRIPT\CTC_VAT_REPORT.xlsx"
output_wstc_path = r"C:\2.2\VAT_REPORT_SCRIPT\WSTC_VAT_REPORT.xlsx"
output_wsic_path = r"C:\2.2\VAT_REPORT_SCRIPT\WSIC_VAT_REPORT.xlsx"
# Concatenate all DataFrames into one after processing all files
def save_to_excel(data_list, output_file_path):
    """Function to save a list of DataFrames to a single Excel file."""
    if data_list:
        combined_df = pd.concat(data_list, ignore_index=True)

        # Function to determine the priority of statuses
        def prioritize_fs_ad(status_series):
            if 'NOK' in status_series.values:
                return 'NOK'
            if 'OK' in status_series.values and r'N/A' in status_series.values:
                return 'OK'
            if 'NOK' in status_series.values and r'N/A' in status_series.values:
                return 'NOK'
            if r'N/A' in status_series.values and 'OK' in status_series.values:
                return 'OK'
            if r'N/A' in status_series.values and 'NOK' in status_series.values:
                return 'NOK'
            if 'OK' in status_series.values and 'NOK' in status_series.values:
                return 'NOK'
            if 'NOK' in status_series.values and 'NOK' in status_series.values:
                return 'NOK'
            if 'NOK' in status_series.values and 'OK' in status_series.values:
                return 'NOK'
            if 'OK' in status_series.values and 'OK' in status_series.values:
                return 'OK'
            return status_series.iloc[0]  # Return the first status as fallback

        # Group by 'FS TIS ID' and aggregate statuses
        grouped = combined_df.groupby('FS TIS ID').agg({
            'Req status (FS/AD)': prioritize_fs_ad,
            'Req status (EMIL code)': prioritize_fs_ad, 
            'Theme': 'first',                    # Get the first occurrence for Theme as well
            'SIL': 'first',                      # Get the first occurrence for SIL Level
            'Implem verif CR': 'first',         
            'Justification': 'first',
            'Base_Line_worked':'first'
        }).reset_index()

        # Prepare the final DataFrame to save
        result_df = grouped[['Theme', 'FS TIS ID', 'SIL', 'Req status (FS/AD)', 'Req status (EMIL code)', 'Implem verif CR', 'Justification', 'Base_Line_worked']]
        result_df.rename(columns={'SIL': 'SIL Level'}, inplace=True)

        # **Sort the DataFrame by 'Theme' in alphabetical order**
        result_df = result_df.sort_values(by='Theme').reset_index(drop=True)

        # Create the output workbook and write the data
        wb = Workbook()
        ws = wb.active
        # Write headers with black background and white text
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black fill
        header_font = Font(color="FFFFFF", bold=True)
        # Write headers
    
        for col_idx, column in enumerate(result_df.columns):
            cell = ws.cell(row=1, column=col_idx + 1, value=column)
            cell.fill = header_fill
            cell.font = header_font  # Set header font color
            # Apply data, color format, and add borders to the cells
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        # Apply data and color format
        for row_idx, row in result_df.iterrows():
            for col_idx, value in enumerate(row):
                cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

                # Apply color formatting based on status values
                if col_idx == result_df.columns.get_loc('Req status (FS/AD)'):
                    if value == 'OK':
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                    elif str(value).strip() == 'NOK':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                    elif str(value).strip() == 'N/A':
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                if col_idx == result_df.columns.get_loc('Req status (EMIL code)'):
                    if str(value).strip() == 'OK':
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                    elif str(value).strip() == 'NOK':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                    elif str(value).strip() == 'N/A':
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                # Apply border to each cell
                cell.border = thin_border
            
    
        # Set column widths for better presentation
        for col in range(1, len(result_df.columns) + 1):  # Loop over the columns
            max_length = 0
            column = get_column_letter(col)  # Convert to Excel column letter (1 -> 'A')
            for cell in ws[column]:  # Check all cells in the column
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))  # Get max length in the column
                except:
                    pass
            adjusted_width = (max_length + 2)  # Adjust for some padding
            ws.column_dimensions[column].width = adjusted_width  # Set the column width
        
        # Save the styled workbook
        wb.save(output_file_path)
        print(f"Combined data has been saved to {output_file_path}")

    else: 
        print("No data was extracted from any files.")
# Save the filtered data to their respective files
save_to_excel(ctc_data, output_ctc_path)
save_to_excel(wstc_data, output_wstc_path)
save_to_excel(wsic_data, output_wsic_path)


#----------------comparing requirements in change log and requirements in CTC vat Repot-----------#
ctc_data = r"C:\2.2\VAT_REPORT_SCRIPT\CTC_VAT_REPORT.xlsx"
wstc_data = r"C:\2.2\VAT_REPORT_SCRIPT\WSTC_VAT_REPORT.xlsx"
wsic_data = r"C:\2.2\VAT_REPORT_SCRIPT\WSIC_VAT_REPORT.xlsx"
# Read the Excel file with multi-level header
file_path = r"C:\2.2\VAT_REPORT_SCRIPT\Changelog\CTC_Change_Log_E_F.xlsm"
data = pd.read_excel(file_path, header=[0, 1, 2, 3, 4], sheet_name='Result', engine='openpyxl')
ctc_original_df = pd.read_excel(ctc_data)

#print("Original Columns in DataFrame:")
#print(data.columns.tolist())  # List all columns for inspection
new_columns = []
tis_id_count = 0

for col in data.columns:
    if 'TIS ID' in col:
        tis_id_count += 1
        new_columns.append(f"TIS ID_{tis_id_count}")  # Make TIS ID unique
    elif 'Theme' in col:
        new_columns.append("Theme")  # Assign a unique name (optional if you use one name for all)
    elif 'SIL' in col:
        new_columns.append("SIL")  # Assign a unique name (optional if you use one name for all)
    elif 'Implementation Language' in col:
        new_columns.append("Implementation Language")
    elif 'Status' in col:
        new_columns.append("Status")        
    else:
        new_columns.append(col)  # Keep other columns unchanged
# Assign the new unique column names
data.columns = new_columns  # Assign the new unique column names

# Define the column for Implementation Process
implementation_process_column = 'Implementation Language'
status = 'Status'
# Filter out rows where Implementation Process is either 'Atomic # Ada SIL4' or 'Atomic # Ada SIL0'
if implementation_process_column in data.columns:
    filtered_data = data[
        (data[implementation_process_column] != 'Atomic # Ada SIL4') & 
        (data[implementation_process_column] != 'Atomic # Ada SIL0') &
        (data[implementation_process_column] != 'C++') &
        (data[status] != 'D')
        
    ]
else:
    print(f"Warning: Column '{implementation_process_column}' not found in the DataFrame.")

# Ensure the relevant columns exist after filtering
fs_tis_id_column_name = 'TIS ID_1'  # Adjust based on your actual DataFrame
theme_column = 'Theme'   # Adjust if necessary
sil_column = 'SIL'       # Adjust if necessary

# Rename the first TIS ID column to 'FIS TIS ID' if it exists
if (fs_tis_id_column_name in filtered_data.columns) and (theme_column in filtered_data.columns) and (sil_column in filtered_data.columns):
    filtered_data.rename(columns={fs_tis_id_column_name: 'FIS TIS ID'}, inplace=True)

    # Create a new DataFrame with 'Theme', 'FIS TIS ID', and 'SIL'
    output_data_ctc = filtered_data[[theme_column, 'FIS TIS ID', sil_column]].copy()

    # Remove duplicates based on the first TIS ID column
    output_data_ctc = output_data_ctc.drop_duplicates(subset='FIS TIS ID')
    
    # Print the resulting DataFrame
    #print("Filtered Data with the First FIS TIS ID:")
    #print(output_data_ctc)

    # Optionally save the output DataFrame to an Excel file
    output_file_path = r"C:\2.2\VAT_REPORT_SCRIPT\Filtered_change_log\filtered_changelog_ctc.xlsx"
    output_data_ctc.to_excel(output_file_path, index=False, sheet_name='Filtered Data')
    print(f"Filtered data has been saved to {output_file_path}.")
else:
    print(f"Warning: Required columns were not found in the filtered DataFrame: {fs_tis_id_column_name}, {theme_column}, {sil_column}.")

output_data_ctc_path = r"C:\2.2\VAT_REPORT_SCRIPT\Filtered_change_log\filtered_changelog_ctc.xlsx"
output_data_ctc1 = pd.read_excel(output_data_ctc_path, engine='openpyxl')
# Ensure the relevant columns exist
if 'FIS TIS ID' in output_data_ctc1.columns and 'FS TIS ID' in ctc_original_df.columns:
    
    # Create sets for comparison
    fis_tis_ids = set(output_data_ctc1['FIS TIS ID'])
    fs_tis_ids = set(ctc_original_df['FS TIS ID'])

    
    missing_in_fs = fis_tis_ids - fs_tis_ids
    #missing_in_fis = fs_tis_ids - fis_tis_ids

    # Create a DataFrame for mismatches including Theme and SIL columns
    mismatch_rows = []  # List for capturing mismatch entries

    # Gather details for mismatches
    for missing_id in missing_in_fs:
        # Get the corresponding Theme and SIL for the missing FIS TIS ID
        row = output_data_ctc[output_data_ctc['FIS TIS ID'] == missing_id]
        if not row.empty:
            theme_value = row['Theme'].values[0]  # Adjust based on your actual DataFrame structure
            sil_value = row['SIL'].values[0]
            mismatch_rows.append({
                'Theme': theme_value,
                'Missing FIS TIS ID': missing_id,
                'SIL': sil_value
            })

    # Create DataFrame from mismatched rows
    mismatch_df = pd.DataFrame(mismatch_rows)
    mismatch_df = mismatch_df.sort_values(by=['Theme'], ascending= True)

    # Save mismatches to a new sheet in the original CTC file
    if not mismatch_df.empty:
        with pd.ExcelWriter(ctc_data, engine='openpyxl', mode='a') as writer:  # Append to existing file
            mismatch_df.to_excel(writer, index=False, sheet_name='Mismatch_FRL_final_ctc')
        #print(f"Mismatches have been written to {ctc_data} in the 'Mismatch_FRL_final_ctc' sheet.")
    else:
        print("No mismatches found between 'FIS TIS ID' and 'FS TIS ID'.")
else:
    print("Warning: Required columns not found in the respective DataFrames.")
#-------------------------------------------------------------------------------------------------------#

#----------------comparing requirements in change log and requirements in WSTC vat Repot-----------#
file_path_wstc = r"C:\2.2\VAT_REPORT_SCRIPT\Changelog\WSTC_Change_Log_E_F 2.xlsm"
data_wstc = pd.read_excel(file_path_wstc, header=[0, 1, 2, 3, 4], sheet_name='Result', engine='openpyxl')
wstc_original_df = pd.read_excel(wstc_data)

#print("Original Columns in DataFrame:")
#print(data_wstc.columns.tolist()) 
new_columns = []
tis_id_count = 0

for col in data_wstc.columns:
    if 'TIS ID' in col:
        tis_id_count += 1
        new_columns.append(f"TIS ID_{tis_id_count}")  # Make TIS ID unique
    elif 'Theme' in col:
        new_columns.append("Theme")  # Assign a unique name (optional if you use one name for all)
    elif 'SIL' in col:
        new_columns.append("SIL")  # Assign a unique name (optional if you use one name for all)
    elif 'Implementation Language' in col:
        new_columns.append("Implementation Language")
    elif 'Status' in col:
        new_columns.append("Status")        
    else:
        new_columns.append(col)  # Keep other columns unchanged
# Assign the new unique column names
data_wstc.columns = new_columns  # Assign the new unique column names

# Define the column for Implementation Process
implementation_process_column = 'Implementation Language'
status = 'Status'
# Filter out rows where Implementation Process is either 'Atomic # Ada SIL4' or 'Atomic # Ada SIL0'
if implementation_process_column in data_wstc.columns:
    filtered_data_wstc = data_wstc[
        (data_wstc[implementation_process_column] != 'Atomic # Ada SIL4') & 
        (data_wstc[implementation_process_column] != 'Atomic # Ada SIL0') &
        (data_wstc[implementation_process_column] != 'C++') &
        (data_wstc[status] != 'D')
        
    ]
else:
    print(f"Warning: Column '{implementation_process_column}' not found in the DataFrame.")

# Ensure the relevant columns exist after filtering
fs_tis_id_column_name = 'TIS ID_1'  # Adjust based on your actual DataFrame
theme_column = 'Theme'   
sil_column = 'SIL'       # Adjust if necessary

# Rename the first TIS ID column to 'FIS TIS ID' if it exists
if (fs_tis_id_column_name in filtered_data_wstc.columns) and (theme_column in filtered_data_wstc.columns) and (sil_column in filtered_data_wstc.columns):
    filtered_data_wstc.rename(columns={fs_tis_id_column_name: 'FIS TIS ID'}, inplace=True)

    # Create a new DataFrame with 'Theme', 'FIS TIS ID', and 'SIL'
    output_data_wstc = filtered_data_wstc[[theme_column, 'FIS TIS ID', sil_column]].copy()

    # Remove duplicates based on the first TIS ID column
    output_data_wstc = output_data_wstc.drop_duplicates(subset='FIS TIS ID')
    
    # Print the resulting DataFrame
    #print("Filtered Data with the First FIS TIS ID:")
    #print(output_data_wstc)

    # Optionally save the output DataFrame to an Excel file
    output_file_wstc = r"C:\2.2\VAT_REPORT_SCRIPT\Filtered_change_log\filtered_changelog_wstc.xlsx"
    output_data_wstc.to_excel(output_file_wstc, index=False, sheet_name='Filtered Data')
    print(f"Filtered data has been saved to {output_file_wstc}.")
else:
    print(f"Warning: Required columns were not found in the filtered DataFrame: {fs_tis_id_column_name}, {theme_column}, {sil_column}.")

output_data_wstc_path = r"C:\2.2\VAT_REPORT_SCRIPT\Filtered_change_log\filtered_changelog_wstc.xlsx"
output_data_wstc1 = pd.read_excel(output_data_wstc_path, engine='openpyxl')
# Ensure the relevant columns exist
if 'FIS TIS ID' in output_data_wstc1.columns and 'FS TIS ID' in wstc_original_df.columns:
    
    # Create sets for comparison
    fis_tis_ids = set(output_data_wstc1['FIS TIS ID'])
    fs_tis_ids = set(wstc_original_df['FS TIS ID'])

 
    missing_in_fs = fis_tis_ids - fs_tis_ids  
    #missing_in_fis = fs_tis_ids - fis_tis_ids  

    # Create a DataFrame for mismatches including Theme and SIL columns
    mismatch_rows = []  # List for capturing mismatch entries

    # Gather details for mismatches
    for missing_id in missing_in_fs:
        # Get the corresponding Theme and SIL for the missing FIS TIS ID
        row = output_data_wstc[output_data_wstc['FIS TIS ID'] == missing_id]
        if not row.empty:
            theme_value = row['Theme'].values[0]  # Adjust based on your actual DataFrame structure
            sil_value = row['SIL'].values[0]
            mismatch_rows.append({
                'Theme': theme_value,
                'Missing FIS TIS ID': missing_id,
                'SIL': sil_value
            })

    # Create DataFrame from mismatched rows
    mismatch_df = pd.DataFrame(mismatch_rows)
    mismatch_df = mismatch_df.sort_values(by=['Theme'], ascending= True)

    # Save mismatches to a new sheet in the original CTC file
    if not mismatch_df.empty:
        with pd.ExcelWriter(wstc_data, engine='openpyxl', mode='a') as writer:  # Append to existing file
            mismatch_df.to_excel(writer, index=False, sheet_name='Mismatch_FRL_final_wstc')
        #print(f"Mismatches have been written to {wstc_data} in the 'Mismatch_FRL_final_wstc' sheet.")
    else:
        print("No mismatches found between 'FIS TIS ID' and 'FS TIS ID'.")
else:
    print("Warning: Required columns not found in the respective DataFrames.")
#-------------------------------------------------------------------------------------------------------#

#----------------comparing requirements in change log and requirements in WSIC vat Repot-----------#
file_path_wsic = r"C:\2.2\VAT_REPORT_SCRIPT\Changelog\WSIC_Change_Log_E_F 2.xlsm"
data_wsic = pd.read_excel(file_path_wsic, header=[0, 1, 2, 3, 4], sheet_name='Result', engine='openpyxl')
wsic_original_df = pd.read_excel(wsic_data)
# Print the original columns to see their structure
#print("Original Columns in DataFrame:")
#print(data_wstc.columns.tolist())  # List all columns for inspection
new_columns = []
tis_id_count = 0

for col in data_wsic.columns:
    if 'TIS ID' in col:
        tis_id_count += 1
        new_columns.append(f"TIS ID_{tis_id_count}")  # Make TIS ID unique
    elif 'Theme' in col:
        new_columns.append("Theme")  # Assign a unique name (optional if you use one name for all)
    elif 'SIL' in col:
        new_columns.append("SIL")  # Assign a unique name (optional if you use one name for all)
    elif 'Implementation Language' in col:
        new_columns.append("Implementation Language")
    elif 'Status' in col:
        new_columns.append("Status")        
    else:
        new_columns.append(col)  # Keep other columns unchanged
# Assign the new unique column names
data_wsic.columns = new_columns  # Assign the new unique column names

# Define the column for Implementation Process
implementation_process_column = 'Implementation Language'
status = 'Status'
# Filter out rows where Implementation Process is either 'Atomic # Ada SIL4' or 'Atomic # Ada SIL0'
if implementation_process_column in data_wsic.columns:
    filtered_data_wsic = data_wsic[
        (data_wsic[implementation_process_column] != 'Atomic # Ada SIL4') & 
        (data_wsic[implementation_process_column] != 'Atomic # Ada SIL0') &
        (data_wsic[implementation_process_column] != 'C++') &
        (data_wsic[status] != 'D')
        
    ]
else:
    print(f"Warning: Column '{implementation_process_column}' not found in the DataFrame.")

# Ensure the relevant columns exist after filtering
fs_tis_id_column_name = 'TIS ID_1'  # Adjust based on your actual DataFrame
theme_column = 'Theme'   # Adjust if necessary
sil_column = 'SIL'       # Adjust if necessary

# Rename the first TIS ID column to 'FIS TIS ID' if it exists
if (fs_tis_id_column_name in filtered_data_wsic.columns) and (theme_column in filtered_data_wsic.columns) and (sil_column in filtered_data_wsic.columns):
    filtered_data_wsic.rename(columns={fs_tis_id_column_name: 'FIS TIS ID'}, inplace=True)

    # Create a new DataFrame with 'Theme', 'FIS TIS ID', and 'SIL'
    output_data_wsic = filtered_data_wsic[[theme_column, 'FIS TIS ID', sil_column]].copy()

    # Remove duplicates based on the first TIS ID column
    output_data_wsic = output_data_wsic.drop_duplicates(subset='FIS TIS ID')
    
    # Print the resulting DataFrame
    #print("Filtered Data with the First FIS TIS ID:")
    #print(output_data_wsic)

    # Optionally save the output DataFrame to an Excel file
    output_file_wsic = r"C:\2.2\VAT_REPORT_SCRIPT\Filtered_change_log\filtered_changelog_wsic.xlsx"
    output_data_wsic.to_excel(output_file_wsic, index=False, sheet_name='Filtered Data')
    print(f"Filtered data has been saved to {output_file_wsic}.")
else:
    print(f"Warning: Required columns were not found in the filtered DataFrame: {fs_tis_id_column_name}, {theme_column}, {sil_column}.")

output_data_wsic_path = r"C:\2.2\VAT_REPORT_SCRIPT\Filtered_change_log\filtered_changelog_wsic.xlsx"
output_data_wsic1 = pd.read_excel(output_data_wsic_path, engine='openpyxl')
# Ensure the relevant columns exist
if 'FIS TIS ID' in output_data_wsic1.columns and 'FS TIS ID' in wsic_original_df.columns:
    
    # Create sets for comparison
    fis_tis_ids = set(output_data_wsic1['FIS TIS ID'])
    fs_tis_ids = set(wsic_original_df['FS TIS ID'])

   
    missing_in_fs = fis_tis_ids - fs_tis_ids  
    #missing_in_fis = fs_tis_ids - fis_tis_ids  

    # Create a DataFrame for mismatches including Theme and SIL columns
    mismatch_rows = []  # List for capturing mismatch entries

    # Gather details for mismatches
    for missing_id in missing_in_fs:
        # Get the corresponding Theme and SIL for the missing FIS TIS ID
        row = output_data_wsic[output_data_wsic['FIS TIS ID'] == missing_id]
        if not row.empty:
            theme_value = row['Theme'].values[0]  # Adjust based on your actual DataFrame structure
            sil_value = row['SIL'].values[0]
            mismatch_rows.append({
                'Theme': theme_value,
                'Missing FIS TIS ID': missing_id,
                'SIL': sil_value
            })

    # Create DataFrame from mismatched rows
    mismatch_df = pd.DataFrame(mismatch_rows)
    mismatch_df = mismatch_df.sort_values(by=['Theme'], ascending= True)

    # Save mismatches to a new sheet in the original CTC file
    if not mismatch_df.empty:
        with pd.ExcelWriter(wsic_data, engine='openpyxl', mode='a') as writer:  # Append to existing file
            mismatch_df.to_excel(writer, index=False, sheet_name='Mismatch_FRL_final_wsic')
        #print(f"Mismatches have been written to {wsic_data} in the 'Mismatch_FRL_final_wsic' sheet.")
    else:
        print("No mismatches found between 'FIS TIS ID' and 'FS TIS ID'.")
else:
    print("Warning: Required columns not found in the respective DataFrames.")
#-------------------------------------------------------------------------------------------------------#
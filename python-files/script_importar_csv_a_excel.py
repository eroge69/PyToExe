
import os
import sys
import pandas as pd
from openpyxl import Workbook

# Ruta de la carpeta con los CSV (misma ruta donde está el exe)
carpeta_csv = os.path.dirname(sys.executable)

# Columnas origen y destino (por índice de Excel: A=0, B=1, ..., Z=25)
mapeo_columnas = {
    0: 1,   # A → B
    14: 6,  # O → G
    15: 7,  # P → H
    10: 9,  # K → J
    2: 10,  # C → K
    11: 16  # L → Q
}

# Obtener lista de archivos CSV ordenados por nombre (formato AAAAMMDD)
archivos_csv = sorted([f for f in os.listdir(carpeta_csv) if f.endswith('.csv')])

# Crear nuevo libro de Excel
wb = Workbook()
ws = wb.active

# Fila inicial
fila_actual = 1

for archivo in archivos_csv:
    ruta_archivo = os.path.join(carpeta_csv, archivo)
    
    # Extraer la fecha del nombre del archivo (formato AAAAMMDD)
    fecha = archivo.split('.')[0]
    
    # Leer CSV como texto (sin detectar tipos)
    df = pd.read_csv(ruta_archivo, dtype=str, header=None)
    
    # Copiar columnas deseadas y añadir la fecha en la columna A
    for fila_idx in range(len(df)):
        ws.cell(row=fila_actual + fila_idx, column=1, value=fecha)  # Añadir la fecha en la columna A
        for col_origen, col_destino in mapeo_columnas.items():
            valor = df.iat[fila_idx, col_origen] if col_origen < df.shape[1] else ''
            ws.cell(row=fila_actual + fila_idx, column=col_destino + 1, value=valor)

    fila_actual += len(df)

# Guardar archivo Excel
wb.save('resultado.xlsx')
print("Archivo 'resultado.xlsx' creado con éxito.")


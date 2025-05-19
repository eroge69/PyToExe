import tkinter as tk
from tkinter import messagebox
import os
import random
import string
from openpyxl import load_workbook

# Setup
window = tk.Tk()
window.title("Otto Cinema")

# Load Excel
workbook_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cinemaexcel.xlsx")

if not os.path.exists(workbook_path):
    messagebox.showerror("File Error", f"Workbook not found at {workbook_path}")
    exit(1)

try:
    wb = load_workbook(workbook_path)
except Exception as e:
    messagebox.showerror("Excel Error", f"Failed to load workbook: {e}")
    exit(1)


mummySeats = wb["Mummy"]
mummySeats2 = wb["Mummy2"]
AlSeats = wb["Al"]
AlSeats2 = wb["Al2"]
PhantomSeats = wb["Phantom"]
PhantomSeats2 = wb["Phantom2"]
ordersSheet = wb["Orders"]

# Data and arrays
selectedArrayMap = {}  # (movie, time) => list of (row, col)
currentMovie = None
currentTime = None
buy_button = None
selectedFood = []
foodVars = {}
seat_price = 0.10
food_prices = {
    "European lo ba bung": 0.15,
    "Chicken feet": 0.12,
    "Fried banana": 0.06,
    "Intestine tacos": 0.15,
    "Ethanoic acid": 0.10,
    "Black and white soda": 0.08,
    "Mayo and ketchup 0.5 L": 0.08
}

# Initialize foodVars with default quantities (0 for each food item)
foodVars = {item: 0 for item in food_prices.keys()}

# Get sheet/array for movie
def get_sheet_and_array(movie, time):
    if movie == "The Mummy":
        return (mummySeats if time == "14:30" else mummySeats2, mummySeatsArray if time == "14:30" else mummySeatsArray2)
    elif movie == "Weird: the Al Yankovic Story":
        return (AlSeats if time == "12:30" else AlSeats2, AlSeatsArray if time == "12:30" else AlSeatsArray2)
    else:
        return (PhantomSeats if time == "16:30" else PhantomSeats2, PhantomSeatsArray if time == "16:30" else PhantomSeatsArray2)

# Load seats
def load_seats(sheet):
    return [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_row=8, min_col=1, max_col=4)]

mummySeatsArray = load_seats(mummySeats)
mummySeatsArray2 = load_seats(mummySeats2)
AlSeatsArray = load_seats(AlSeats)
AlSeatsArray2 = load_seats(AlSeats2)
PhantomSeatsArray = load_seats(PhantomSeats)
PhantomSeatsArray2 = load_seats(PhantomSeats2)

# Frames
mainFrame = tk.Frame(window)
mainFrame.pack(fill="both", expand=True)
homepage = tk.Frame(mainFrame)
timesForMummy = tk.Frame(mainFrame)
timesForAl = tk.Frame(mainFrame)
timesForPhantom = tk.Frame(mainFrame)
seatFrame = tk.Frame(mainFrame)
foodFrame = tk.Frame(mainFrame)
receiptSideFrame = tk.Frame(mainFrame, width=200)
receiptSideFrame.grid(row=0, column=1, sticky="ns")
checkoutFrame = tk.Frame(window)

# Variables
selectedSeatsVar = tk.StringVar()

# Utility Functions
def seek_and_destroy():
    for sheet in [mummySeats, mummySeats2, AlSeats, AlSeats2, PhantomSeats, PhantomSeats2]:
        for row in sheet.iter_rows(min_row=2, max_row=8, min_col=1, max_col=4):
            for cell in row:
                if cell.value == "1":
                    cell.value = "0"
    wb.save(workbook_path)

def elimination():
    if messagebox.askokcancel("Quit", "Are you sure you want to quit?"):
        try:
            seek_and_destroy()
            wb.save(workbook_path)
        except Exception as e:
            print(f"Warning: Could not save workbook on exit. {e}")
        window.destroy()

def retfunc(homeV):
    global currentMovie, currentTime
    currentMovie = None
    currentTime = None
    for f in [seatFrame, timesForMummy, timesForAl, timesForPhantom, homepage, foodFrame, checkoutFrame]:
        f.grid_forget()
    if homeV == "Mummy":
        timesForMummy.grid()
    elif homeV == "Al Yankovic":
        timesForAl.grid()
    elif homeV == "Phantom":
        timesForPhantom.grid()
    else:
        homepage.grid()

def updateSelectedSeatsLabel():
    global buy_button
    seat_lines = []
    for (movie, time), seats in selectedArrayMap.items():
        for r, c in seats:
            seat_lines.append(f"{movie} at {time}: Row {r + 1}, Seat {c + 1}")

    food_lines = [f"{item} x {quantity}" for item, quantity in foodVars.items() if quantity > 0]
    total_price = (
        sum(len(seats) for seats in selectedArrayMap.values()) * seat_price
        + sum(food_prices[f] * qty for f, qty in foodVars.items())
    )

    # List all seats without truncation
    text = "Seats:\n" + "\n".join(seat_lines)
    if food_lines:
        text += "\n\nFood:\n" + "\n".join(food_lines)
    text += f"\n\nTotal: £{total_price:.2f}"

    selectedSeatsVar.set(text)

    # Enable or disable the Buy button
    if seat_lines:  # Seats selected
        buy_button.config(state="normal")
    else:  # No seats selected
        buy_button.config(state="disabled")


def openSeats(movie, showtime):
    global currentMovie, currentTime
    currentMovie = movie
    currentTime = showtime

    key = (movie, showtime)
    if key not in selectedArrayMap:
        selectedArrayMap[key] = []

    selectedArray = selectedArrayMap[key]

    for f in [seatFrame, timesForMummy, timesForAl, timesForPhantom, homepage, foodFrame, checkoutFrame]:
        f.grid_forget()

    for widget in seatFrame.winfo_children():
        widget.destroy()
    seatFrame.grid()

    sheet, seats = get_sheet_and_array(movie, showtime)

    tk.Label(seatFrame, text=f"{movie} - {showtime}", height=3).grid(row=0, columnspan=4)

    def select_seat(r, c):
        # Get the current value of the seat
        val = str(seats[r][c]) if seats[r][c] is not None else "0"

        # Check if the seat is already sold
        if val == "2":
            messagebox.showerror("Seat Unavailable", "This seat is already sold. Please choose another seat.")
            return  # Exit the function early if the seat is sold

        # Handle seat selection
        if val == "0":  # Seat is available
            seats[r][c] = "1"
            if (r, c) not in selectedArrayMap[key]:
                selectedArrayMap[key].append((r, c))
            sheet.cell(row=r + 2, column=c + 1, value="1")
            seat_buttons[r][c].config(bg="yellow")  # Mark the seat as selected
        elif val == "1":  # Seat is already selected by the user
            seats[r][c] = "0"
            if (r, c) in selectedArrayMap[key]:
                selectedArrayMap[key].remove((r, c))
            sheet.cell(row=r + 2, column=c + 1, value="0")
            seat_buttons[r][c].config(bg="green")  # Unmark the seat

        # Save changes and update UI
        wb.save(workbook_path)
        updateSelectedSeatsLabel()

    tk.Button(seatFrame, text="Return", width=25, height=3, command=lambda: retfunc(movie.split()[0])).grid(row=1, column=0, columnspan=2, pady=20)
    tk.Button(seatFrame, text="Exit", width=25, height=3, command=elimination).grid(row=1, column=2, columnspan=2, pady=10)

    global seat_buttons
    seat_buttons = []
    for r in range(len(seats)):
        row_buttons = []
        for c in range(len(seats[r])):
            val = str(seats[r][c]) if seats[r][c] is not None else "0"
            color = "green" if val == "0" else "yellow" if val == "1" else "red"

            # Ensure all buttons are selectable
            btn = tk.Button(
                seatFrame,
                text=f"{r + 1}-{c + 1}",
                bg=color,
                width=5,
                height=2,
                command=lambda r=r, c=c: select_seat(r, c)
            )
            btn.grid(column=c, row=r + 2, padx=2, pady=5)
            row_buttons.append(btn)
        seat_buttons.append(row_buttons)


# Finalize Order
def buy():
    mainFrame.pack_forget()
    checkoutFrame.pack(fill="both", expand=True)
    for widget in checkoutFrame.winfo_children():
        widget.destroy()

    tk.Label(checkoutFrame, text="Confirm your order below").pack(pady=10)
    tk.Label(checkoutFrame, textvariable=selectedSeatsVar, justify="left").pack(pady=5)
    tk.Button(checkoutFrame, text="Finalize", command=finalizeOrder).pack(pady=10)
    tk.Button(checkoutFrame, text="Cancel", command=lambda: [checkoutFrame.pack_forget(), mainFrame.pack()]).pack()

def finalizeOrder():
    all_orders = []
    has_seats = False

    for (movie, time), seats in selectedArrayMap.items():
        if seats:
            has_seats = True
            sheet, seat_array = get_sheet_and_array(movie, time)
            seat_list = [f"Row {r+1} Seat {c+1}" for r, c in seats]

            for row, col in seats:
                seat_array[row][col] = "2"
                sheet.cell(row=row + 2, column=col + 1, value="2")

            all_orders.append({
                "movie": movie,
                "time": time,
                "seats": seat_list
            })

    if not has_seats:
        messagebox.showwarning("No Seats", "You haven't selected any seats.")
        return

    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    total_price = (
        sum(len(order["seats"]) * seat_price for order in all_orders)
        + sum(food_prices[item] * count for item, count in foodVars.items())
    )

    seat_summary = "; ".join(
        f"{len(order['seats'])} seats for {order['movie']} at {order['time']} ({', '.join(order['seats'])})"
        for order in all_orders
    )
    food_summary = ", ".join([f"{item} x{count}" for item, count in foodVars.items() if count > 0])

    ordersSheet.append([
        code,
        seat_summary,
        food_summary if food_summary else "None",
        f"£{total_price:.2f}"
    ])

    wb.save(workbook_path)
    messagebox.showinfo("Order Confirmed", f"Your code: {code}\nPlease pay at the counter.")
    window.destroy()

# Food Menu
def updateFoodSelection():
    selectedFood.clear()
    for item, quantity in foodVars.items():
        if quantity > 0:
            selectedFood.extend([item] * quantity)
    updateSelectedSeatsLabel()


def openFoodMenu():
    for widget in foodFrame.winfo_children():
        widget.destroy()

    for f in [seatFrame, homepage, timesForMummy, timesForAl, timesForPhantom]:
        f.grid_forget()
    foodFrame.grid()

    def increaseFood(item, quantity_label):
        if foodVars[item] < 10:  # Cap at 10
            foodVars[item] += 1
            quantity_label.config(text=f"Quantity: {foodVars[item]}")
            updateFoodSelection()
        else:
            messagebox.showwarning("Limit Reached", f"You can only order up to 10 of {item}.")

    def decreaseFood(item, quantity_label):
        if foodVars[item] > 0:
            foodVars[item] -= 1
            quantity_label.config(text=f"Quantity: {foodVars[item]}")
            updateFoodSelection()

    for item in food_prices.keys():
        # Display food item and quantity
        foodFrameItem = tk.Frame(foodFrame)
        foodFrameItem.pack(anchor="w", padx=10, pady=5)

        tk.Label(foodFrameItem, text=f"{item} (£{food_prices[item]:.2f})").pack(side="left", padx=5)

        quantity_label = tk.Label(foodFrameItem, text=f"Quantity: {foodVars[item]}", width=10)
        quantity_label.pack(side="left", padx=5)

        # Increase and decrease buttons
        tk.Button(foodFrameItem, text="+", width=3, command=lambda i=item, q_label=quantity_label: increaseFood(i, q_label)).pack(side="right", padx=5)
        tk.Button(foodFrameItem, text="-", width=3, command=lambda i=item, q_label=quantity_label: decreaseFood(i, q_label)).pack(side="right", padx=5)

    tk.Button(foodFrame, text="Return to Home", command=lambda: retfunc(0)).pack(pady=10)
    tk.Button(foodFrame, text="Exit", command=elimination).pack()



# Homepage
tk.Label(homepage, text="Otto Cinema", height=2).grid()
tk.Label(homepage, text="Welcome to Otto Cinema!").grid()
tk.Label(homepage, text="Here you can reserve a seat for a movie.").grid()
tk.Label(homepage, text="Movies:", height=3).grid()
tk.Button(homepage, text="The Mummy", width=50, height=3, command=lambda: retfunc("Mummy")).grid()
tk.Button(homepage, text="Weird: the Al Yankovic Story", width=50, height=3, command=lambda: retfunc("Al Yankovic")).grid()
tk.Button(homepage, text="Phantom of the Opera", width=50, height=3, command=lambda: retfunc("Phantom")).grid()
tk.Button(homepage, text="Food", width=50, height=3, command=openFoodMenu).grid()
tk.Button(homepage, text="Exit", width=50, height=3, command=elimination).grid()

# Times
def addTimeButtons(frame, movie, times):
    tk.Label(frame, text=movie, height=3).grid()
    for t in times:
        tk.Button(frame, text=t, width=50, height=3, command=lambda m=movie, ti=t: openSeats(m, ti)).grid()
    tk.Button(frame, text="Return", width=50, height=3, command=lambda: retfunc(0)).grid()
    tk.Button(frame, text="Exit", width=50, height=3, command=elimination).grid()

addTimeButtons(timesForMummy, "The Mummy", ["14:30", "21:30"])
addTimeButtons(timesForAl, "Weird: the Al Yankovic Story", ["12:30", "19:30"])
addTimeButtons(timesForPhantom, "Phantom of the Opera", ["16:30", "23:30"])

# Side Frame
tk.Label(receiptSideFrame, text="Selected Seats").grid(row=0, column=0)
tk.Label(receiptSideFrame, text="And Foods:").grid(row=1, column=0)
selectedSeatsLabel = tk.Label(receiptSideFrame, textvariable=selectedSeatsVar, justify="left", anchor="w")
selectedSeatsLabel.grid(row=2, column=0, padx=5, pady=5)
buy_button = tk.Button(receiptSideFrame, text="Buy", command=buy, state="disabled")
buy_button.grid(row=3, column=0, pady=10)


homepage.grid()
window.protocol("WM_DELETE_WINDOW", elimination)
window.mainloop()

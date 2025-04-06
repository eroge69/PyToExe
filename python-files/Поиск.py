import requests
import os
import time
import random
from pystyle import *
from colorama import Fore, Style, init
from console.utils import wait_key
from console import fg, bg, fx
from colorama import init, Fore, Back, Style
from pprint import pprint
from random import choice
from pyfiglet import Figlet
from console.utils import clear_screen
from console.screen import sc
import glob

VK_TOKEN = "0af157510af157510af15751aa0a89e69600af10af157516a0bc15996e74fe2b440998c"
VK_API_URL = "https://api.vk.com/method/users.get"

tokens = ["api"] 
api_url = "https://leakosintapi.com/"

men = """
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣀⡠⢤⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⡴⠟⠃⠀⠀⠙⣄⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⠋⠀⠀⠀⠀⠀⠀⠘⣆⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢠⠾⢛⠒⠀⠀⠀⠀⠀⠀⠀⢸⡆⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣶⣄⡈⠓⢄⠠⡀⠀⠀⠀⣄⣷⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣿⣷⠀⠈⠱⡄⠑⣌⠆⠀⠀⡜⢻⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⡿⠳⡆⠐⢿⣆⠈⢿⠀⠀⡇⠘⡆⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢿⣿⣷⡇⠀⠀⠈⢆⠈⠆⢸⠀⠀⢣⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠘⣿⣿⣿⣧⠀⠀⠈⢂⠀⡇⠀⠀⢨⠓⣄⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣸⣿⣿⣿⣦⣤⠖⡏⡸⠀⣀⡴⠋⠀⠈⠢⡀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢠⣾⠁⣹⣿⣿⣿⣷⣾⠽⠖⠊⢹⣀⠄⠀⠀⠀⠈⢣⡀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⡟⣇⣰⢫⢻⢉⠉⠀⣿⡆⠀⠀⡸⡏⠀⠀⠀⠀⠀⠀⢇
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢨⡇⡇⠈⢸⢸⢸⠀⠀⡇⡇⠀⠀⠁⠻⡄⡠⠂⠀⠀⠀⠘
⢤⣄⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢠⠛⠓⡇⠀⠸⡆⢸⠀⢠⣿⠀⠀⠀⠀⣰⣿⣵⡆⠀⠀⠀⠀
⠈⢻⣷⣦⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⡿⣦⣀⡇⠀⢧⡇⠀⠀⢺⡟⠀⠀⠀⢰⠉⣰⠟⠊⣠⠂⠀⡸
⠀⠀⢻⣿⣿⣷⣦⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⢧⡙⠺⠿⡇⠀⠘⠇⠀⠀⢸⣧⠀⠀⢠⠃⣾⣌⠉⠩⠭⠍⣉⡇
⠀⠀⠀⠻⣿⣿⣿⣿⣿⣦⣀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣠⣞⣋⠀⠈⠀⡳⣧⠀⠀⠀⠀⠀⢸⡏⠀⠀⡞⢰⠉⠉⠉⠉⠉⠓⢻⠃
⠀⠀⠀⠀⠹⣿⣿⣿⣿⣿⣿⣷⡄⠀⠀⢀⣀⠠⠤⣤⣤⠤⠞⠓⢠⠈⡆⠀⢣⣸⣾⠆⠀⠀⠀⠀⠀⢀⣀⡼⠁⡿⠈⣉⣉⣒⡒⠢⡼⠀
⠀⠀⠀⠀⠀⠘⣿⣿⣿⣿⣿⣿⣿⣎⣽⣶⣤⡶⢋⣤⠃⣠⡦⢀⡼⢦⣾⡤⠚⣟⣁⣀⣀⣀⣀⠀⣀⣈⣀⣠⣾⣅⠀⠑⠂⠤⠌⣩⡇⠀
⠀⠀⠀⠀⠀⠀⠘⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡁⣺⢁⣞⣉⡴⠟⡀⠀⠀⠀⠁⠸⡅⠀⠈⢷⠈⠏⠙⠀⢹⡛⠀⢉⠀⠀⠀⣀⣀⣼⡇⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠈⠻⣿⣿⣿⣿⣿⣿⣿⣿⣽⣿⡟⢡⠖⣡⡴⠂⣀⣀⣀⣰⣁⣀⣀⣸⠀⠀⠀⠀⠈⠁⠀⠀⠈⠀⣠⠜⠋⣠⠁⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠙⢿⣿⣿⣿⡟⢿⣿⣿⣷⡟⢋⣥⣖⣉⠀⠈⢁⡀⠤⠚⠿⣷⡦⢀⣠⣀⠢⣄⣀⡠⠔⠋⠁⠀⣼⠃⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠻⣿⣿⡄⠈⠻⣿⣿⢿⣛⣩⠤⠒⠉⠁⠀⠀⠀⠀⠀⠉⠒⢤⡀⠉⠁⠀⠀⠀⠀⠀⢀⡿⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠙⢿⣤⣤⠴⠟⠋⠉⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠑⠤⠀⠀⠀⠀⠀⢩⠇⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
    ╔════════════════════════════════════════════╗
    ║                @Vovanhickto                ║
    ╠════════════════════════════════════════════╣
    ║       [01] Поиск       [02] Osint Vk       ║
    ║       [03] Бд (файлы)  [00] Выход          ║
    ╚════════════════════════════════════════════╝"""

meni = """

⠀⠀⠀⠀⠀⠀⠀⣰⠀⠀⣠⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣄⠀⠀⣆⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⣼⠃⠀⢰⠏⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⡆⠀⠸⣧⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⣸⡇⠀⢠⡟⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢻⡄⠀⢸⣇⠀⠀⠀⠀⠀
⠀⠀⠀⠀⢰⡿⠀⠀⣼⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣧⠀⠀⢿⡆⠀⠀⠀⠀
⠀⠀⠀⢀⣾⡇⠀⢰⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⡆⠀⢸⣿⡀⠀⠀⠀
⠀⠀⠀⢸⣿⠁⠀⣸⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣿⣧⠀⠈⣿⣇⠀⠀⠀
⠀⠀⢀⣿⡟⠀⠀⢿⣿⣶⠶⠶⠶⠶⢾⣦⡀⡴⢀⣀⣦⢀⣴⡷⠶⠶⠶⠶⣶⣿⡿⠀⠀⢹⣿⡀⠀⠀
⠀⠀⢸⣿⡇⠀⠀⣀⣉⣤⣴⣶⣶⣶⣶⣽⣿⣿⣿⣿⣿⣿⣯⣶⣶⣶⣶⣦⣤⣉⣀⠀⠀⢸⣿⡇⠀⠀
⠀⠀⠸⢿⣷⣶⠿⠟⠋⠉⠀⣠⣄⣤⣭⣿⣿⣿⣿⣿⣿⡿⣿⣭⣥⣠⣄⠀⠉⠉⠻⠿⣷⣾⡿⠇⠀⠀
⠀⠀⠀⠀⠉⠁⠀⠀⢀⣠⣾⡿⠟⢋⣹⣿⢿⣿⣿⣿⣿⡿⣿⣯⡙⠻⠿⣷⣄⡀⠀⠀⠈⠉⠁⠀⠀⠀
⠀⠀⠀⠀⠀⠀⢀⣴⠾⠋⠁⣠⣴⡾⠋⠁⣾⣿⣿⣿⣿⣷⠈⠙⢿⣦⣄⠈⠙⠷⣦⡀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⢀⣠⣶⠟⠁⠀⠀⠀⣿⣿⠁⠀⢰⣿⣿⣿⣿⣿⣿⡆⠀⠈⣿⣿⠀⠀⠀⠉⠻⣶⣄⡀⠀⠀⠀
⢰⣤⣶⣿⠟⠁⠀⠀⠀⠀⠀⣿⣿⠀⠀⠈⣿⣿⣿⣿⣿⣿⠁⠀⠀⣿⣿⠀⠀⠀⠀⠀⠈⠻⣿⣶⣤⡆
⢸⣿⣿⠁⠀⠀⠀⠀⠀⠀⠀⣿⣿⠀⠀⠀⢸⣿⣿⣿⣿⡏⠀⠀⠀⣿⣿⠀⠀⠀⠀⠀⠀⠀⠈⣿⣿⡇
⠀⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⠀⠀⠀⠈⢿⣿⣿⡿⠁⠀⠀⠀⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⠀
⠀⢿⣿⠀⠀⠀⠀⠀⠀⠀⠀⣿⣿⠀⠀⠀⠀⠸⣿⣿⠃⠀⠀⠀⠀⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⣿⡿⠀
⠀⢸⣿⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⠀⠀⠀⠀⠀⠈⠃⠀⠀⠀⠀⠀⣿⡇⠀⠀⠀⠀⠀⠀⠀⠀⣿⡇⠀
⠀⠈⣿⡇⠀⠀⠀⠀⠀⠀⠀⢸⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣿⡇⠀⠀⠀⠀⠀⠀⠀⢸⣿⠃⠀
⠀⠀⢸⣇⠀⠀⠀⠀⠀⠀⠀⠘⣿⡆⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢨⣿⠇⠀⠀⠀⠀⠀⠀⠀⣸⡏⠀⠀
⠀⠀⠈⢿⡄⠀⠀⠀⠀⠀⠀⠀⢿⡇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⡿⠀⠀⠀⠀⠀⠀⠀⢠⡿⠁⠀⠀
⠀⠀⠀⠘⣧⠀⠀⠀⠀⠀⠀⠀⢸⣧⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣸⡇⠀⠀⠀⠀⠀⠀⠀⣼⠃⠀⠀⠀
⠀⠀⠀⠀⠸⣇⠀⠀⠀⠀⠀⠀⠀⢿⡄⠀⠀⠀⠀⠀⠀⠀⠀⢀⡿⠀⠀⠀⠀⠀⠀⠀⣰⠇⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠈⠀⠀⠀⠀⠀⠀⠀⠸⣇⠀⠀⠀⠀⠀⠀⠀⠀⢸⠇⠀⠀⠀⠀⠀⠀⠀⠁⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢹⡄⠀⠀⠀⠀⠀⠀⢠⡏⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢳⠀⠀⠀⠀⠀⠀⡞⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
              ╔═════════════╗
              ║ Нажми Enter   ║
              ╚═════════════╝
"""
Anime.Fade(Center.Center(meni), Colors.red_to_green, Colorate.Vertical, interval=0.0045, enter=True)

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def print_section(title):
    print(Colorate.Horizontal(Colors.red_to_green, f"\n{'=' * 50}"))
    print(Colorate.Horizontal(Colors.red_to_green, f"{title.center(50)}"))
    print(Colorate.Horizontal(Colors.red_to_green, f"{'=' * 50}\n"))

def print_data(data):
    for key, value in data.items():
        print(f"{key.ljust(20)}: {value}")
    print(Colorate.Horizontal(Colors.red_to_green, "\n" + "-" * 50 + "\n"))

def input_colored(prompt, color):
    return input(Colorate.Horizontal(color, prompt))

def search():
    search_query = input_colored("    Введите запрос для поиска: ", Colors.red_to_green)
    token = choice(tokens)

    data = {
        "token": token,
        "request": search_query,
        "limit": 100,
        "lang": "ru"
    }

    try:
        response = requests.post(api_url, json=data)
        response.raise_for_status()
        result = response.json()
    except requests.exceptions.RequestException as e:
        print(Colorate.Horizontal(Colors.red_to_green, f"  Ошибка при выполнении запроса: {e}"))
        return
    except ValueError as e:
        print(Colorate.Horizontal(Colors.red_to_green, f"  Ошибка при обработке JSON: {e}"))
        print(Colorate.Horizontal(Colors.red_to_green, f"  Ответ сервера: {response.text}"))
        return

    if "List" in result:
        if "No results found" in result["List"]:
            print_section("В базе данных нет совпадений")
            return
        if not result["List"]:
            print_section("В базе данных нет совпадений")
            return
        for source, info in result["List"].items():
            print_section(f"Найдено в '{source}':")
            if "Data" in info:
                for entry in info["Data"]:
                    print_data(entry)
    else:
        print_section("В базе данных нет совпадений")

import csv
import json
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed

def search_data_in_file(file_path, search_query):
    found_data = []
    try:
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8-sig') as file:
                reader = csv.reader(file)
                headers = next(reader)
                for row in reader:
                    full_name = ' '.join(row[2:5]).strip('\\"').lower()
                    if search_query.lower() in full_name:
                        found_data.append(dict(zip(headers, row)))
                    else:
                        for value in row:
                            if search_query.lower() in value.lower():
                                found_data.append(dict(zip(headers, row)))
                                break
        elif file_path.endswith('.xlsx'):
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                full_name = ' '.join(str(cell) for cell in row[2:5]).strip('\\"').lower()
                if search_query.lower() in full_name:
                    found_data.append(dict(zip(headers, row)))
                else:
                    for value in row:
                        if search_query.lower() in str(value).lower():
                            found_data.append(dict(zip(headers, row)))
                            break
        elif file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as file:
                lines = file.readlines()
                for line in lines:
                    if search_query.lower() in line.lower():
                        found_data.append({'Text': line.strip()})
        elif file_path.endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                for item in data:
                    if isinstance(item, dict):
                        for key, value in item.items():
                            if search_query.lower() in str(value).lower():
                                found_data.append(item)
                                break
        elif file_path.endswith('.xml'):
            tree = ET.parse(file_path)
            root = tree.getroot()
            for elem in root.iter():
                if search_query.lower() in elem.text.lower():
                    found_data.append({elem.tag: elem.text})
        elif file_path.endswith('.html'):
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                matches = re.findall(r'<[^>]+>|[^<]+', content)
                for match in matches:
                    if search_query.lower() in match.lower():
                        found_data.append({'HTML Content': match.strip()})
        else:
            print(Colorate.Horizontal(Colors.red_to_green, f"Неподдерживаемый формат файла: {file_path}"))
    except Exception as e:
        print(Colorate.Horizontal(Colors.red_to_green, f"Ошибка при обработке файла {file_path}: {e}"))
    return found_data

def search_in_files():
    search_query = input_colored("    Введите запрос для поиска: ", Colors.red_to_green)
    base_dir = "База данных"
    if not os.path.exists(base_dir):
        print_section("Папка 'База данных' не найдена.")
        return

    supported_formats = ["*.csv", "*.xlsx", "*.txt", "*.json", "*.xml", "*.html"]
    files = []
    for file_format in supported_formats:
        files.extend(glob.glob(os.path.join(base_dir, file_format)))

    if not files:
        print_section("Нет файлов для поиска в папке 'База данных'.")
        return

    all_found_data = {}
    total_results = 0
    start_time = time.time()

    with ThreadPoolExecutor() as executor:
        futures = {executor.submit(search_data_in_file, file_path, search_query): file_path for file_path in files}
        for future in as_completed(futures):
            file_path = futures[future]
            try:
                data = future.result()
                if data:
                    all_found_data[file_path] = data
                    total_results += len(data)
            except Exception as e:
                print(Colorate.Horizontal(Colors.red_to_green, f"Ошибка при обработке файла {file_path}: {e}"))

    elapsed_time = time.time() - start_time
    time_str = f"{elapsed_time:.2f} секунд" if elapsed_time < 60 else f"{int(elapsed_time // 60)} минут {elapsed_time % 60:.2f} секунд"

    for file_path, data in all_found_data.items():
        print_section(f"Файл: {os.path.basename(file_path)}")
        if len(data) > 100:
            print(Colorate.Horizontal(Colors.red_to_green, f"Размер информации слишком большой ({len(data)} записей), смотрите файл: output.txt"))
            with open("output.txt", "a", encoding="utf-8") as f:
                f.write(f"Файл: {file_path}\n")
                for found_row in data:
                    f.write(str(found_row) + "\n")
        else:
            for found_row in data:
                print_data(found_row)

    print(Colorate.Horizontal(Colors.red_to_green, f"\nВсего найдено запросов: {total_results}"))
    print(Colorate.Horizontal(Colors.red_to_green, f"Время выполнения: {time_str}"))

def input_colored(prompt, color):
    return input(Colorate.Horizontal(color, prompt))

def vk_osint():
    vk_user_id = input_colored("    Введите ID или username пользователя ВКонтакте: ", Colors.red_to_green)
    params = {
        "access_token": VK_TOKEN,
        "v": "5.131",
        "user_ids": vk_user_id,
        "fields": "first_name,last_name,status,sex,bdate,city,country,photo_max_orig,mobile_phone,home_phone,connections,site,about,activities,books,games,movies,music,tv,quotes,schools,universities,career,military,relation,personal,counters"
    }

    try:
        response = requests.get(VK_API_URL, params=params)
        response.raise_for_status()
        data = response.json()
    except requests.exceptions.RequestException as e:
        print(Colorate.Horizontal(Colors.red_to_green, f"  Ошибка при выполнении запроса: {e}"))
        return
    except ValueError as e:
        print(Colorate.Horizontal(Colors.red_to_green, f"  Ошибка при обработке JSON: {e}"))
        print(Colorate.Horizontal(Colors.red_to_green, f"  Ответ сервера: {response.text}"))
        return

    if "response" not in data or not data["response"]:
        print_section("Пользователь не найден или произошла ошибка.")
        return

    user = data["response"][0]
    if "deactivated" in user:
        print_section("Этот профиль ВКонтакте удалён.")
        return

    first_name = user.get("first_name", "")
    last_name = user.get("last_name", "")
    status = user.get("status", "")
    sex = "Ж" if user.get("sex") == 1 else ("М" if user.get("sex") == 2 else "")
    bdate = user.get("bdate", "")
    city = user.get("city", {}).get("title", "")
    country = user.get("country", {}).get("title", "")
    photo_url = user.get("photo_max_orig", "")
    mobile_phone = user.get("mobile_phone", "")
    home_phone = user.get("home_phone", "")
    site = user.get("site", "")
    about = user.get("about", "")
    activities = user.get("activities", "")
    books = user.get("books", "")
    games = user.get("games", "")
    movies = user.get("movies", "")
    music = user.get("music", "")
    tv = user.get("tv", "")
    quotes = user.get("quotes", "")
    relation = user.get("relation", "")
    relation_map = {
        1: "Не женат/не замужем",
        2: "Есть друг/есть подруга",
        3: "Помолвлен/помолвлена",
        4: "Женат/замужем",
        5: "Всё сложно",
        6: "В активном поиске",
        7: "Влюблён/влюблена",
        8: "В гражданском браке"
    }
    relation = relation_map.get(relation, "")
    personal = user.get("personal", {})
    political = personal.get("political", "")
    political_map = {
        1: "Коммунистические",
        2: "Социалистические",
        3: "Умеренные",
        4: "Либеральные",
        5: "Консервативные",
        6: "Монархические",
        7: "Ультраконсервативные",
        8: "Индифферентные",
        9: "Либертарианские"
    }
    political = political_map.get(political, "")
    religion = personal.get("religion", "")
    inspired_by = personal.get("inspired_by", "")
    people_main = personal.get("people_main", "")
    people_main_map = {
        1: "Ум и креативность",
        2: "Доброта и честность",
        3: "Красота и здоровье",
        4: "Власть и богатство",
        5: "Смелость и упорство",
        6: "Юмор и жизнелюбие"
    }
    people_main = people_main_map.get(people_main, "")
    life_main = personal.get("life_main", "")
    life_main_map = {
        1: "Семейная жизнь",
        2: "Карьера и деньги",
        3: "Развлечения и отдых",
        4: "Наука и исследования",
        5: "Совершенствование мира",
        6: "Саморазвитие",
        7: "Красота и искусство",
        8: "Слава и влияние"
    }
    life_main = life_main_map.get(life_main, "")
    smoking = personal.get("smoking", "")
    smoking_map = {
        1: "Резко негативное",
        2: "Негативное",
        3: "Компромиссное",
        4: "Нейтральное",
        5: "Положительное"
    }
    smoking = smoking_map.get(smoking, "")
    alcohol = personal.get("alcohol", "")
    alcohol_map = {
        1: "Резко негативное",
        2: "Негативное",
        3: "Компромиссное",
        4: "Нейтральное",
        5: "Положительное"
    }
    alcohol = alcohol_map.get(alcohol, "")
    counters = user.get("counters", {})
    followers_count = counters.get("followers", 0)
    friends_count = counters.get("friends", 0)
    photos_count = counters.get("photos", 0)
    videos_count = counters.get("videos", 0)
    audios_count = counters.get("audios", 0)
    notes_count = counters.get("notes", 0)
    career = user.get("career", [])  
    military = user.get("military", [])
    schools = user.get("schools", [])
    universities = user.get("universities", [])

    print_section("Информация о профиле ВК")
    print_if_not_empty("- Имя:", first_name)
    print_if_not_empty("- Фамилия:", last_name)
    print_if_not_empty("- Статус:", status)
    print_if_not_empty("- Пол:", sex)
    print_if_not_empty("- Дата рождения:", bdate)
    print_if_not_empty("- Страна:", country)
    print_if_not_empty("- Город:", city)
    print_if_not_empty("- Мобильный телефон:", mobile_phone)
    print_if_not_empty("- Домашний телефон:", home_phone)
    print_if_not_empty("- Сайт:", site)
    print_if_not_empty("- О себе:", about)
    print_if_not_empty("- Интересы:", activities)
    print_if_not_empty("- Любимые книги:", books)
    print_if_not_empty("- Любимые игры:", games)
    print_if_not_empty("- Любимые фильмы:", movies)
    print_if_not_empty("- Любимая музыка:", music)
    print_if_not_empty("- Любимые телешоу:", tv)
    print_if_not_empty("- Любимые цитаты:", quotes)
    print_if_not_empty("- Семейное положение:", relation)
    print_if_not_empty("- Политические предпочтения:", political)
    print_if_not_empty("- Религия:", religion)
    print_if_not_empty("- Вдохновляется:", inspired_by)
    print_if_not_empty("- Главное в людях:", people_main)
    print_if_not_empty("- Главное в жизни:", life_main)
    print_if_not_empty("- Отношение к курению:", smoking)
    print_if_not_empty("- Отношение к алкоголю:", alcohol)
    print_if_not_empty("- Количество подписчиков:", followers_count)
    print_if_not_empty("- Количество друзей:", friends_count)
    print_if_not_empty("- Количество фотографий:", photos_count)
    print_if_not_empty("- Количество видео:", videos_count)
    print_if_not_empty("- Количество аудиозаписей:", audios_count)
    print_if_not_empty("- Количество заметок:", notes_count)

    if photo_url:
        print_if_not_empty("- Фотография:", photo_url)

    if career:  
        print_section("Карьера")
        for job in career:
            company = job.get("company", "")
            position = job.get("position", "")
            from_year = job.get("from", "")
            until_year = job.get("until", "")
            print_if_not_empty("  - Компания:", company)
            print_if_not_empty("  - Должность:", position)
            print_if_not_empty("  - Год начала:", from_year)
            print_if_not_empty("  - Год окончания:", until_year)

    if military:
        print_section("Военная служба")
        for service in military:
            unit = service.get("unit", "")
            unit_id = service.get("unit_id", "")
            from_year = service.get("from", "")
            until_year = service.get("until", "")
            print_if_not_empty("  - Отряд:", unit)
            print_if_not_empty("  - ID отряда:", unit_id)
            print_if_not_empty("  - Год начала:", from_year)
            print_if_not_empty("  - Год окончания:", until_year)

    if schools:
        print_section("Учебные заведения")
        for school in schools:
            name = school.get("name", "")
            year_from = school.get("year_from", "")
            year_to = school.get("year_to", "")
            print_if_not_empty("  - Название:", name)
            print_if_not_empty("  - Год начала:", year_from)
            print_if_not_empty("  - Год окончания:", year_to)

    if universities:
        print_section("Университеты")
        for university in universities:
            name = university.get("name", "")
            faculty = university.get("faculty_name", "")
            graduation = university.get("graduation", "")
            print_if_not_empty("  - Название:", name)
            print_if_not_empty("  - Факультет:", faculty)
            print_if_not_empty("  - Год окончания:", graduation)

def print_if_not_empty(label, value):
    if value:
        print(f"{label} {value}")

def main_menu():
    while True:
        time.sleep(0.1)
        clear_screen()
        print(Colorate.Horizontal(Colors.red_to_green, men))
        choice = input_colored("    Выберите действие: ", Colors.red_to_green)

        if choice == "1":
            search()
            input_colored("  Нажмите Enter для возврата в меню...", Colors.red_to_green)
        elif choice == "2":
            vk_osint()
            input_colored("  Нажмите Enter для возврата в меню...", Colors.red_to_green)
        elif choice == "3":
            search_in_files()
            input_colored("  Нажмите Enter для возврата в меню...", Colors.red_to_green)
        elif choice == "0":
            print(Colorate.Horizontal(Colors.red_to_green, "  Выход из программы."))
            break
        else:
            print(Colorate.Horizontal(Colors.red_to_green, "  Неверный выбор. Попробуйте снова."))
            input_colored("  Нажмите Enter для повторного ввода...", Colors.red_to_green)

if __name__ == "__main__":
    main_menu()
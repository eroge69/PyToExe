import openpyxl
import datetime
import PySimpleGUI as sg
import os
import sys

def update_maintenance_schedule(date, operator, machine, problem, filename='C:\\Users\\dp161\\Desktop\\Maintenance Schedule 0.1.xlsx'):
    """
    Updates the maintenance schedule spreadsheet with new data.

    Args:
        date (datetime): The date and time of the maintenance request.
        operator (str): The name of the operator.
        machine (str): The name of the machine.
        problem (str): A brief description of the problem.
        filename (str): The name of the Excel file.
    """

    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        # Create a new workbook and sheet if the file doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Date"
        sheet["B1"] = "Operator"
        sheet["C1"] = "Machine"
        sheet["D1"] = "Priority"
        sheet["E1"] = "Assigned To"
        sheet["F1"] = "Due Date"
        sheet["G1"] = "Notes"
        sheet["H1"] = "Status"
        sheet.column_dimensions["G"].width = 50 # Adjust notes column width

    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value=date)
    sheet.cell(row=row, column=2, value=operator)
    sheet.cell(row=row, column=3, value=machine)
    sheet.cell(row=row, column=4, value="Medium")  # Default priority to medium
    sheet.cell(row=row, column=7, value=problem)
    sheet.cell(row=row, column=8, value="Open")

    workbook.save(filename)
    sg.popup("Maintenance request added successfully!")

def main():

    sg.theme('LightBlue2')  # Add a touch of color

    layout = [
        [sg.Text("Maintenance Request Form")],
        [sg.Text("Date and Time:"), sg.Input(default_text=datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), key="-DATE-")],
        [sg.Text("Operator:"), sg.Input(key="-OPERATOR-")],
        [sg.Text("Machine:"), sg.Input(key="-MACHINE-")],
        [sg.Text("Problem Description:"), sg.Multiline(key="-PROBLEM-", size=(40, 5))],
        [sg.Button("Submit"), sg.Button("Cancel")]
    ]

    window = sg.Window("Maintenance", layout)

    while True:
        event, values = window.read()
        if event == "Cancel" or event == sg.WIN_CLOSED:
            break
        if event == "Submit":
            try:
                date_str = values["-DATE-"]
                date = datetime.datetime.strptime(date_str, "%Y-%m-%d %H:%M")
                operator = values["-OPERATOR-"]
                machine = values["-MACHINE-"]
                problem = values["-PROBLEM-"]

                if not operator or not machine or not problem:
                    sg.popup_error("Please fill in all fields.")
                    continue

                update_maintenance_schedule(date, operator, machine, problem)
                window["-DATE-"].update(datetime.datetime.now().strftime("%Y-%m-%d %H:%M")) #reset date to current time
                window["-OPERATOR-"].update("")
                window["-MACHINE-"].update("")
                window["-PROBLEM-"].update("")

            except ValueError:
                sg.popup_error("Invalid date/time format. Use YYYY-MM-DD HH:MM.")

    window.close()

if __name__ == "__main__":
    if getattr(sys, 'frozen', False):
        # Running as compiled exe
        os.chdir(sys._MEIPASS)
    main()
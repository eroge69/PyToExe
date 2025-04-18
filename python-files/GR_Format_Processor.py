
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import shutil
import os

GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

COLUMNS_TO_REMOVE = [
    'Gumruk Beyan Tutar', 'Hacim Kg', 'Shipper EORI', 'Latitude', 'Cari Kod', 'Yükseklik',
    'Yabanci Fatura No', 'ATR No', 'Fatura', 'Servis', 'NOT', 'Longitude', 'Shipper VAT ID',
    'Platform', 'Boy', 'Mikro İhracat', 'En', 'Yabanci Fatura Tarihi', 'ATR Tarihi'
]

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("GR Format Excel İşleyici")
        self.root.geometry("600x400")
        self.root.configure(bg="#f2f2f2")

        tk.Label(root, text="Excel Dosyası Seç:", font=("Arial", 12), bg="#f2f2f2").pack(pady=10)
        tk.Button(root, text="Gözat", command=self.select_file, font=("Arial", 10), bg="#4CAF50", fg="white").pack()

        self.status_label = tk.Label(root, text="", font=("Arial", 10), bg="#f2f2f2", fg="blue")
        self.status_label.pack(pady=20)

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            try:
                self.process_excel(file_path)
                self.status_label.config(text="İşlem başarıyla tamamlandı!")
                messagebox.showinfo("Başarılı", "GR Format, e-Arşiv ve ATR Listesi oluşturuldu.")
            except Exception as e:
                messagebox.showerror("Hata", f"İşlem sırasında hata oluştu:\n{e}")

    def process_excel(self, file_path):
        wb = load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        def get_col_idx(name):
            return headers.index(name) + 1 if name in headers else None

        ref2_idx = get_col_idx("Ref.2")
        customs_value_idx = get_col_idx("Customs Value")
        awb_no_idx = get_col_idx("Awb No")
        hs_code_idx = get_col_idx("HS CODE")

        atr_required_awbs = []
        max_rows = ws.max_row

        for row in range(2, max_rows + 1):
            if ref2_idx:
                ws.cell(row=row, column=ref2_idx).fill = GREEN_FILL
            if customs_value_idx:
                cv_cell = ws.cell(row=row, column=customs_value_idx)
                cv_cell.fill = YELLOW_FILL
                try:
                    val = float(str(cv_cell.value).replace(",", "."))
                    if val >= 150:
                        awb = ws.cell(row=row, column=awb_no_idx).value
                        if awb:
                            atr_required_awbs.append(str(awb))
                except:
                    pass

            hs_cell = ws.cell(row=row, column=hs_code_idx)
            if hs_cell.value in [None, ""]:
                hs_cell.value = "620342110000"

        last_col = ws.max_column + 1
        ws.cell(row=1, column=last_col).value = "IOSS"
        for row in range(2, max_rows + 1):
            ws.cell(row=row, column=last_col).value = "IM5286025168"

        updated_headers = [cell.value for cell in ws[1]]
        remove_indexes = [i + 1 for i, h in enumerate(updated_headers) if h in COLUMNS_TO_REMOVE]
        for idx in sorted(remove_indexes, reverse=True):
            ws.delete_cols(idx)

        base_dir = os.path.dirname(file_path)
        base_name = os.path.splitext(os.path.basename(file_path))[0]

        gr_format_path = os.path.join(base_dir, f"{base_name}_GR_FORMAT.xlsx")
        e_arsiv_path = os.path.join(base_dir, f"{base_name}_e_arsiv.xlsx")
        atr_list_path = os.path.join(base_dir, f"{base_name}_atr_awb_list.txt")

        wb.save(gr_format_path)
        shutil.copy(file_path, e_arsiv_path)
        with open(atr_list_path, "w") as f:
            for awb in sorted(set(atr_required_awbs)):
                f.write(f"{awb}\n")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()

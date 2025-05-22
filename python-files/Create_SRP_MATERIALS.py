import os
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

#PART 1 - CREATING THE USED MATERIAL FILE

# Excel-Dateiname
excel_filename = 'FEM-USED-MATERIALS.xlsx'

# Datenstrukturen
isotropic_data = {}
orthotropic_data = {}

# Temporäre Puffer
pending_density = None

# Regex-Patterns
material_pattern = re.compile(r'\*MATERIAL, *NAME *= *(.+)', re.IGNORECASE)
density_pattern = re.compile(r'\*DENSITY\s*,?', re.IGNORECASE)
elastic_pattern = re.compile(r'\*ELASTIC(?:, *TYPE *= *(ISOTROPIC|LAMINA))?', re.IGNORECASE)
plastic_pattern = re.compile(r'\*PLASTIC(?:, *HARDENING *= *ISOTROPIC)?', re.IGNORECASE)

# Dichte umrechnen: t/mm³ → g/cm³
def convert_density(density_value_t_per_mm3):
    try:
        return float(density_value_t_per_mm3) * 1e9
    except ValueError:
        return None

# Dateien im aktuellen Verzeichnis durchlaufen
for filename in os.listdir('.'):
    if filename.lower().endswith('.inp'):
        with open(filename, 'r', encoding='utf-8') as file:
            current_material = None
            material_type = None
            pending_density = None
            lines = file.readlines()

            i = 0
            while i < len(lines):
                line = lines[i].strip()

                # Material erkennen
                material_match = material_pattern.match(line)
                if material_match:
                    current_material = material_match.group(1).strip().strip('"')
                    material_type = None
                    pending_density = None

                    if current_material.lower().endswith("_pseudo"):
                        current_material = None
                        i += 1
                        continue

                # Dichte puffern
                elif density_pattern.match(line) and current_material:
                    i += 1
                    if i < len(lines):
                        density_line = lines[i].strip().rstrip(',')
                        values = [v.strip() for v in re.split(r'[,\s]+', density_line) if v.strip()]
                        if values:
                            density = convert_density(values[0])
                            pending_density = round(density, 6) if density is not None else None

                # ELASTIC verarbeiten
                elif elastic_pattern.match(line) and current_material:
                    elastic_type_match = elastic_pattern.match(line)
                    material_type = elastic_type_match.group(1).upper() if elastic_type_match.group(1) else 'ISOTROPIC'

                    if material_type == 'LAMINA':
                        orthotropic_data.setdefault(current_material, {
                            "Density (g/cm³)": pending_density,
                            "E1 (MPa)": None,
                            "E2 (MPa)": None,
                            "Nu12": None,
                            "G12 (MPa)": None,
                            "G13 (MPa)": None,
                            "G23 (MPa)": None
                        })
                        i += 1
                        if i < len(lines):
                            try:
                                parts = [float(p) for p in re.split(r'[,\s]+', lines[i].strip()) if p]
                                orthotropic_data[current_material]["E1 (MPa)"] = parts[0]
                                orthotropic_data[current_material]["E2 (MPa)"] = parts[1]
                                orthotropic_data[current_material]["Nu12"] = parts[2]
                                orthotropic_data[current_material]["G12 (MPa)"] = parts[3]
                                orthotropic_data[current_material]["G13 (MPa)"] = parts[4]
                                orthotropic_data[current_material]["G23 (MPa)"] = parts[5]
                            except:
                                pass

                    elif material_type == 'ISOTROPIC':
                        isotropic_data.setdefault(current_material, {
                            "Density (g/cm³)": pending_density,
                            "Young's Modulus (MPa)": None,
                            "Poisson's Ratio": None,
                            "Yield Stress (MPa)": "NOT IN INP",
                            "UTS (MPa)": "NOT IN INP"
                        })
                        i += 1
                        if i < len(lines):
                            try:
                                parts = [float(p) for p in re.split(r'[,\s]+', lines[i].strip()) if p]
                                isotropic_data[current_material]["Young's Modulus (MPa)"] = parts[0]
                                isotropic_data[current_material]["Poisson's Ratio"] = parts[1]
                            except:
                                pass

                # PLASTIC – nur für isotrope Materialien
                elif plastic_pattern.match(line) and current_material:
                    i += 1
                    if i < len(lines):
                        try:
                            # Extrahiere die Spannungswerte aus den beiden Zeilen
                            plastic_line = lines[i].strip()
                            parts = [float(v) for v in plastic_line.split(',') if v.strip()]

                            if parts and material_type == 'ISOTROPIC':
                                # Der Yield Stress (MPa) befindet sich in der ersten Zeile
                                isotropic_data[current_material]["Yield Stress (MPa)"] = parts[0]

                                # Nächste Zeile extrahieren für UTS (steht in der zweiten Zeile von PLASTIC)
                                i += 1
                                if i < len(lines):
                                    plastic_line_2 = lines[i].strip()
                                    parts_2 = [float(v) for v in plastic_line_2.split(',') if v.strip()]

                                    if parts_2:
                                        # Der UTS (MPa) befindet sich in der ersten Zahl der zweiten Zeile
                                        isotropic_data[current_material]["UTS (MPa)"] = parts_2[0]

                        except Exception as e:
                            print(f"Fehler beim Extrahieren von PLASTIC für {current_material}: {e}")

                i += 1

# Excel-Datei schreiben
with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='w') as writer:
    if isotropic_data:
        df_iso = pd.DataFrame.from_dict(isotropic_data, orient='index')
        df_iso.index.name = "Material"
        df_iso.sort_index(inplace=True)
        df_iso.to_excel(writer, sheet_name="Isotropic Materials")

    if orthotropic_data:
        df_ortho = pd.DataFrame.from_dict(orthotropic_data, orient='index')
        df_ortho.index.name = "Material"
        df_ortho.sort_index(inplace=True)
        df_ortho.to_excel(writer, sheet_name="Orthotropic Materials")

# Spaltenbreiten anpassen
wb = load_workbook(excel_filename)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 2
wb.save(excel_filename)

print("✅ FEM-USED-MATERIALS.xlsx wurde erfolgreich erstellt und Spaltenbreiten angepasst.")


# PART 2 - COMPARING THE FILE WITH THE DATABASE

# Dynamisch den Pfad zum Material-Datenverzeichnis herausfinden
user_name = os.getlogin()
database_path = f"C:/Users/{user_name}/CMM_ABAQUS_MATERIAL_DATABASE/Material_Database.xlsx"

# Pfad zu der FEM-USED-MATERIALS Datei (liegt im selben Ordner wie das Skript)
fem_used_materials_path = os.path.join(os.getcwd(), "FEM-USED-MATERIALS.xlsx")

# Öffne die Dateien
fem_used_materials = openpyxl.load_workbook(fem_used_materials_path)
material_database = openpyxl.load_workbook(database_path)

# Überprüfen, ob der Reiter "Orthotropic Materials" vorhanden ist
if "Orthotropic Materials" not in fem_used_materials.sheetnames:
    raise ValueError("Der Reiter 'Orthotropic Materials' fehlt in der Datei FEM-USED-MATERIALS.xlsx")
if "Orthotropic Materials" not in material_database.sheetnames:
    raise ValueError("Der Reiter 'Orthotropic Materials' fehlt in der Datei Material_Database.xlsx")

# Lade die Reiter der relevanten Blätter
fem_used_orthotropic = fem_used_materials["Orthotropic Materials"]
fem_used_isotropic = fem_used_materials["Isotropic Materials"]

database_orthotropic = material_database["Orthotropic Materials"]
database_isotropic = material_database["Isotropic Materials"]

# Erstelle ein neues Workbook für die Ausgabedatei
materials_for_srp = openpyxl.Workbook()

# Erstelle die Reiter für "Orthotropic Materials" und "Isotropic Materials"
isotropic_sheet = materials_for_srp.create_sheet("Isotropic Materials")
orthotropic_sheet = materials_for_srp.create_sheet("Orthotropic Materials")

# Kopiere die gesamte erste Zeile (Header) von der Material_Database in die neuen Reiter
def copy_full_header(source_sheet, target_sheet):
    max_col = source_sheet.max_column
    for col in range(1, max_col + 1):
        target_sheet.cell(row=1, column=col).value = source_sheet.cell(row=1, column=col).value

# Kopiere die vollständigen Header
copy_full_header(database_isotropic, isotropic_sheet)
copy_full_header(database_orthotropic, orthotropic_sheet)

# Formatierung für fett
bold_font = Font(bold=True)

# Funktion, um Zeile 1 und Spalte A fett zu machen
def make_bold_headers(sheet, max_col):
    # Fett für Zeile 1 (Header)
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = bold_font
        # Füge einen Rahmen hinzu
        cell.border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )
    
    # Fett für die erste Spalte (Spalte A)
    for row in range(2, sheet.max_row + 1):  # Ab Zeile 2, da Zeile 1 Header ist
        cell = sheet.cell(row=row, column=1)
        cell.font = bold_font
        # Rahmen für Zelle hinzufügen
        cell.border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )

# Orange Farbe für Duplikate
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# Funktion zum Anpassen der Spaltenbreiten
def adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Holen Sie sich die Spaltenbezeichnung (z.B. "A", "B", "C")
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # Fügt etwas Puffer hinzu
        sheet.column_dimensions[column].width = adjusted_width

# Rahmen für alle Zellen mit Einträgen hinzufügen
def add_borders(sheet):
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.border = Border(
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
                left=Side(border_style="thin"),
                right=Side(border_style="thin")
            )

# Vergleiche die Zeilen in "Isotropic Materials" (Spalten B-F) mit jeder Zeile aus der "Material_Database"
def compare_isotropic():
    # Iteriere über alle Zeilen in FEM-USED-MATERIALS
    for fem_row in fem_used_isotropic.iter_rows(min_row=2, min_col=2, max_col=6, values_only=True):  # B bis F
        fem_values = tuple(fem_row)  # Werte aus FEM-USED-MATERIALS speichern

        existing_entries = set()  # Set für Duplikate

        # Iteriere über alle Zeilen in Material_Database
        for db_row in database_isotropic.iter_rows(min_row=2, values_only=True):
            db_values = tuple(db_row[1:6])  # Spalten B bis F der Material_Database

            # Vergleiche FEM-Werte mit Material-Datenbank
            if fem_values == db_values:
                # Füge die gesamte Zeile aus der Material-Datenbank in die Ergebnisdatei ein
                new_row = isotropic_sheet.max_row + 1
                for col, value in enumerate(db_row, 1):
                    isotropic_sheet.cell(row=new_row, column=col).value = value

                # Markiere als orange, wenn es bereits eine Übereinstimmung gibt
                if fem_values in existing_entries:
                    for cell in isotropic_sheet[new_row]:
                        cell.fill = orange_fill
                else:
                    existing_entries.add(fem_values)

# Vergleiche die Zeilen in "Orthotropic Materials" (Spalten B-H)
def compare_orthotropic():
    # Iteriere über alle Zeilen in FEM-USED-MATERIALS (Orthotropic Materials)
    for fem_row in fem_used_orthotropic.iter_rows(min_row=2, min_col=2, max_col=8, values_only=True):  # B bis H
        fem_values = tuple(fem_row)  # Werte aus FEM-USED-MATERIALS speichern

        existing_entries = set()  # Set für Duplikate

        # Iteriere über alle Zeilen in Material_Database (Orthotropic Materials)
        for db_row in database_orthotropic.iter_rows(min_row=2, values_only=True):
            db_values = tuple(db_row[1:8])  # Spalten B bis H der Material_Database

            # Vergleiche FEM-Werte mit Material-Datenbank
            if fem_values == db_values:
                # Füge die gesamte Zeile aus der Material-Datenbank in die Ergebnisdatei ein
                new_row = orthotropic_sheet.max_row + 1
                for col, value in enumerate(db_row, 1):
                    orthotropic_sheet.cell(row=new_row, column=col).value = value

                # Markiere als orange, wenn es bereits eine Übereinstimmung gibt
                if fem_values in existing_entries:
                    for cell in orthotropic_sheet[new_row]:
                        cell.fill = orange_fill
                else:
                    existing_entries.add(fem_values)

# Vergleiche die Daten
compare_isotropic()
compare_orthotropic()

# Entferne das leere Standard-Sheet
del materials_for_srp["Sheet"]

# Spaltenbreite für beide Sheets anpassen
adjust_column_width(isotropic_sheet)  # Für Isotropic Materials
adjust_column_width(orthotropic_sheet)  # Für Orthotropic Materials

# Header und Spalte A fett machen
make_bold_headers(isotropic_sheet, isotropic_sheet.max_column)  # Für Isotropic Materials
make_bold_headers(orthotropic_sheet, orthotropic_sheet.max_column)  # Für Orthotropic Materials

# Rahmen für alle Zellen mit Einträgen hinzufügen
add_borders(isotropic_sheet)
add_borders(orthotropic_sheet)

# Speichere die neue Excel-Datei
output_path = os.path.join(os.getcwd(), "Materials_for_SRP.xlsx")
materials_for_srp.save(output_path)

print(f"Die Datei 'Materials_for_SRP.xlsx' wurde erfolgreich erstellt und gespeichert: {output_path}")

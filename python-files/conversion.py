
from openpyxl import Workbook
from pathlib import Path
import re, sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QFileDialog,
    QTableWidgetItem, QMessageBox, QVBoxLayout, QProgressBar
)
from ui_form import Ui_Form

_ILLEGAL_XML_CHARS_RE = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')

def encodage(s: str) -> str:
    return _ILLEGAL_XML_CHARS_RE.sub('', s)

def verif(element_actu):
    try:
        float(element_actu)
        return True
    except ValueError:
        return False
    
def message():
    msg = QMessageBox()
    msg.setWindowTitle("Terminé !")
    msg.setText(f"La conversion du fichier texte est terminé avec succès !")
    msg.exec_()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.ui.cancel_button.clicked.connect(self.close)
        self.ui.valider.clicked.connect(self.valider)
        self.ui.path_txt_button.clicked.connect(self.chemin_txt)
        self.ui.path_xlsx_button.clicked.connect(self.chemin_xlsx)

        self.chemin_txt_path = None
        self.chemin_dir_xlsx_path = None

    def chemin_xlsx(self):
        xlsx_path = QFileDialog.getExistingDirectory(
            parent=self,
            caption="Choisir un dossier de destination",
            directory="",
            options=QFileDialog.ShowDirsOnly
        )
        if xlsx_path:
            self.chemin_dir_xlsx_path = Path(xlsx_path)

    def chemin_txt(self):
        txt_path, _ = QFileDialog.getOpenFileName(
            parent=self,
            caption="Choisir un fichier",
            directory=".",
            filter="Fichiers texte (*.txt)"
        )
        if txt_path:
            self.chemin_txt_path = Path(txt_path)

    def valider(self):
        print("activer start")
        wb = Workbook()
        ws = wb.active
        chemin = Path(self.chemin_txt_path)
        liste_glob = []
        a = 0
        with open(chemin, "r") as fichier:
            for lignes in fichier:
                if a <= 6:
                    liste_glob.append(lignes.replace("\n", "").split(" : "))
                else:
                    liste_glob.append(lignes.replace("\n", "").split("\t"))
                a += 1
        b = 1
        i = 0
        for element in liste_glob:
            nb_colonne = len(element)
            while i < nb_colonne:
                element_actu = element[0+i]
                element_actu = encodage(element_actu)
                if verif(element_actu):
                    element_actu = float(element_actu)
                ws.cell(row=b, column=1 + i, value=element_actu)
                i += 1
            i = 0
            b += 1

        ligne_finale = len(liste_glob)
        ws[f"B{ligne_finale + 1}"] = f"=AVERAGE(B9:B{ligne_finale})"
        ws[f"C{ligne_finale + 1}"] = f"=AVERAGE(C9:C{ligne_finale})"
        ws[f"D{ligne_finale + 1}"] = f"=AVERAGE(D9:D{ligne_finale})"

        dest = self.chemin_dir_xlsx_path / f"{self.chemin_txt_path.stem}.xlsx"
        wb.save(dest)
        message()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
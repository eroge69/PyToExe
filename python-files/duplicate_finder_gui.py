import tkinter as tk
from tkinter import filedialog, ttk, scrolledtext, messagebox
import pandas as pd
import itertools
from collections import defaultdict
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Reuse functions from the original script
def read_excel_file(file_path, scan_all_sheets=False):
    """
    Read Excel file and return DataFrame(s)
    If scan_all_sheets is True, returns a dictionary of DataFrames, one for each sheet
    If scan_all_sheets is False, returns a single DataFrame from the first sheet
    """
    try:
        if scan_all_sheets:
            # Read all sheets into a dictionary of DataFrames
            sheet_dict = pd.read_excel(file_path, sheet_name=None)
            return sheet_dict, None
        else:
            # Read only the first sheet
            return pd.read_excel(file_path), None
    except Exception as e:
        return None, str(e)


def extract_combinations(row_data, min_size=1, max_size=5):
    """
    Extract all possible combinations of size min_size to max_size from row_data
    """
    all_combinations = []
    # Filter out non-numeric values and convert to integers where possible
    numeric_values = []
    for val in row_data:
        if pd.notna(val):  # Check if value is not NaN
            try:
                numeric_values.append(int(val))
            except (ValueError, TypeError):
                # Skip non-numeric values
                pass
    
    # Generate combinations of different sizes
    for size in range(min_size, min(max_size + 1, len(numeric_values) + 1)):
        for combo in itertools.combinations(numeric_values, size):
            # Sort to ensure consistent ordering
            all_combinations.append(tuple(sorted(combo)))
    
    return all_combinations


def find_duplicate_combinations(data, min_size=1, max_size=5, scan_all_sheets=False):
    """
    Find combinations that appear in multiple rows
    If scan_all_sheets is True, data should be a dictionary of DataFrames
    If scan_all_sheets is False, data should be a single DataFrame
    """
    # Dictionary to track combinations and the rows they appear in
    combination_rows = defaultdict(list)
    
    if scan_all_sheets:
        # Process each sheet
        for sheet_name, df in data.items():
            # Process each row in this sheet
            for row_idx, row in df.iterrows():
                # Extract combinations from this row
                row_combinations = extract_combinations(row.values, min_size, max_size)
                
                # Track which row each combination appears in, including sheet name
                for combo in row_combinations:
                    # row_idx + 2 to account for header and 1-based indexing
                    combination_rows[combo].append((sheet_name, row_idx + 2))
    else:
        # Process a single DataFrame
        df = data
        for row_idx, row in df.iterrows():
            # Extract combinations from this row
            row_combinations = extract_combinations(row.values, min_size, max_size)
            
            # Track which row each combination appears in
            for combo in row_combinations:
                combination_rows[combo].append(row_idx + 2)
    
    # Filter to only combinations that appear in multiple rows
    duplicate_combinations = {combo: rows for combo, rows in combination_rows.items() if len(rows) > 1}
    
    return duplicate_combinations


def format_results(duplicate_combinations, scan_all_sheets=False):
    """
    Format the results for display
    """
    if not duplicate_combinations:
        return "No duplicate combinations found."
    
    results = ["Duplicate combinations found:"]
    for combo, rows in sorted(duplicate_combinations.items(), key=lambda x: (len(x[0]), x[0])):
        combo_str = ", ".join(str(num) for num in combo)
        
        if scan_all_sheets:
            # Format with sheet names
            locations = []
            for sheet_name, row in rows:
                locations.append(f"Sheet '{sheet_name}' Row {row}")
            rows_str = ", ".join(locations)
            results.append(f"Combination [{combo_str}] appears in: {rows_str}")
        else:
            # Format with just row numbers
            rows_str = ", ".join(str(row) for row in rows)
            results.append(f"Combination [{combo_str}] appears in rows: {rows_str}")
    
    return "\n".join(results)


def write_results_to_excel(file_path, duplicate_combinations, output_file=None, scan_all_sheets=False):
    """
    Write results to a new sheet in the Excel file or create a new file
    """
    if output_file is None:
        # Create output filename based on input filename
        base_name, ext = os.path.splitext(file_path)
        output_file = f"{base_name}_results{ext}"
    
    # Create a DataFrame from the results
    results_data = []
    for combo, rows in duplicate_combinations.items():
        combo_str = ", ".join(str(num) for num in combo)
        
        if scan_all_sheets:
            # Format with sheet names
            locations = []
            sheet_row_map = {}
            
            for sheet_name, row in rows:
                locations.append(f"Sheet '{sheet_name}' Row {row}")
                
                # Group by sheet for easier reference
                if sheet_name not in sheet_row_map:
                    sheet_row_map[sheet_name] = []
                sheet_row_map[sheet_name].append(str(row))
            
            # Create a formatted string for each sheet
            sheet_details = []
            for sheet, sheet_rows in sheet_row_map.items():
                sheet_details.append(f"Sheet '{sheet}': Rows {', '.join(sheet_rows)}")
            
            sheet_rows_str = "; ".join(sheet_details)
            
            results_data.append({
                "Combination": combo_str,
                "Size": len(combo),
                "Appears in": sheet_rows_str,
                "Occurrence Count": len(rows)
            })
        else:
            # Format with just row numbers
            rows_str = ", ".join(str(row) for row in rows)
            results_data.append({
                "Combination": combo_str,
                "Size": len(combo),
                "Appears in Rows": rows_str,
                "Occurrence Count": len(rows)
            })
    
    # Create DataFrame and sort by combination size and occurrence count
    results_df = pd.DataFrame(results_data)
    if not results_df.empty:
        results_df = results_df.sort_values(by=["Size", "Occurrence Count"], ascending=[True, False])
    
    # Write to Excel
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Copy original data
            if scan_all_sheets:
                # Copy all sheets
                sheet_dict = pd.read_excel(file_path, sheet_name=None)
                for sheet_name, df in sheet_dict.items():
                    df.to_excel(writer, sheet_name=f'Original - {sheet_name}', index=False)
            else:
                # Copy just the first sheet
                original_df = pd.read_excel(file_path)
                original_df.to_excel(writer, sheet_name='Original Data', index=False)
            
            # Write results
            if not results_df.empty:
                results_df.to_excel(writer, sheet_name='Duplicate Combinations', index=False)
            else:
                pd.DataFrame({"Message": ["No duplicate combinations found."]}).to_excel(
                    writer, sheet_name='Duplicate Combinations', index=False)
        
        return True, output_file
    except Exception as e:
        return False, str(e)


def highlight_rows_in_excel(file_path, duplicate_combinations, output_file=None, scan_all_sheets=False):
    """
    Highlight rows in the original Excel file that contain duplicate combinations
    """
    if output_file is None:
        # Create output filename based on input filename
        base_name, ext = os.path.splitext(file_path)
        output_file = f"{base_name}_highlighted{ext}"
    
    try:
        # Copy the original workbook
        wb = load_workbook(file_path)
        
        # Create a fill pattern for highlighting
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        if scan_all_sheets:
            # Group rows to highlight by sheet
            sheet_rows = defaultdict(set)
            
            for rows in duplicate_combinations.values():
                for sheet_name, row_idx in rows:
                    sheet_rows[sheet_name].add(row_idx)
            
            # Highlight rows in each sheet
            for sheet_name, rows_to_highlight in sheet_rows.items():
                # Try to get the sheet by name
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Highlight the rows
                    for row_idx in rows_to_highlight:
                        for cell in ws[row_idx]:  # row_idx is already 1-based
                            cell.fill = highlight_fill
        else:
            # Highlight rows in the active sheet
            ws = wb.active
            
            # Get all rows that contain duplicate combinations
            rows_to_highlight = set()
            for rows in duplicate_combinations.values():
                rows_to_highlight.update(rows)
            
            # Highlight the rows (accounting for header row)
            for row_idx in rows_to_highlight:
                for cell in ws[row_idx]:  # row_idx is already 1-based
                    cell.fill = highlight_fill
        
        # Save the workbook
        wb.save(output_file)
        return True, output_file
    except Exception as e:
        return False, str(e)


class DuplicateFinderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Duplicate Combination Finder")
        self.root.geometry("800x600")
        self.root.minsize(800, 600)
        
        self.file_path = None
        self.df = None
        self.duplicate_combinations = None
        
        self.create_widgets()
        self.setup_layout()
    
    def create_widgets(self):
        # File selection frame
        self.file_frame = ttk.LabelFrame(self.root, text="File Selection")
        self.file_path_var = tk.StringVar()
        self.file_path_entry = ttk.Entry(self.file_frame, textvariable=self.file_path_var, width=50)
        self.browse_button = ttk.Button(self.file_frame, text="Browse", command=self.browse_file)
        
        # Options frame
        self.options_frame = ttk.LabelFrame(self.root, text="Options")
        
        # Min size option
        self.min_size_var = tk.IntVar(value=1)
        self.min_size_label = ttk.Label(self.options_frame, text="Minimum Combination Size:")
        self.min_size_spinbox = ttk.Spinbox(self.options_frame, from_=1, to=10, textvariable=self.min_size_var, width=5)
        
        # Max size option
        self.max_size_var = tk.IntVar(value=5)
        self.max_size_label = ttk.Label(self.options_frame, text="Maximum Combination Size:")
        self.max_size_spinbox = ttk.Spinbox(self.options_frame, from_=2, to=10, textvariable=self.max_size_var, width=5)
        
        # Scan mode option
        self.scan_all_sheets_var = tk.BooleanVar(value=False)
        self.scan_all_sheets_check = ttk.Checkbutton(
            self.options_frame, 
            text="Scan all sheets together", 
            variable=self.scan_all_sheets_var
        )
        
        # Process button
        self.process_button = ttk.Button(self.options_frame, text="Find Duplicates", command=self.process_file)
        
        # Results frame
        self.results_frame = ttk.LabelFrame(self.root, text="Results")
        self.results_text = scrolledtext.ScrolledText(self.results_frame, wrap=tk.WORD, width=80, height=20)
        self.results_text.config(state=tk.DISABLED)
        
        # Save options frame
        self.save_frame = ttk.LabelFrame(self.root, text="Save Options")
        
        # Save options
        self.save_results_var = tk.BooleanVar(value=True)
        self.save_results_check = ttk.Checkbutton(self.save_frame, text="Save Results to Excel", variable=self.save_results_var)
        
        self.highlight_rows_var = tk.BooleanVar(value=True)
        self.highlight_rows_check = ttk.Checkbutton(self.save_frame, text="Highlight Rows in Excel", variable=self.highlight_rows_var)
        
        # Save button
        self.save_button = ttk.Button(self.save_frame, text="Save", command=self.save_results)
        self.save_button.config(state=tk.DISABLED)
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
    
    def setup_layout(self):
        # File frame layout
        self.file_frame.pack(fill=tk.X, padx=10, pady=5)
        self.file_path_entry.pack(side=tk.LEFT, padx=5, pady=5, expand=True, fill=tk.X)
        self.browse_button.pack(side=tk.RIGHT, padx=5, pady=5)
        
        # Options frame layout
        self.options_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Create a grid for options
        self.min_size_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.min_size_spinbox.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        self.max_size_label.grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
        self.max_size_spinbox.grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)
        
        self.scan_all_sheets_check.grid(row=1, column=0, columnspan=4, padx=5, pady=5, sticky=tk.W)
        
        self.process_button.grid(row=0, column=4, rowspan=2, padx=20, pady=5)
        
        # Results frame layout
        self.results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.results_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Save frame layout
        self.save_frame.pack(fill=tk.X, padx=10, pady=5)
        self.save_results_check.pack(side=tk.LEFT, padx=5, pady=5)
        self.highlight_rows_check.pack(side=tk.LEFT, padx=20, pady=5)
        self.save_button.pack(side=tk.RIGHT, padx=5, pady=5)
        
        # Status bar layout
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Set initial status
        self.status_var.set("Ready. Please select an Excel file.")
        
        # Set initial results text
        self.update_results_text("No file processed yet. Please select an Excel file and click 'Find Duplicates'.")
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            self.file_path = file_path
            self.file_path_var.set(file_path)
            self.status_var.set(f"File selected: {os.path.basename(file_path)}")
    
    def update_results_text(self, text):
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, text)
        self.results_text.config(state=tk.DISABLED)
    
    def process_file(self):
        if not self.file_path:
            messagebox.showerror("Error", "Please select an Excel file first.")
            return
        
        if not os.path.exists(self.file_path):
            messagebox.showerror("Error", f"File '{self.file_path}' not found.")
            return
        
        # Read Excel file
        self.status_var.set("Reading Excel file...")
        self.root.update_idletasks()
        
        # Get scan mode
        scan_all_sheets = self.scan_all_sheets_var.get()
        
        self.df, error_msg = read_excel_file(self.file_path, scan_all_sheets=scan_all_sheets)
        if self.df is None:
            messagebox.showerror("Error", f"Error reading Excel file: {error_msg}")
            self.status_var.set("Error reading file.")
            return
        
        # Get min and max sizes
        min_size = self.min_size_var.get()
        max_size = self.max_size_var.get()
        
        # Get scan mode
        scan_all_sheets = self.scan_all_sheets_var.get()
        
        # Validate sizes
        if min_size > max_size:
            messagebox.showerror("Error", "Minimum size cannot be greater than maximum size.")
            return
        
        # Find duplicate combinations
        if scan_all_sheets:
            total_rows = sum(len(df) for df in self.df.values())
            self.status_var.set(f"Processing {len(self.df)} sheets with a total of {total_rows} rows...")
        else:
            self.status_var.set(f"Processing {len(self.df)} rows...")
        self.root.update_idletasks()
        
        self.duplicate_combinations = find_duplicate_combinations(self.df, min_size, max_size, scan_all_sheets=scan_all_sheets)
        
        # Display results
        results = format_results(self.duplicate_combinations, scan_all_sheets=scan_all_sheets)
        self.update_results_text(results)
        
        # Enable save button if duplicates found
        if self.duplicate_combinations:
            self.save_button.config(state=tk.NORMAL)
            self.status_var.set(f"Found {len(self.duplicate_combinations)} duplicate combinations.")
        else:
            self.save_button.config(state=tk.DISABLED)
            self.status_var.set("No duplicate combinations found.")
    
    def save_results(self):
        if not self.duplicate_combinations:
            messagebox.showinfo("Info", "No duplicate combinations to save.")
            return
        
        save_results = self.save_results_var.get()
        highlight_rows = self.highlight_rows_var.get()
        
        if not (save_results or highlight_rows):
            messagebox.showinfo("Info", "Please select at least one save option.")
            return
        
        success_messages = []
        error_messages = []
        
        # Get scan mode
        scan_all_sheets = self.scan_all_sheets_var.get()
        
        if save_results:
            self.status_var.set("Saving results to Excel...")
            self.root.update_idletasks()
            
            success, result = write_results_to_excel(self.file_path, self.duplicate_combinations, scan_all_sheets=scan_all_sheets)
            if success:
                success_messages.append(f"Results written to {os.path.basename(result)}")
            else:
                error_messages.append(f"Error writing to Excel: {result}")
        
        if highlight_rows:
            self.status_var.set("Highlighting rows in Excel...")
            self.root.update_idletasks()
            
            success, result = highlight_rows_in_excel(self.file_path, self.duplicate_combinations, scan_all_sheets=scan_all_sheets)
            if success:
                success_messages.append(f"Highlighted workbook saved to {os.path.basename(result)}")
            else:
                error_messages.append(f"Error highlighting Excel file: {result}")
        
        # Show results
        if success_messages:
            messagebox.showinfo("Success", "\n".join(success_messages))
        
        if error_messages:
            messagebox.showerror("Error", "\n".join(error_messages))
        
        self.status_var.set("Save operations completed.")


def main():
    root = tk.Tk()
    app = DuplicateFinderApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
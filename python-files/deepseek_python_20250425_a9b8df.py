import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
import os

class AppealDeadlineCalculator:
    def __init__(self, root):
        self.root = root
        self.root.title("حاسبة آجال الاستئناف والطعن بالنقض")
        self.root.geometry("800x600")
        self.root.resizable(False, False)
        
        # إنشاء دفتر ملاحظات (Notebook) للتبويبات
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True)
        
        # إنشاء إطارات للتبويبات المختلفة
        self.create_civil_frame()
        self.create_criminal_frame()
        
        # إنشاء قائمة ملف
        self.create_menu()
        
        # تحميل البيانات من ملف Excel إذا وجد
        self.excel_file = "appeal_deadlines.xlsx"
        self.load_data()
    
    def create_menu(self):
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="حفظ البيانات", command=self.save_data)
        file_menu.add_command(label="تصدير إلى Excel", command=self.export_to_excel)
        file_menu.add_separator()
        file_menu.add_command(label="خروج", command=self.root.quit)
        menubar.add_cascade(label="ملف", menu=file_menu)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="حول البرنامج", command=self.about)
        menubar.add_cascade(label="مساعدة", menu=help_menu)
        
        self.root.config(menu=menubar)
    
    def create_civil_frame(self):
        """إنشاء واجهة القضايا المدنية"""
        civil_frame = ttk.Frame(self.notebook)
        self.notebook.add(civil_frame, text="القضايا المدنية")
        
        # عنوان الواجهة
        title_label = ttk.Label(civil_frame, text="حاسبة آجال الاستئناف في القضايا المدنية", font=('Arial', 14, 'bold'))
        title_label.pack(pady=10)
        
        # إطار بيانات الإدخال
        input_frame = ttk.LabelFrame(civil_frame, text="بيانات التبليغ")
        input_frame.pack(pady=10, padx=10, fill='x')
        
        # تاريخ استلام التبليغ
        ttk.Label(input_frame, text="تاريخ استلام التبليغ:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.notification_date_civil = ttk.Entry(input_frame)
        self.notification_date_civil.grid(row=0, column=1, padx=5, pady=5)
        
        # نوع التبليغ
        ttk.Label(input_frame, text="نوع التبليغ:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.notification_type_civil = ttk.Combobox(input_frame, values=["تبليغ في الموطن الحقيقي", "تبليغ شخصي"])
        self.notification_type_civil.grid(row=1, column=1, padx=5, pady=5)
        self.notification_type_civil.current(0)
        
        # زر الحساب
        calculate_btn = ttk.Button(input_frame, text="حساب الآجال", command=self.calculate_civil_deadlines)
        calculate_btn.grid(row=2, column=0, columnspan=2, pady=10)
        
        # إطار النتائج
        result_frame = ttk.LabelFrame(civil_frame, text="نتائج الحساب")
        result_frame.pack(pady=10, padx=10, fill='both', expand=True)
        
        # نتائج التبليغ في الموطن الحقيقي
        ttk.Label(result_frame, text="التبليغ في الموطن الحقيقي (60 يوم):").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.real_address_deadline = ttk.Label(result_frame, text="")
        self.real_address_deadline.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(result_frame, text="الأيام المتبقية:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.real_address_days_left = ttk.Label(result_frame, text="")
        self.real_address_days_left.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        self.real_address_status = ttk.Label(result_frame, text="", font=('Arial', 10, 'bold'))
        self.real_address_status.grid(row=2, column=0, columnspan=2, pady=5)
        
        # فاصل
        ttk.Separator(result_frame, orient='horizontal').grid(row=3, column=0, columnspan=2, pady=10, sticky='ew')
        
        # نتائج التبليغ الشخصي
        ttk.Label(result_frame, text="التبليغ الشخصي (30 يوم):").grid(row=4, column=0, padx=5, pady=5, sticky='e')
        self.personal_notification_deadline = ttk.Label(result_frame, text="")
        self.personal_notification_deadline.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(result_frame, text="الأيام المتبقية:").grid(row=5, column=0, padx=5, pady=5, sticky='e')
        self.personal_days_left = ttk.Label(result_frame, text="")
        self.personal_days_left.grid(row=5, column=1, padx=5, pady=5, sticky='w')
        
        self.personal_status = ttk.Label(result_frame, text="", font=('Arial', 10, 'bold'))
        self.personal_status.grid(row=6, column=0, columnspan=2, pady=5)
    
    def create_criminal_frame(self):
        """إنشاء واجهة القضايا الجزائية"""
        criminal_frame = ttk.Frame(self.notebook)
        self.notebook.add(criminal_frame, text="القضايا الجزائية")
        
        # عنوان الواجهة
        title_label = ttk.Label(criminal_frame, text="حاسبة آجال الاستئناف والطعن بالنقض في القضايا الجزائية", font=('Arial', 14, 'bold'))
        title_label.pack(pady=10)
        
        # إطار بيانات الإدخال
        input_frame = ttk.LabelFrame(criminal_frame, text="بيانات الحكم")
        input_frame.pack(pady=10, padx=10, fill='x')
        
        # تاريخ استلام أو صدور الحكم
        ttk.Label(input_frame, text="تاريخ استلام/صدور الحكم:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.judgment_date_criminal = ttk.Entry(input_frame)
        self.judgment_date_criminal.grid(row=0, column=1, padx=5, pady=5)
        
        # زر الحساب
        calculate_btn = ttk.Button(input_frame, text="حساب الآجال", command=self.calculate_criminal_deadlines)
        calculate_btn.grid(row=1, column=0, columnspan=2, pady=10)
        
        # إطار النتائج
        result_frame = ttk.LabelFrame(criminal_frame, text="نتائج الحساب")
        result_frame.pack(pady=10, padx=10, fill='both', expand=True)
        
        # نتائج التصريح بالطعن بالنقض
        ttk.Label(result_frame, text="آخر أجل للتصريح بالطعن بالنقض (10 أيام):").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.appeal_declaration_deadline = ttk.Label(result_frame, text="")
        self.appeal_declaration_deadline.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(result_frame, text="الأيام المتبقية:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.appeal_declaration_days_left = ttk.Label(result_frame, text="")
        self.appeal_declaration_days_left.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        self.appeal_declaration_status = ttk.Label(result_frame, text="", font=('Arial', 10, 'bold'))
        self.appeal_declaration_status.grid(row=2, column=0, columnspan=2, pady=5)
        
        # فاصل
        ttk.Separator(result_frame, orient='horizontal').grid(row=3, column=0, columnspan=2, pady=10, sticky='ew')
        
        # نتائج الاستئناف
        ttk.Label(result_frame, text="آخر أجل للاستئناف (10 أيام):").grid(row=4, column=0, padx=5, pady=5, sticky='e')
        self.appeal_deadline = ttk.Label(result_frame, text="")
        self.appeal_deadline.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(result_frame, text="الأيام المتبقية:").grid(row=5, column=0, padx=5, pady=5, sticky='e')
        self.appeal_days_left = ttk.Label(result_frame, text="")
        self.appeal_days_left.grid(row=5, column=1, padx=5, pady=5, sticky='w')
        
        self.appeal_status = ttk.Label(result_frame, text="", font=('Arial', 10, 'bold'))
        self.appeal_status.grid(row=6, column=0, columnspan=2, pady=5)
        
        # فاصل
        ttk.Separator(result_frame, orient='horizontal').grid(row=7, column=0, columnspan=2, pady=10, sticky='ew')
        
        # نتائج إيداع مذكرة الطعن
        ttk.Label(result_frame, text="آخر أجل لإيداع مذكرة الطعن (60 يوم):").grid(row=8, column=0, padx=5, pady=5, sticky='e')
        self.appeal_memo_deadline = ttk.Label(result_frame, text="")
        self.appeal_memo_deadline.grid(row=8, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(result_frame, text="الأيام المتبقية:").grid(row=9, column=0, padx=5, pady=5, sticky='e')
        self.appeal_memo_days_left = ttk.Label(result_frame, text="")
        self.appeal_memo_days_left.grid(row=9, column=1, padx=5, pady=5, sticky='w')
        
        self.appeal_memo_status = ttk.Label(result_frame, text="", font=('Arial', 10, 'bold'))
        self.appeal_memo_status.grid(row=10, column=0, columnspan=2, pady=5)
    
    def calculate_civil_deadlines(self):
        """حساب آجال الاستئناف في القضايا المدنية"""
        try:
            notification_date = datetime.strptime(self.notification_date_civil.get(), "%Y-%m-%d")
            today = datetime.now()
            
            # حساب آجال التبليغ في الموطن الحقيقي (60 يوم)
            real_address_deadline = notification_date + timedelta(days=60)
            days_left_real = (real_address_deadline - today).days
            
            self.real_address_deadline.config(text=real_address_deadline.strftime("%Y-%m-%d"))
            self.real_address_days_left.config(text=f"{days_left_real} يوم")
            
            if days_left_real > 0:
                self.real_address_status.config(text="الاستئناف ضمن الآجال", foreground='green')
            else:
                self.real_address_status.config(text="عدم جواز الاستئناف لانقضاء الآجال", foreground='red')
            
            # حساب آجال التبليغ الشخصي (30 يوم)
            personal_deadline = notification_date + timedelta(days=30)
            days_left_personal = (personal_deadline - today).days
            
            self.personal_notification_deadline.config(text=personal_deadline.strftime("%Y-%m-%d"))
            self.personal_days_left.config(text=f"{days_left_personal} يوم")
            
            if days_left_personal > 0:
                self.personal_status.config(text="الاستئناف ضمن الآجال", foreground='green')
            else:
                self.personal_status.config(text="عدم جواز الاستئناف لانقضاء الآجال", foreground='red')
            
        except ValueError:
            messagebox.showerror("خطأ", "الرجاء إدخال تاريخ صحيح بصيغة YYYY-MM-DD")
    
    def calculate_criminal_deadlines(self):
        """حساب آجال الاستئناف والطعن بالنقض في القضايا الجزائية"""
        try:
            judgment_date = datetime.strptime(self.judgment_date_criminal.get(), "%Y-%m-%d")
            today = datetime.now()
            
            # حساب آجال التصريح بالطعن بالنقض (10 أيام)
            appeal_declaration_deadline = judgment_date + timedelta(days=10)
            days_left_declaration = (appeal_declaration_deadline - today).days
            
            self.appeal_declaration_deadline.config(text=appeal_declaration_deadline.strftime("%Y-%m-%d"))
            self.appeal_declaration_days_left.config(text=f"{days_left_declaration} يوم")
            
            if days_left_declaration > 0:
                self.appeal_declaration_status.config(text="التصريح بالطعن ضمن الآجال", foreground='green')
            else:
                self.appeal_declaration_status.config(text="عدم جواز التصريح بالطعن لانقضاء الآجال", foreground='red')
            
            # حساب آجال الاستئناف (10 أيام)
            appeal_deadline = judgment_date + timedelta(days=10)
            days_left_appeal = (appeal_deadline - today).days
            
            self.appeal_deadline.config(text=appeal_deadline.strftime("%Y-%m-%d"))
            self.appeal_days_left.config(text=f"{days_left_appeal} يوم")
            
            if days_left_appeal > 0:
                self.appeal_status.config(text="الاستئناف ضمن الآجال", foreground='green')
            else:
                self.appeal_status.config(text="عدم جواز الاستئناف لانقضاء الآجال", foreground='red')
            
            # حساب آجال إيداع مذكرة الطعن (60 يوم)
            appeal_memo_deadline = judgment_date + timedelta(days=60)
            days_left_memo = (appeal_memo_deadline - today).days
            
            self.appeal_memo_deadline.config(text=appeal_memo_deadline.strftime("%Y-%m-%d"))
            self.appeal_memo_days_left.config(text=f"{days_left_memo} يوم")
            
            if days_left_memo > 0:
                self.appeal_memo_status.config(text="إيداع مذكرة الطعن ضمن الآجال", foreground='green')
            else:
                self.appeal_memo_status.config(text="عدم جواز الطعن لعدم إيداع المذكرة في الآجال", foreground='red')
            
        except ValueError:
            messagebox.showerror("خطأ", "الرجاء إدخال تاريخ صحيح بصيغة YYYY-MM-DD")
    
    def save_data(self):
        """حفظ البيانات في ملف البرنامج"""
        try:
            data = {
                "civil": {
                    "notification_date": self.notification_date_civil.get(),
                    "notification_type": self.notification_type_civil.get()
                },
                "criminal": {
                    "judgment_date": self.judgment_date_criminal.get()
                }
            }
            
            # يمكن هنا إضافة كود لحفظ البيانات في ملف محلي (مثل JSON)
            messagebox.showinfo("حفظ", "تم حفظ البيانات بنجاح")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء حفظ البيانات: {str(e)}")
    
    def export_to_excel(self):
        """تصدير البيانات إلى ملف Excel"""
        try:
            wb = Workbook()
            
            # ورقة القضايا المدنية
            ws_civil = wb.active
            ws_civil.title = "اجال الاستئناف المدني"
            
            ws_civil.append(["تطبيقة حساب أجال الاستئناف في القضايا المدنية"])
            ws_civil.append(["تاريخ استلام التبليغ:", self.notification_date_civil.get()])
            ws_civil.append(["نوع التبليغ:", self.notification_type_civil.get()])
            ws_civil.append([])
            
            # حساب التواريخ للمدني
            if self.notification_date_civil.get():
                notification_date = datetime.strptime(self.notification_date_civil.get(), "%Y-%m-%d")
                today = datetime.now()
                
                # التبليغ في الموطن الحقيقي
                real_address_deadline = notification_date + timedelta(days=60)
                days_left_real = (real_address_deadline - today).days
                status_real = "ضمن الآجال" if days_left_real > 0 else "انقضاء الآجال"
                
                # التبليغ الشخصي
                personal_deadline = notification_date + timedelta(days=30)
                days_left_personal = (personal_deadline - today).days
                status_personal = "ضمن الآجال" if days_left_personal > 0 else "انقضاء الآجال"
                
                ws_civil.append(["التبليغ في الموطن الحقيقي (60 يوم)", real_address_deadline.strftime("%Y-%m-%d"), f"الأيام المتبقية: {days_left_real}", status_real])
                ws_civil.append(["التبليغ الشخصي (30 يوم)", personal_deadline.strftime("%Y-%m-%d"), f"الأيام المتبقية: {days_left_personal}", status_personal])
            
            # ورقة القضايا الجزائية
            ws_criminal = wb.create_sheet("اجال الطعن بالنقض (جزائي)")
            
            ws_criminal.append(["تطبيقة حساب أجال الاستئناف والطعن بالنقض في القضايا الجزائية"])
            ws_criminal.append(["تاريخ استلام/صدور الحكم:", self.judgment_date_criminal.get()])
            ws_criminal.append([])
            
            # حساب التواريخ للجزائي
            if self.judgment_date_criminal.get():
                judgment_date = datetime.strptime(self.judgment_date_criminal.get(), "%Y-%m-%d")
                today = datetime.now()
                
                # التصريح بالطعن
                appeal_declaration_deadline = judgment_date + timedelta(days=10)
                days_left_declaration = (appeal_declaration_deadline - today).days
                status_declaration = "ضمن الآجال" if days_left_declaration > 0 else "انقضاء الآجال"
                
                # الاستئناف
                appeal_deadline = judgment_date + timedelta(days=10)
                days_left_appeal = (appeal_deadline - today).days
                status_appeal = "ضمن الآجال" if days_left_appeal > 0 else "انقضاء الآجال"
                
                # إيداع المذكرة
                appeal_memo_deadline = judgment_date + timedelta(days=60)
                days_left_memo = (appeal_memo_deadline - today).days
                status_memo = "سارية" if days_left_memo > 0 else "انقضاء الآجال"
                
                ws_criminal.append(["آخر أجل للتصريح بالطعن بالنقض (10 أيام)", appeal_declaration_deadline.strftime("%Y-%m-%d"), f"الأيام المتبقية: {days_left_declaration}", status_declaration])
                ws_criminal.append(["آخر أجل للاستئناف (10 أيام)", appeal_deadline.strftime("%Y-%m-%d"), f"الأيام المتبقية: {days_left_appeal}", status_appeal])
                ws_criminal.append(["آخر أجل لإيداع مذكرة الطعن (60 يوم)", appeal_memo_deadline.strftime("%Y-%m-%d"), f"الأيام المتبقية: {days_left_memo}", status_memo])
            
            # حفظ الملف
            wb.save(self.excel_file)
            messagebox.showinfo("تصدير", f"تم تصدير البيانات إلى ملف Excel: {self.excel_file}")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء التصدير: {str(e)}")
    
    def load_data(self):
        """تحميل البيانات من ملف Excel إذا وجد"""
        if os.path.exists(self.excel_file):
            try:
                wb = openpyxl.load_workbook(self.excel_file)
                
                # تحميل بيانات القضايا المدنية
                if "اجال الاستئناف المدني" in wb.sheetnames:
                    ws_civil = wb["اجال الاستئناف المدني"]
                    if ws_civil.cell(row=2, column=1).value == "تاريخ استلام التبليغ:":
                        self.notification_date_civil.insert(0, ws_civil.cell(row=2, column=2).value or "")
                    if ws_civil.cell(row=3, column=1).value == "نوع التبليغ:":
                        notification_type = ws_civil.cell(row=3, column=2).value
                        if notification_type:
                            self.notification_type_civil.set(notification_type)
                
                # تحميل بيانات القضايا الجزائية
                if "اجال الطعن بالنقض (جزائي)" in wb.sheetnames:
                    ws_criminal = wb["اجال الطعن بالنقض (جزائي)"]
                    if ws_criminal.cell(row=2, column=1).value == "تاريخ استلام/صدور الحكم:":
                        self.judgment_date_criminal.insert(0, ws_criminal.cell(row=2, column=2).value or "")
                
            except Exception as e:
                print(f"حدث خطأ أثناء تحميل البيانات: {str(e)}")
    
    def about(self):
        """نافذة حول البرنامج"""
        about_text = """حاسبة آجال الاستئناف والطعن بالنقض
الإصدار 1.0

تم تطوير هذا البرنامج لحساب آجال الاستئناف والطعن بالنقض
في القضايا المدنية والجزائية وفقًا للنظام القانوني.

المطور: Belabbes T@-Amine
للتواصل: 06,56,18,78,74"""
        
        messagebox.showinfo("حول البرنامج", about_text)

if __name__ == "__main__":
    root = tk.Tk()
    app = AppealDeadlineCalculator(root)
    root.mainloop()
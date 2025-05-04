
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def safe_str(val):
    return "" if pd.isna(val) else str(val).strip()

def compara_si_aplica_reguli(fisier2024, fisier2025):
    df2024 = pd.read_excel(fisier2024)
    df2025 = pd.read_excel(fisier2025)

    df2024["ID"] = df2024["ID"].astype(str)
    df2025["ID"] = df2025["ID"].astype(str)

    df = pd.merge(df2024, df2025, on="ID", how="outer", suffixes=("_2024", "_2025"))

    def regula_1(row):
        cat_2024 = safe_str(row.get("cat_use_2024")).upper()
        cat_2025 = safe_str(row.get("cat_use_2025")).upper()
        eco_2024 = safe_str(row.get("Eco_2024")).lower()
        eco_2025 = safe_str(row.get("Eco_2025")).lower()
        if cat_2024 in ["PP", "PPN"] or cat_2025 in ["PP", "PPN"]:
            if eco_2024 == "conv 1" and eco_2025 == "conv 2":
                return "CORECT"
            else:
                return "INCORECT"
        return ""

    def regula_2(row):
        agro_2024 = safe_str(row.get("agro_env_2024")).upper()
        ang_2024 = safe_str(row.get("Ang_2024")).lower()
        ang_2025 = safe_str(row.get("Ang_2025")).lower()
        if agro_2024 == "B22":
            if ang_2024 == "ang 1" and (ang_2025 == "ang 2" or ang_2025 == ""):
                return "CORECT"
            else:
                return "INCORECT"
        return ""

    def regula_3(row):
        agro_2024 = safe_str(row.get("agro_env_2024")).upper()
        agro_2025 = safe_str(row.get("agro_env_2025")).upper()
        ang_2025 = safe_str(row.get("Ang_2025")).lower()
        if agro_2025 == "B22" and agro_2024 != "B22":
            if ang_2025 == "ang 1" or ang_2025 == "":
                return "CORECT"
            else:
                return "INCORECT"
        return ""

    def regula_4(row):
        id_2024_parcela = row.get("parcel_nr_2024")
        id_2025_parcela = row.get("parcel_nr_2025")
        eco_2025 = safe_str(row.get("Eco_2025")).lower()
        if pd.isna(id_2024_parcela) and not pd.isna(id_2025_parcela):
            if eco_2025 == "conv 1":
                return "CORECT"
            else:
                return "INCORECT"
        return ""

    df["Regula 1"] = df.apply(regula_1, axis=1)
    df["Regula 2"] = df.apply(regula_2, axis=1)
    df["Regula 3"] = df.apply(regula_3, axis=1)
    df["Regula 4"] = df.apply(regula_4, axis=1)

    out_file = "Comparativ_Colorat.xlsx"
    df.to_excel(out_file, index=False)

    wb = load_workbook(out_file)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col_index = {name: idx + 1 for idx, name in enumerate(header)}

    perechi = [
        ("parcel_nr_2024", "parcel_nr_2025"),
        ("bloc_nr_2024", "bloc_nr_2025"),
        ("cat_use_2024", "cat_use_2025"),
        ("crop_code_2024", "crop_code_2025"),
        ("crop_name_2024", "crop_name_2025"),
        ("area_dec_2024", "area_dec_2025"),
        ("agro_env_2024", "agro_env_2025"),
        ("Eco_2024", "Eco_2025"),
        ("Ang_2024", "Ang_2025")
    ]

    fill_diff = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_incorect = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        for col_2024, col_2025 in perechi:
            val_2024 = ws.cell(row=row, column=col_index[col_2024]).value
            val_2025 = ws.cell(row=row, column=col_index[col_2025]).value
            if str(val_2024).strip() != str(val_2025).strip():
                ws.cell(row=row, column=col_index[col_2025]).fill = fill_diff

    for row in range(2, ws.max_row + 1):
        for col_name in ["Regula 1", "Regula 2", "Regula 3", "Regula 4"]:
            val = ws.cell(row=row, column=col_index[col_name]).value
            if val == "INCORECT":
                ws.cell(row=row, column=col_index[col_name]).fill = fill_incorect

    wb.save(out_file)
    messagebox.showinfo("Gata", "Comparativul a fost salvat ca Comparativ_Colorat.xlsx")

def select_files():
    file_2024 = filedialog.askopenfilename(title="Selectează fișierul 2024", filetypes=[("Excel files", "*.xlsx")])
    file_2025 = filedialog.askopenfilename(title="Selectează fișierul 2025", filetypes=[("Excel files", "*.xlsx")])

    if file_2024 and file_2025:
        compara_si_aplica_reguli(file_2024, file_2025)

root = tk.Tk()
root.title("Comparativ APIA 2024 vs 2025")
root.geometry("400x200")

btn = tk.Button(root, text="Încarcă 2024 și 2025 și Compară", command=select_files, font=("Arial", 14))
btn.pack(expand=True)

root.mainloop()

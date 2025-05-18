import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import xml.etree.ElementTree as ET
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
ALL_CLASSES = ['ACBPR',
'ALD',
'ALD_R',
'AMGR',
'AMLEPR',
'ANBIOT',
'ANBIOTCARR',
'ANR',
'ANRPRL',
'ANRPRW',
'ANTL',
'ANTL_R',
'AOM',
'APEQM',
'APEQM_R',
'APUCCH_FDD',
'AgentDiscovery',
'BBADM',
'BBMOD',
'BBMOD_R',
'BBPOOL',
'BRGPRT',
'BWP_PROFILE',
'BtsState',
'CABINET',
'CABINET_R',
'CABLINK',
'CABLINK_R',
'CACERT_R',
'CADPR',
'CAGENB',
'CAPCFG',
'CAPLIM',
'CAPR',
'CAREL',
'CELLMAPPING',
'CERTH',
'CERTHENT',
'CHANNEL',
'CHANNELGROUP',
'CLOCK',
'CLOCK_R',
'CMD',
'CMP',
'CMPECDSA',
'CMPECDSA_R',
'CMPFH',
'CMPFHEXT',
'CMP_R',
'CONNECTOR_R',
'CREL',
'CRLH',
'CRLH_R',
'CTRLTS',
'CUPLANEFHNW',
'DNS',
'DRX',
'DSCP2PCPMAP',
'DSCP2QMAP',
'DSCPTOQMAP',
'EAC_IN',
'EAC_OUT',
'EAC_R',
'ECPRIMASTER',
'ECPRIMASTERSTATUS_R',
'ENDCDMEASCONF',
'EQM',
'EQM_R',
'ETHAPP',
'ETHIF',
'ETHIF_R',
'ETHLK',
'ETHLK_R',
'ETHMIRROR',
'ETHSVC',
'EXCCU',
'EXCENBF',
'EXENBF',
'EXEUCE',
'EXUCE',
'FAN_R',
'FEATCADM',
'FEATLADM',
'FIREWALL',
'FMCADM',
'FSTSCH',
'GFIM',
'GNFL',
'GNSSE',
'GNSSE_R',
'GTPU',
'HW',
'HWTOP',
'HWTOP_R',
'IBRGPRT',
'IFGDPR',
'IFGPR',
'INVUNIT',
'IPADDRESSV4',
'IPADDRESSV4_R',
'IPAPP',
'IPIF',
'IPIF_R',
'IPNO',
'IPRT',
'IPRTV6',
'IPRTV6_R',
'IPRT_R',
'IPSECC',
'IPSECC_R',
'IRFIM',
'ISHPR',
'L2SWI',
'LANE',
'LANE_R',
'LBPUCCHRDPR',
'LCELL',
'LCELNR',
'LNADJ',
'LNADJGNB',
'LNADJL',
'LNADJW',
'LNBTS',
'LNBTS_FDD',
'LNCEL',
'LNCEL_FDD',
'LNHOG',
'LNHOIF',
'LNHOW',
'LNHOX',
'LNMME',
'LNREL',
'LNRELGNBCELL',
'LNRELW',
'LOAM',
'LOAMPRF',
'LOAM_R',
'LOGLINK',
'LOGLINK_R',
'LTAC',
'LTEENB',
'LTRACE',
'LUAC',
'LUAC_R',
'MNL',
'MNLENT',
'MNL_R',
'MPLANENW',
'MPLANENW_R',
'MPUCCH_FDD',
'MRBTS',
'MRBTSDESC',
'MTRACE',
'NBIOTPR',
'NBIOT_FDD',
'NECERT_R',
'NETACT',
'NOTE',
'NOTES',
'NRADJNRCELL',
'NRANR',
'NRANRPR',
'NRBTS',
'NRCELL',
'NRCELLGRP',
'NRCTRLTS',
'NRCUUP',
'NRDCDPR',
'NRDLMUMIMO',
'NRDRB',
'NRDRB_5QI',
'NRDRB_MAC',
'NRDRB_PDCP',
'NRDRB_QCI',
'NRDRB_RLC_AM',
'NRDRB_RLC_UM',
'NRDRX',
'NRDU',
'NRMEASDPR',
'NRMTRACECU',
'NRPGRP',
'NRPLMN',
'NRPMCCP',
'NRPMRNL',
'NRREL',
'NRRFSP_PROFILE',
'NRSCTP',
'NRSPID_PROFILE',
'NRSYSINFO_PROFILE',
'NRSYSINFO_PROFILE_NSA',
'NRULCLPC_PROFILE',
'NRULMUMIMO',
'NRX2LINK_TRUST',
'NTP',
'NTP_R',
'NeIdentityData',
'NeIntegrationData',
'PCP2QMAP',
'PDCCH',
'PDCCH_CONFIG_COMMON',
'PDCCH_CONFIG_DEDICATED',
'PDSCH',
'PFOE_R',
'PHYANT',
'PHYANTU_R',
'PHYANT_R',
'PKTFLTR',
'PLMN',
'PMCADM',
'PMCADM_R',
'PMCCP',
'PMMNL',
'PMPLM',
'PMRNL',
'PMRPQH',
'PMTNL',
'PMTNLINT',
'PM_Period',
'PM_Schedule',
'PM_ScheduleItem',
'POWERGROUP_R',
'PSGRP',
'PTPMASTER',
'QOS',
'REDRT',
'RETU',
'RETU_R',
'RFREQ',
'RFRES',
'RFTHLD',
'RIM',
'RMOD',
'RMOD_R',
'RSL',
'RSL_R',
'RTPOL',
'SCTP',
'SDRX',
'SECADM',
'SECADM_R',
'SFP',
'SFP_R',
'SIB',
'SMOD',
'SMOD_R',
'SWRESETREQ',
'SWRESETRES',
'SYNC',
'SYNCE',
'SYNCE_R',
'TAC',
'TCE',
'TCEADM',
'TEST',
'TIME',
'TNL',
'TNLSVC',
'TNL_R',
'TOP',
'TOPF',
'TOPF_R',
'TOPP',
'TOPP_R',
'TRACKINGAREA',
'TRBLCADM',
'TRBLCADM_R',
'TRMOD_R',
'TRSNW',
'TWAMP',
'TWAMPREFLECT',
'UFFIM',
'ULCOMP',
'VLANIF',

]  # <--- Dán danh sách class vào đây, như bạn đã cung cấp

def get_tag_name(tag):
    return tag.split('}')[-1] if '}' in tag else tag

def extract_mrbts_id(dist_name):
    match = re.search(r'MRBTS-(\d+)', dist_name)
    return match.group(1) if match else ''

def extract_and_write_to_excel(xml_file_path, output_excel_path, selected_classes):
    context = ET.iterparse(xml_file_path, events=('start', 'end'))
    _, root = next(context)

    class_data = defaultdict(list)
    headers_by_class = defaultdict(set)

    for event, elem in context:
        if event == 'end' and get_tag_name(elem.tag) == 'managedObject':
            mo_class = elem.attrib.get('class')
            if mo_class not in selected_classes:
                elem.clear()
                continue

            dist_name = elem.attrib.get('distName')
            mo_id = elem.attrib.get('id', "")
            operation = elem.attrib.get('operation', "")

            mrbts_id = extract_mrbts_id(dist_name)

            base_info = {'MRBTS': mrbts_id, 'distName': dist_name, 'id': mo_id, 'operation': operation}
            flat_fields = {}
            list_fields = {}

            for child in elem:
                tag = get_tag_name(child.tag)
                if tag == 'p':
                    name = child.attrib.get('name')
                    flat_fields[name] = child.text
                    headers_by_class[mo_class].add(name)
                elif tag == 'list':
                    list_name = child.attrib.get('name', '')
                    headers_by_class[mo_class].add(list_name)
                    flat_fields[list_name] = "List"
                    contains_item = any(get_tag_name(e.tag) == 'item' for e in child)
                    if contains_item:
                        for item in child:
                            if get_tag_name(item.tag) != 'item':
                                continue
                            for p in item:
                                if get_tag_name(p.tag) != 'p':
                                    continue
                                pname = p.attrib.get('name')
                                key = f"item-{list_name}-{pname}" if pname else f"item-{list_name}"
                                value = p.text.strip() if p.text else ""
                                if key not in list_fields:
                                    list_fields[key] = []
                                list_fields[key].append(value)
                                headers_by_class[mo_class].add(key)
                    else:
                        for i, p in enumerate(child):
                            if get_tag_name(p.tag) != 'p':
                                continue
                            value = p.text.strip() if p.text else ""
                            key = f"{list_name}_{i+1}" if len(child) > 1 else list_name
                            flat_fields[key] = value
                            headers_by_class[mo_class].add(key)

            row = base_info.copy()
            row.update(flat_fields)
            for key, values in list_fields.items():
                row[key] = ";".join(values)
            class_data[mo_class].append(row)
            elem.clear()

    wb = Workbook()
    menu_ws = wb.active
    menu_ws.title = "Menu"

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    header_fill = PatternFill(start_color="3399FF", end_color="3399FF", fill_type="solid")
    header_align = Alignment(text_rotation=90, vertical='center')
    data_font = Font(name="Arial", size=9, color="000000")
    data_align = Alignment(horizontal='left')

    for class_name, rows in class_data.items():
        ws = wb.create_sheet(title=class_name[:31])
        headers = ['MRBTS', 'distName', 'id', 'operation'] + sorted(
            h for h in headers_by_class[class_name] if h not in ['MRBTS', 'distName', 'id', 'operation']
        )

        ws['A1'] = 'Menu'
        ws['A1'].hyperlink = "#Menu!A1"
        ws['A1'].style = "Hyperlink"
        ws['A1'].font = data_font

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for r_idx, row_data in enumerate(rows, start=3):
            for c_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                try:
                    if "." in value:
                        value = float(value)
                    else:
                        value = int(value)
                except:
                    pass
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = data_font
                cell.alignment = data_align

        r = menu_ws.max_row + 1
        menu_ws.cell(row=r, column=1, value=class_name)
        menu_ws.cell(row=r, column=1).hyperlink = f"#{class_name}!A2"
        menu_ws.cell(row=r, column=1).style = "Hyperlink"
        menu_ws.cell(row=r, column=1).font = data_font

    menu_ws['A1'] = 'Danh sách các class (Sheets)'
    menu_ws['A1'].font = header_font
    menu_ws['A1'].fill = header_fill
    menu_ws['A1'].alignment = Alignment(horizontal='center')

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(output_excel_path)

import tkinter.ttk as ttk  # cần thêm nếu chưa có

def run_gui():
    def move_selected(source, target):
        items = list(source.curselection())
        for i in reversed(items):
            val = source.get(i)
            if val not in target.get(0, tk.END):
                target.insert(tk.END, val)
            source.delete(i)

    def move_all(source, target):
        for val in source.get(0, tk.END):
            if val not in target.get(0, tk.END):
                target.insert(tk.END, val)
        source.delete(0, tk.END)

    def select_input_file():
        path = filedialog.askopenfilename(filetypes=[("XML files", "*.xml")])
        if path:
            input_path.set(path)

    def select_output_file():
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if path:
            output_path.set(path)

    def update_progress(percent):
        progress_bar["value"] = percent
        root.update_idletasks()

    def convert_file():
        in_file = input_path.get()
        out_file = output_path.get()
        selected_classes = list(class_selected.get(0, tk.END))
        if not in_file or not out_file:
            messagebox.showerror("Lỗi", "Vui lòng chọn file XML và nơi lưu Excel.")
            return
        if not selected_classes:
            messagebox.showerror("Lỗi", "Chưa có class nào được chọn.")
            return
        try:
            update_progress(5)
            extract_and_write_to_excel(in_file, out_file, selected_classes)
            update_progress(100)
            messagebox.showinfo("Thành công", f"Đã lưu file Excel:\n{out_file}")
            update_progress(0)
        except Exception as e:
            messagebox.showerror("Lỗi xử lý", str(e))
            update_progress(0)

    root = tk.Tk()
    root.title("XML to Excel Exporter")
    root.geometry("900x680")  # Tăng chiều cao
    root.resizable(False, False)

    input_path = tk.StringVar()
    output_path = tk.StringVar()

    top_frame = tk.Frame(root)
    top_frame.pack(pady=10, fill="x")

    tk.Label(top_frame, text="Chọn file XML:").grid(row=0, column=0, sticky='e', padx=5)
    tk.Entry(top_frame, textvariable=input_path, width=70).grid(row=0, column=1)
    tk.Button(top_frame, text="Browse", command=select_input_file).grid(row=0, column=2, padx=5)

    tk.Label(top_frame, text="Lưu file Excel:").grid(row=1, column=0, sticky='e', padx=5)
    tk.Entry(top_frame, textvariable=output_path, width=70).grid(row=1, column=1)
    tk.Button(top_frame, text="Browse", command=select_output_file).grid(row=1, column=2, padx=5)

    label_frame = tk.Frame(root)
    label_frame.pack(pady=5)
    tk.Label(label_frame, text="Danh sách toàn bộ class", font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=60)
    tk.Label(label_frame, text="Các class muốn xuất dữ liệu", font=('Arial', 10, 'bold')).grid(row=0, column=2, padx=60)

    middle_frame = tk.Frame(root)
    middle_frame.pack()

    frame_left = tk.Frame(middle_frame)
    frame_left.grid(row=0, column=0, padx=10)
    scrollbar_left = tk.Scrollbar(frame_left)
    class_available = tk.Listbox(frame_left, selectmode=tk.MULTIPLE, width=40, height=18, yscrollcommand=scrollbar_left.set)
    scrollbar_left.config(command=class_available.yview)
    scrollbar_left.pack(side=tk.RIGHT, fill=tk.Y)
    class_available.pack(side=tk.LEFT)

    button_frame = tk.Frame(middle_frame)
    button_frame.grid(row=0, column=1, padx=20)
    tk.Button(button_frame, text=">", width=5, command=lambda: move_selected(class_available, class_selected)).pack(pady=5)
    tk.Button(button_frame, text=">>", width=5, command=lambda: move_all(class_available, class_selected)).pack(pady=5)
    tk.Button(button_frame, text="<", width=5, command=lambda: move_selected(class_selected, class_available)).pack(pady=5)
    tk.Button(button_frame, text="<<", width=5, command=lambda: move_all(class_selected, class_available)).pack(pady=5)

    frame_right = tk.Frame(middle_frame)
    frame_right.grid(row=0, column=2, padx=10)
    scrollbar_right = tk.Scrollbar(frame_right)
    class_selected = tk.Listbox(frame_right, selectmode=tk.MULTIPLE, width=40, height=18, yscrollcommand=scrollbar_right.set)
    scrollbar_right.config(command=class_selected.yview)
    scrollbar_right.pack(side=tk.RIGHT, fill=tk.Y)
    class_selected.pack(side=tk.LEFT)

    bottom_frame = tk.Frame(root)
    bottom_frame.pack(pady=10)
    tk.Button(bottom_frame, text="Bắt đầu xử lý", command=convert_file, bg="green", fg="white", width=25).pack()

    # Thanh tiến trình
    progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
    progress_bar.pack(pady=10)

    for cls in sorted(ALL_CLASSES):
        class_available.insert(tk.END, cls)

    root.mainloop()

if __name__ == "__main__":
    run_gui()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Vadusha CRM — Современное одностраничное PyQt5 приложение.
Автор: ChatGPT
Функции:
- Splash screen с анимацией
- Темная/светлая тема с переключателем
- Управление задачами: добавить, редактировать, удалить, отметить выполнено
- Просмотр задачи в отдельном окне
- Импорт/экспорт в Excel
- Анимации и приятный UI
"""

import sys
import sqlite3
import datetime
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QLineEdit, QTextEdit, QComboBox, QDateEdit,
    QFileDialog, QMessageBox, QDialog, QCheckBox, QHeaderView, QSpacerItem, QSizePolicy
)
from PyQt5.QtGui import QColor, QFont, QBrush, QIcon, QCursor
from PyQt5.QtCore import Qt, QTimer, QPropertyAnimation, QEasingCurve, pyqtSignal
from openpyxl import Workbook, load_workbook


DB_FILE = "vadusha_crm.db"
USERS = ["Анастасия", "Николай", "Анна", "Татьяна", "Вадим"]

# --- Инициализация базы данных ---
def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user TEXT,
            category TEXT,
            description TEXT,
            link TEXT,
            deadline TEXT,
            urgency TEXT,
            done INTEGER DEFAULT 0,
            responsible TEXT,
            comment TEXT,
            attachment TEXT
        )
    """)
    conn.commit()
    conn.close()


# --- Splash Screen с плавным появлением и исчезанием ---
class Splash(QWidget):
    splash_done = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setFixedSize(500, 300)
        self.setStyleSheet("""
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #1e1e2f, stop:1 #282c34);
        """)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.label = QLabel("Vadusha CRM")
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setFont(QFont("Segoe UI", 48, QFont.Bold))
        self.label.setStyleSheet("color: #61dafb;")
        layout.addStretch()
        layout.addWidget(self.label)
        layout.addStretch()

        self.setWindowOpacity(0)
        self.anim_in()

    def anim_in(self):
        self.anim = QPropertyAnimation(self, b"windowOpacity")
        self.anim.setDuration(1600)
        self.anim.setStartValue(0)
        self.anim.setEndValue(1)
        self.anim.setEasingCurve(QEasingCurve.InOutQuad)
        self.anim.finished.connect(self.hold)
        self.anim.start()

    def hold(self):
        QTimer.singleShot(1200, self.anim_out)

    def anim_out(self):
        self.anim2 = QPropertyAnimation(self, b"windowOpacity")
        self.anim2.setDuration(1000)
        self.anim2.setStartValue(1)
        self.anim2.setEndValue(0)
        self.anim2.setEasingCurve(QEasingCurve.InOutQuad)
        self.anim2.finished.connect(self.finish)
        self.anim2.start()

    def finish(self):
        self.splash_done.emit()
        self.close()


# --- Диалог просмотра задачи ---
class ViewTaskDialog(QDialog):
    def __init__(self, task_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Просмотр задачи")
        self.setMinimumWidth(600)
        self.task_data = task_data
        self.build_ui()

    def build_ui(self):
        layout = QVBoxLayout(self)

        fields = [
            ("Пользователь:", self.task_data[1]),
            ("Категория:", self.task_data[2]),
            ("Описание:", self.task_data[3]),
            ("Ссылка:", self.task_data[4]),
            ("Срок исполнения:", self.task_data[5]),
            ("Срочность:", self.task_data[6]),
            ("Выполнено:", "Да" if self.task_data[7] else "Нет"),
            ("Ответственный:", self.task_data[8]),
            ("Комментарий:", self.task_data[9]),
            ("Вложение:", self.task_data[10]),
        ]

        for label_text, value in fields:
            lbl = QLabel(f"<b>{label_text}</b> {value if value else '-'}")
            lbl.setWordWrap(True)
            layout.addWidget(lbl)

        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close, alignment=Qt.AlignRight)


# --- Диалог добавления / редактирования задачи ---
class TaskDialog(QDialog):
    def __init__(self, parent=None, task=None, user=""):
        super().__init__(parent)
        self.setWindowTitle("Добавить / Редактировать задачу")
        self.setMinimumSize(450, 600)
        self.task = task
        self.user = user
        self.build_ui()
        if task:
            self.load_task()

    def build_ui(self):
        v = QVBoxLayout(self)

        self.cat = QComboBox()
        self.cat.addItems([
            "Обращения граждан (ответственные)",
            "Обращения граждан (комментарии)",
            "Доверенности",
            "ЛНА"
        ])

        self.desc = QTextEdit()
        self.desc.setPlaceholderText("Описание задачи")

        self.link = QLineEdit()
        self.link.setPlaceholderText("Ссылка (URL или путь)")

        self.deadline = QDateEdit()
        self.deadline.setCalendarPopup(True)
        self.deadline.setDate(datetime.date.today())

        self.urgency = QComboBox()
        self.urgency.addItems(["Низкая", "Средняя", "Высокая"])

        self.resp = QComboBox()
        self.resp.addItems(USERS)

        self.comment = QTextEdit()
        self.comment.setPlaceholderText("Комментарий")

        self.attach = QLineEdit()
        self.attach.setPlaceholderText("Путь к файлу")

        self.btn_file = QPushButton("Обзор...")
        self.btn_file.clicked.connect(self.select_file)

        self.chk_done = QCheckBox("Выполнено")

        self.btn_save = QPushButton("Сохранить")
        self.btn_save.clicked.connect(self.save)

        # Layout
        v.addWidget(QLabel("Категория:"))
        v.addWidget(self.cat)
        v.addWidget(QLabel("Описание:"))
        v.addWidget(self.desc)
        v.addWidget(QLabel("Ссылка:"))
        v.addWidget(self.link)
        v.addWidget(QLabel("Срок исполнения:"))
        v.addWidget(self.deadline)
        v.addWidget(QLabel("Срочность:"))
        v.addWidget(self.urgency)
        v.addWidget(QLabel("Ответственный:"))
        v.addWidget(self.resp)
        v.addWidget(QLabel("Комментарий:"))
        v.addWidget(self.comment)

        h_attach = QHBoxLayout()
        h_attach.addWidget(self.attach)
        h_attach.addWidget(self.btn_file)
        v.addLayout(h_attach)

        v.addWidget(self.chk_done)
        v.addStretch()
        v.addWidget(self.btn_save)

    def select_file(self):
        p, _ = QFileDialog.getOpenFileName(self, "Выберите файл")
        if p:
            self.attach.setText(p)

    def load_task(self):
        t = self.task
        self.cat.setCurrentText(t[2])
        self.desc.setPlainText(t[3])
        self.link.setText(t[4])
        self.deadline.setDate(datetime.datetime.strptime(t[5], "%Y-%m-%d").date())
        self.urgency.setCurrentText(t[6])
        self.resp.setCurrentText(t[8])
        self.comment.setPlainText(t[9])
        self.attach.setText(t[10])
        self.chk_done.setChecked(t[7] == 1)

    def save(self):
        if not self.desc.toPlainText().strip():
            QMessageBox.warning(self, "Ошибка", "Описание задачи не может быть пустым.")
            return

        data = (
            self.user,
            self.cat.currentText(),
            self.desc.toPlainText(),
            self.link.text(),
            self.deadline.date().toString("yyyy-MM-dd"),
            self.urgency.currentText(),
            1 if self.chk_done.isChecked() else 0,
            self.resp.currentText(),
            self.comment.toPlainText(),
            self.attach.text()
        )
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        if self.task:
            cur.execute("""
                UPDATE tasks SET
                    user=?, category=?, description=?, link=?, deadline=?, urgency=?,
                    done=?, responsible=?, comment=?, attachment=?
                WHERE id=?
            """, data + (self.task[0],))
        else:
            cur.execute("""
                INSERT INTO tasks
                (user, category, description, link, deadline, urgency, done, responsible, comment, attachment)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, data)
        conn.commit()
        conn.close()
        self.accept()


# --- Главное окно ---
class MainWindow(QWidget):
    def __init__(self, user="User"):
        super().__init__()
        self.user = user
        self.setWindowTitle("Vadusha CRM")
        self.resize(1024, 700)
        self.build_ui()
        self.load_tasks()
        self.dark_mode = False
        self.apply_theme()

    def build_ui(self):
        main_layout = QVBoxLayout(self)

        # Заголовок с переключателем темы
        header = QHBoxLayout()
        self.lbl_title = QLabel(f"Vadusha CRM — Добро пожаловать, {self.user}!")
        self.lbl_title.setFont(QFont("Segoe UI", 20, QFont.Bold))
        self.btn_toggle_theme = QPushButton("Тёмная тема")
        self.btn_toggle_theme.setCursor(QCursor(Qt.PointingHandCursor))
        self.btn_toggle_theme.clicked.connect(self.toggle_theme)
        header.addWidget(self.lbl_title)
        header.addStretch()
        header.addWidget(self.btn_toggle_theme)
        main_layout.addLayout(header)

        # Фильтр по категории
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Фильтр по категории:"))
        self.filter_combo = QComboBox()
        self.filter_combo.addItem("Все")
        self.filter_combo.addItems([
            "Обращения граждан (ответственные)",
            "Обращения граждан (комментарии)",
            "Доверенности",
            "ЛНА"
        ])
        self.filter_combo.currentIndexChanged.connect(self.load_tasks)
        filter_layout.addStretch()
        main_layout.addLayout(filter_layout)
        filter_layout.addWidget(self.filter_combo)

        # Таблица задач
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "Категория", "Описание", "Срок", "Срочность", "Выполнено", "Ответственный", "Комментарий", "Действия"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(self.table.SelectRows)
        self.table.setEditTriggers(self.table.NoEditTriggers)
        main_layout.addWidget(self.table)

        # Кнопки управления
        btn_layout = QHBoxLayout()
        self.btn_add = QPushButton("Добавить задачу")
        self.btn_add.clicked.connect(self.add_task)
        self.btn_edit = QPushButton("Редактировать задачу")
        self.btn_edit.clicked.connect(self.edit_task)
        self.btn_delete = QPushButton("Удалить задачу")
        self.btn_delete.clicked.connect(self.delete_task)
        self.btn_view = QPushButton("Просмотреть задачу")
        self.btn_view.clicked.connect(self.view_task)
        self.btn_export = QPushButton("Экспорт в Excel")
        self.btn_export.clicked.connect(self.export_excel)
        self.btn_import = QPushButton("Импорт из Excel")
        self.btn_import.clicked.connect(self.import_excel)

        btn_layout.addWidget(self.btn_add)
        btn_layout.addWidget(self.btn_edit)
        btn_layout.addWidget(self.btn_delete)
        btn_layout.addWidget(self.btn_view)
        btn_layout.addWidget(self.btn_export)
        btn_layout.addWidget(self.btn_import)
        main_layout.addLayout(btn_layout)

    def apply_theme(self):
        if self.dark_mode:
            self.setStyleSheet("""
                QWidget {
                    background-color: #121212;
                    color: #e0e0e0;
                    font-family: "Segoe UI";
                }
                QPushButton {
                    background-color: #1e88e5;
                    color: white;
                    border-radius: 6px;
                    padding: 6px 12px;
                }
                QPushButton:hover {
                    background-color: #1565c0;
                }
                QTableWidget {
                    background-color: #1e1e1e;
                    gridline-color: #333;
                }
                QHeaderView::section {
                    background-color: #272727;
                    padding: 4px;
                    border: none;
                    font-weight: bold;
                }
                QLineEdit, QTextEdit, QComboBox, QDateEdit {
                    background-color: #1e1e1e;
                    border: 1px solid #444;
                    border-radius: 4px;
                    padding: 4px;
                    color: #ddd;
                }
            """)
            self.btn_toggle_theme.setText("Светлая тема")
        else:
            self.setStyleSheet("""
                QWidget {
                    background-color: #f0f0f0;
                    color: #202020;
                    font-family: "Segoe UI";
                }
                QPushButton {
                    background-color: #1976d2;
                    color: white;
                    border-radius: 6px;
                    padding: 6px 12px;
                }
                QPushButton:hover {
                    background-color: #1565c0;
                }
                QTableWidget {
                    background-color: white;
                    gridline-color: #ddd;
                }
                QHeaderView::section {
                    background-color: #e0e0e0;
                    padding: 4px;
                    border: none;
                    font-weight: bold;
                }
                QLineEdit, QTextEdit, QComboBox, QDateEdit {
                    background-color: white;
                    border: 1px solid #ccc;
                    border-radius: 4px;
                    padding: 4px;
                    color: #202020;
                }
            """)
            self.btn_toggle_theme.setText("Тёмная тема")

    def toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self.apply_theme()

    def load_tasks(self):
        self.table.setRowCount(0)
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        category_filter = self.filter_combo.currentText()
        if category_filter == "Все":
            cur.execute("SELECT * FROM tasks ORDER BY deadline ASC")
        else:
            cur.execute("SELECT * FROM tasks WHERE category=? ORDER BY deadline ASC", (category_filter,))
        tasks = cur.fetchall()
        conn.close()

        for row_num, task in enumerate(tasks):
            self.table.insertRow(row_num)
            # Категория
            self.table.setItem(row_num, 0, QTableWidgetItem(task[2]))
            # Описание (сократить до 40 символов)
            desc = task[3] if len(task[3]) < 40 else task[3][:37] + "..."
            item_desc = QTableWidgetItem(desc)
            item_desc.setToolTip(task[3])
            self.table.setItem(row_num, 1, item_desc)
            # Срок
            self.table.setItem(row_num, 2, QTableWidgetItem(task[5]))
            # Срочность
            item_urg = QTableWidgetItem(task[6])
            if task[6] == "Высокая":
                item_urg.setBackground(QColor("#e57373"))
            elif task[6] == "Средняя":
                item_urg.setBackground(QColor("#ffb74d"))
            else:
                item_urg.setBackground(QColor("#aed581"))
            self.table.setItem(row_num, 3, item_urg)
            # Выполнено
            done_item = QTableWidgetItem("Да" if task[7] else "Нет")
            if task[7]:
                done_item.setForeground(QBrush(QColor("green")))
            else:
                # Если просрочено (срок меньше текущей даты) - красный цвет
                deadline_date = datetime.datetime.strptime(task[5], "%Y-%m-%d").date()
                if deadline_date < datetime.date.today():
                    done_item.setForeground(QBrush(QColor("red")))
            self.table.setItem(row_num, 4, done_item)
            # Ответственный
            self.table.setItem(row_num, 5, QTableWidgetItem(task[8]))
            # Комментарий (сократить)
            comment = task[9] if task[9] and len(task[9]) < 40 else (task[9][:37] + "...") if task[9] else "-"
            self.table.setItem(row_num, 6, QTableWidgetItem(comment))
            # Действия: кнопки редактирования и удаления
            btn_view = QPushButton("Просмотреть")
            btn_view.setCursor(QCursor(Qt.PointingHandCursor))
            btn_view.clicked.connect(lambda _, t=task: self.view_task_direct(t))
            self.table.setCellWidget(row_num, 7, btn_view)

    def add_task(self):
        dlg = TaskDialog(self, user=self.user)
        if dlg.exec():
            self.load_tasks()

    def get_selected_task(self):
        selected = self.table.selectedItems()
        if not selected:
            return None
        row = selected[0].row()
        desc = self.table.item(row, 1).toolTip() or self.table.item(row, 1).text()
        deadline = self.table.item(row, 2).text()
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        # Выбираем задачу по описанию и сроку (уникальность не гарантируется, можно улучшить)
        cur.execute("SELECT * FROM tasks WHERE description=? AND deadline=?", (desc, deadline))
        task = cur.fetchone()
        conn.close()
        return task

    def edit_task(self):
        task = self.get_selected_task()
        if not task:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите задачу для редактирования.")
            return
        dlg = TaskDialog(self, task=task, user=self.user)
        if dlg.exec():
            self.load_tasks()

    def delete_task(self):
        task = self.get_selected_task()
        if not task:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите задачу для удаления.")
            return
        reply = QMessageBox.question(
            self, "Подтвердите удаление",
            "Вы действительно хотите удалить выбранную задачу?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            conn = sqlite3.connect(DB_FILE)
            cur = conn.cursor()
            cur.execute("DELETE FROM tasks WHERE id=?", (task[0],))
            conn.commit()
            conn.close()
            self.load_tasks()

    def view_task(self):
        task = self.get_selected_task()
        if not task:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите задачу для просмотра.")
            return
        dlg = ViewTaskDialog(task, self)
        dlg.exec()

    def view_task_direct(self, task):
        dlg = ViewTaskDialog(task, self)
        dlg.exec()

    def export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить Excel", "", "Excel Files (*.xlsx)")
        if not path:
            return
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT * FROM tasks")
        tasks = cur.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Задачи Vadusha CRM"
        headers = ["ID", "Пользователь", "Категория", "Описание", "Ссылка", "Срок", "Срочность", "Выполнено", "Ответственный", "Комментарий", "Вложение"]
        ws.append(headers)

        for task in tasks:
            ws.append(task)

        try:
            wb.save(path)
            QMessageBox.information(self, "Успех", "Экспорт завершён успешно.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{e}")

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Выберите Excel для импорта", "", "Excel Files (*.xlsx)")
        if not path:
            return
        try:
            wb = load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows[0] != ("ID", "Пользователь", "Категория", "Описание", "Ссылка", "Срок", "Срочность", "Выполнено", "Ответственный", "Комментарий", "Вложение"):
                QMessageBox.warning(self, "Ошибка", "Неверный формат файла.")
                return
            conn = sqlite3.connect(DB_FILE)
            cur = conn.cursor()
            for row in rows[1:]:
                # Проверяем, есть ли уже задача с таким ID (если ID совпадает, обновляем, иначе вставляем)
                cur.execute("SELECT id FROM tasks WHERE id=?", (row[0],))
                exists = cur.fetchone()
                if exists:
                    cur.execute("""
                        UPDATE tasks SET
                            user=?, category=?, description=?, link=?, deadline=?, urgency=?, done=?, responsible=?, comment=?, attachment=?
                        WHERE id=?
                    """, row[1:] + (row[0],))
                else:
                    cur.execute("""
                        INSERT INTO tasks (id, user, category, description, link, deadline, urgency, done, responsible, comment, attachment)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, row)
            conn.commit()
            conn.close()
            QMessageBox.information(self, "Успех", "Импорт завершён успешно.")
            self.load_tasks()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось импортировать файл:\n{e}")



def main():
    init_db()
    app = QApplication(sys.argv)

    splash = Splash()
    splash.splash_done.connect(lambda: show_main(app, splash))
    splash.show()

    sys.exit(app.exec())


def show_main(app, splash):
    main_win = MainWindow(user="Вадим")
    main_win.show()
    # Убедимся, что splash закрыт
    splash.close()


if __name__ == "__main__":
    main()

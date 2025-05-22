import os
import re
import getpass
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def convert_density(density_value_t_per_mm3):
    return float(density_value_t_per_mm3) * 1e9  # t/mm³ → g/cm³

def extract_material_data(input_folder):
    lamina_data = []
    isotropic_data = []

    for filename in os.listdir(input_folder):
        if filename.endswith(".inp"):
            file_path = os.path.join(input_folder, filename)
            with open(file_path, "r", encoding="utf-8") as f:
                lines = f.readlines()

            current_material = None
            current_type = None
            density = None

            # ISOTROPIC
            e_modul = None
            poisson_ratio = None
            yield_stress = None
            uts = None

            # LAMINA
            e11 = None
            e22 = None
            nu12 = None
            g12 = None
            g13 = None
            g23 = None
            fail_stress_s1 = None
            fail_stress_s2 = None
            fail_stress_s3 = None
            fail_stress_s4 = None
            fail_stress_s5 = None

            i = 0
            while i < len(lines):
                line = lines[i].strip()

                if line.startswith("*MATERIAL"):
                    # Save previous material
                    if current_material and current_type:
                        entry = {
                            "Material": current_material,
                            "Density (g/cm³)": round(convert_density(density), 6) if density else None
                        }

                        if current_type == "ISOTROPIC":
                            entry["Young's Modulus (MPa)"] = e_modul
                            entry["Poisson's Ratio"] = poisson_ratio
                            entry["Yield Stress (MPa)"] = yield_stress
                            entry["UTS (MPa)"] = uts
                            isotropic_data.append(entry)

                        elif current_type == "LAMINA":
                            entry["E11"] = e11
                            entry["E22"] = e22
                            entry["NU12"] = nu12
                            entry["G12"] = g12
                            entry["G13"] = g13
                            entry["G23"] = g23
                            entry["S1 (MPa)"] = fail_stress_s1
                            entry["S2 (MPa)"] = fail_stress_s2
                            entry["S3 (MPa)"] = fail_stress_s3
                            entry["S4 (MPa)"] = fail_stress_s4
                            entry["S5 (MPa)"] = fail_stress_s5
                            lamina_data.append(entry)

                    # Start new material
                    match = re.search(r"\*MATERIAL, *NAME *= *(.*)", line)
                    current_material = match.group(1).strip() if match else None
                    current_type = None
                    density = None
                    e_modul = poisson_ratio = yield_stress = uts = None
                    e11 = e22 = nu12 = g12 = g13 = g23 = None
                    fail_stress_s1 = fail_stress_s2 = fail_stress_s3 = fail_stress_s4 = fail_stress_s5 = None

                elif line.startswith("*DENSITY"):
                    i += 1
                    values = re.split(r"[,\s]+", lines[i].strip())
                    if values:
                        try:
                            density = float(values[0])
                        except ValueError:
                            pass

                elif line.startswith("*ELASTIC"):
                    match = re.search(r"TYPE *= *(LAMINA|ISOTROPIC)", line.upper())
                    current_type = match.group(1) if match else None

                    i += 1
                    values = []
                    while i < len(lines) and not lines[i].strip().startswith("*"):
                        values += re.split(r"[,\s]+", lines[i].strip())
                        i += 1
                    i -= 1  # step back after loop

                    try:
                        if current_type == "ISOTROPIC" and len(values) >= 2:
                            e_modul = float(values[0])
                            poisson_ratio = float(values[1])
                        elif current_type == "LAMINA" and len(values) >= 6:
                            e11 = float(values[0])
                            e22 = float(values[1])
                            nu12 = float(values[2])
                            g12 = float(values[3])
                            g13 = float(values[4])
                            g23 = float(values[5])
                    except ValueError:
                        pass

                elif line.startswith("*PLASTIC"):
                    plastic_data = []
                    i += 1
                    while i < len(lines) and not lines[i].strip().startswith("*"):
                        plastic_data += re.split(r"[,\s]+", lines[i].strip())
                        i += 1
                    i -= 1
                    try:
                        if len(plastic_data) >= 1:
                            yield_stress = float(plastic_data[0])
                        if len(plastic_data) >= 4:
                            uts = float(plastic_data[3])
                    except ValueError:
                        pass

                elif line.startswith("*FAIL STRESS"):
                    i += 1
                    values = re.split(r"[,\s]+", lines[i].strip())
                    try:
                        if len(values) >= 5:
                            fail_stress_s1 = float(values[0])
                            fail_stress_s2 = float(values[1])
                            fail_stress_s3 = float(values[2])
                            fail_stress_s4 = float(values[3])
                            fail_stress_s5 = float(values[4])
                    except ValueError:
                        pass

                i += 1

            # Save last material
            if current_material and current_type:
                entry = {
                    "Material": current_material,
                    "Density (g/cm³)": round(convert_density(density), 6) if density else None
                }

                if current_type == "ISOTROPIC":
                    entry["Young's Modulus (MPa)"] = e_modul
                    entry["Poisson's Ratio"] = poisson_ratio
                    entry["Yield Stress (MPa)"] = yield_stress
                    entry["UTS (MPa)"] = uts
                    isotropic_data.append(entry)

                elif current_type == "LAMINA":
                    entry["E11"] = e11
                    entry["E22"] = e22
                    entry["NU12"] = nu12
                    entry["G12"] = g12
                    entry["G13"] = g13
                    entry["G23"] = g23
                    entry["S1 (MPa)"] = fail_stress_s1
                    entry["S2 (MPa)"] = fail_stress_s2
                    entry["S3 (MPa)"] = fail_stress_s3
                    entry["S4 (MPa)"] = fail_stress_s4
                    entry["S5 (MPa)"] = fail_stress_s5
                    lamina_data.append(entry)

    return lamina_data, isotropic_data

def create_excel(lamina_data, isotropic_data, output_path):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        pd.DataFrame(lamina_data).to_excel(writer, index=False, sheet_name='Orthotropic Materials')
        pd.DataFrame(isotropic_data).to_excel(writer, index=False, sheet_name='Isotropic Materials')

    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for col_idx, column_cells in enumerate(sheet.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)

    wb.save(output_path)

def main():
    current_user = getpass.getuser()
    output_folder = fr"C:\Users\{current_user}\CMM_ABAQUS_MATERIAL_DATABASE"
    os.makedirs(output_folder, exist_ok=True)

    input_folder = fr"\\spe-ch-md9.net\hin\programs\CMM\Shareable-Services\FEM_ExportCards\Abaqus\{current_user}"
    output_file = os.path.join(output_folder, "Material_Database.xlsx")

    if not os.path.exists(input_folder):
        print(f"Input folder not found: {input_folder}")
        return

    print(f"Processing materials from: {input_folder}")
    lamina_data, isotropic_data = extract_material_data(input_folder)
    print(f"Found {len(lamina_data)} orthotropic materials.")
    print(f"Found {len(isotropic_data)} isotropic materials.")

    create_excel(lamina_data, isotropic_data, output_file)
    print(f"Excel file created: {output_file}")

if __name__ == "__main__":
    main()

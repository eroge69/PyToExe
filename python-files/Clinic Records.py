import tkinter as tk
from tkinter import messagebox, ttk
import openpyxl
import os
from datetime import datetime

EXCEL_FILE = "patients.xlsx"

PRESCRIPTIONS = [
    "PCM", "NISE", "CETZINE", "B.M.H", "AVIL", "DICLO", "FAMO", "FA", "CPM", "DEXA",
    "DOMSTAL", "LOPERA", "CYCLOPAM", "OCID", "DERIPHYLINE", "SB", "PHENARGAN", "MEFTAL",
    "STEMETIL", "UNIENZYME", "ALPRAX", "DULCOLAX", "CONS", "DB", "DRESSING", "EAR DROP",
    "EYE DROP", "INJECTION", "NASAL DROP", "NEBULIZER", "OINMENT"
]

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Patients"
        headers = ["ID", "Date", "Name", "Age", "Gender", "Prescription", "DID", "Amount"]
        sheet.append(headers)
        wb.save(EXCEL_FILE)

def load_patients():
    patients = []
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["Patients"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            patients.append(row)
    return patients

def save_patient(patient):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["Patients"]
    ids = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]
    new_id = max(ids) + 1 if ids else 1
    sheet.append([new_id] + patient)
    wb.save(EXCEL_FILE)

def update_patient(record_id, updated_data):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["Patients"]
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == record_id:
            for i, value in enumerate(updated_data):
                row[i + 1].value = value
            break
    wb.save(EXCEL_FILE)

def delete_patient(record_id):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["Patients"]
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == record_id:
            sheet.delete_rows(i)
            break
    wb.save(EXCEL_FILE)

class PatientApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Patient Record Management")
        self.root.geometry("1050x700")
        self.root.configure(bg="#e6f2ff")

        style = ttk.Style()
        style.configure("Treeview", background="#f2f2f2", foreground="black", rowheight=25, fieldbackground="#f2f2f2")
        style.map('Treeview', background=[('selected', '#b3d9ff')])

        title = tk.Label(root, text="Patient Record Management", font=("Helvetica", 20, "bold"), bg="#e6f2ff", fg="#003366")
        title.pack(pady=10)

        form_frame = tk.Frame(root, bg="#e6f2ff")
        form_frame.pack(pady=10)

        labels = ["Name", "Age", "Gender", "Prescription", "DID", "Amount"]
        for i, text in enumerate(labels):
            tk.Label(form_frame, text=text, bg="#e6f2ff", font=("Arial", 10, "bold")).grid(row=i, column=0, sticky="w", padx=10, pady=5)

        self.name_entry = tk.Entry(form_frame, width=30)
        self.age_entry = tk.Entry(form_frame, width=30)
        self.gender_var = tk.StringVar()
        self.gender_combo = ttk.Combobox(form_frame, textvariable=self.gender_var, values=["Male", "Female", "Other"], state="readonly", width=28)

        self.prescription_listbox = tk.Listbox(form_frame, selectmode="multiple", height=6, exportselection=False, width=28)
        for item in PRESCRIPTIONS:
            self.prescription_listbox.insert(tk.END, item)

        self.did_entry = tk.Entry(form_frame, width=30)
        self.amount_entry = tk.Entry(form_frame, width=30)

        widgets = [self.name_entry, self.age_entry, self.gender_combo, self.prescription_listbox, self.did_entry, self.amount_entry]
        for i, widget in enumerate(widgets):
            widget.grid(row=i, column=1, pady=5, padx=5)

        btn_frame = tk.Frame(root, bg="#e6f2ff")
        btn_frame.pack(pady=10)
        buttons = [
            ("Add Patient", self.add_patient),
            ("Update Selected", self.update_selected),
            ("Delete Selected", self.delete_selected),
            ("Refresh List", self.load_patient_list)
        ]
        for i, (text, command) in enumerate(buttons):
            tk.Button(btn_frame, text=text, command=command, width=20, bg="#cce5ff", fg="black", relief="raised").grid(row=0, column=i, padx=5)

        search_frame = tk.Frame(root, bg="#e6f2ff")
        search_frame.pack(pady=10)
        tk.Label(search_frame, text="Search by Name:", bg="#e6f2ff", font=("Arial", 10)).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var, width=40)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_entry.bind("<KeyRelease>", lambda e: self.load_patient_list())

        self.tree = ttk.Treeview(root, columns=("ID", "Date", "Name", "Age", "Gender", "Prescription", "DID", "Amount"), show="headings")
        for col in ("ID", "Date", "Name", "Age", "Gender", "Prescription", "DID", "Amount"):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor="center")
        self.tree.pack(pady=10, fill="both", expand=True)

        self.load_patient_list()

    def get_selected_prescriptions(self):
        return ", ".join([self.prescription_listbox.get(i) for i in self.prescription_listbox.curselection()])

    def add_patient(self):
        name = self.name_entry.get().strip()
        age = self.age_entry.get().strip()
        gender = self.gender_var.get()
        prescription = self.get_selected_prescriptions()
        did = self.did_entry.get().strip()
        amount = self.amount_entry.get().strip()

        if not name or not age or not gender:
            messagebox.showerror("Error", "Name, Age, and Gender are required!")
            return

        try:
            age = int(age)
            amount = float(amount) if amount else 0.0
        except:
            messagebox.showerror("Error", "Age must be a number and Amount must be numeric!")
            return

        date_str = datetime.now().strftime("%Y-%m-%d")
        patient = [date_str, name, age, gender, prescription, did, amount]
        save_patient(patient)

        messagebox.showinfo("Success", "Patient added successfully!")
        self.clear_fields()
        self.load_patient_list()

    def update_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "No patient selected")
            return

        item = self.tree.item(selected[0])['values']
        record_id = item[0]

        name = self.name_entry.get().strip()
        age = self.age_entry.get().strip()
        gender = self.gender_var.get()
        prescription = self.get_selected_prescriptions()
        did = self.did_entry.get().strip()
        amount = self.amount_entry.get().strip()

        try:
            age = int(age)
            amount = float(amount) if amount else 0.0
        except:
            messagebox.showerror("Error", "Age must be a number and Amount must be numeric!")
            return

        updated_data = [item[1], name, age, gender, prescription, did, amount]
        update_patient(record_id, updated_data)
        messagebox.showinfo("Success", "Patient record updated!")
        self.clear_fields()
        self.load_patient_list()

    def delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "No patient selected")
            return
        item = self.tree.item(selected[0])['values']
        record_id = item[0]
        delete_patient(record_id)
        messagebox.showinfo("Deleted", "Patient record deleted")
        self.load_patient_list()

    def load_patient_list(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        patients = load_patients()
        search_text = self.search_var.get().lower()
        for p in patients:
            if search_text in str(p[2]).lower():
                self.tree.insert("", tk.END, values=p)

    def clear_fields(self):
        self.name_entry.delete(0, tk.END)
        self.age_entry.delete(0, tk.END)
        self.gender_var.set('')
        self.prescription_listbox.selection_clear(0, tk.END)
        self.did_entry.delete(0, tk.END)
        self.amount_entry.delete(0, tk.END)

if __name__ == "__main__":
    init_excel()
    root = tk.Tk()
    app = PatientApp(root)
    root.mainloop()
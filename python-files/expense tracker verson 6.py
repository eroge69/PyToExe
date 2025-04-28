# file: expense_tracker_excel_gui.py

import os
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import simpledialog
from openpyxl import Workbook, load_workbook
import tempfile
import platform
import subprocess

DATA_FILE = 'expenses.xlsx'
SHEET_NAME = 'Transactions'

def initialize_excel():
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(['Date', 'Type', 'Amount', 'Description'])
        wb.save(DATA_FILE)

def load_transactions():
    wb = load_workbook(DATA_FILE)
    ws = wb[SHEET_NAME]
    return list(ws.iter_rows(min_row=2, values_only=True))

def add_transaction_gui(root, amount_entry, type_var, description_entry):
    try:
        amount = float(amount_entry.get())
        t_type = type_var.get()
        description = description_entry.get()
        if t_type not in ('income', 'expense'):
            messagebox.showerror("Error", "Please select Income or Expense.")
            return
        if not description:
            description = "N/A"
        wb = load_workbook(DATA_FILE)
        ws = wb[SHEET_NAME]
        ws.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            t_type,
            amount,
            description
        ])
        wb.save(DATA_FILE)
        messagebox.showinfo("Success", "Transaction added successfully.")
        root.destroy()
        main_gui()
    except ValueError:
        messagebox.showerror("Error", "Invalid amount. Please enter a valid number.")

def view_transactions_gui():
    transactions = load_transactions()
    view_window = tk.Toplevel()
    view_window.title("All Transactions")

    if not transactions:
        messagebox.showinfo("Transactions", "No transactions found.")
        return

    for idx, (date, t_type, amount, description) in enumerate(transactions, start=1):
        tk.Label(view_window, text=f"{idx}. {date} - {t_type.capitalize()} - ${amount:.2f} - {description}").pack(anchor='w')

    tk.Button(view_window, text="Home", command=view_window.destroy).pack(pady=5)

def view_balance_gui():
    transactions = load_transactions()
    balance = 0.0
    for date, t_type, amount, description in transactions:
        if t_type == 'income':
            balance += amount
        else:
            balance -= amount
    balance_window = tk.Toplevel()
    balance_window.title("Balance")
    tk.Label(balance_window, text=f"Current Balance: ${balance:.2f}").pack(pady=10)
    tk.Button(balance_window, text="Home", command=balance_window.destroy).pack(pady=5)

def clear_transactions():
    confirm = messagebox.askyesno("Confirm", "Are you sure you want to clear all transactions?")
    if confirm:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(['Date', 'Type', 'Amount', 'Description'])
        wb.save(DATA_FILE)
        messagebox.showinfo("Cleared", "All transactions have been cleared.")

def generate_report():
    transactions = load_transactions()
    if not transactions:
        messagebox.showinfo("Report", "No transactions to generate report.")
        return
    temp_report = tempfile.NamedTemporaryFile(delete=False, suffix='.txt')
    with open(temp_report.name, 'w') as f:
        f.write("Expense Tracker Report\n")
        f.write("=======================\n\n")
        for idx, (date, t_type, amount, description) in enumerate(transactions, start=1):
            f.write(f"{idx}. {date} - {t_type.capitalize()} - ${amount:.2f} - {description}\n")
        f.write("\nReport generated successfully.")

    if platform.system() == "Windows":
        os.startfile(temp_report.name, 'print')
    else:
        subprocess.run(['lp', temp_report.name])

def main_gui():
    initialize_excel()
    transactions = load_transactions()

    root = tk.Tk()
    root.title("Personal Expense Tracker (Excel Version)")

    tk.Label(root, text=f"Loaded {len(transactions)} existing transactions.", pady=10).pack()

    frame = tk.Frame(root)
    frame.pack(pady=10)

    tk.Label(frame, text="Amount:").grid(row=0, column=0, padx=5, pady=5)
    amount_entry = tk.Entry(frame)
    amount_entry.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(frame, text="Type:").grid(row=1, column=0, padx=5, pady=5)
    type_var = tk.StringVar()
    tk.Radiobutton(frame, text="Income", variable=type_var, value='income').grid(row=1, column=1, sticky='w')
    tk.Radiobutton(frame, text="Expense", variable=type_var, value='expense').grid(row=1, column=2, sticky='w')

    tk.Label(frame, text="Description:").grid(row=2, column=0, padx=5, pady=5)
    description_entry = tk.Entry(frame)
    description_entry.grid(row=2, column=1, columnspan=2, padx=5, pady=5)

    tk.Button(root, text="Add Transaction", command=lambda: add_transaction_gui(root, amount_entry, type_var, description_entry), width=30).pack(pady=5)
    tk.Button(root, text="View All Transactions", command=view_transactions_gui, width=30).pack(pady=5)
    tk.Button(root, text="View Balance", command=view_balance_gui, width=30).pack(pady=5)
    tk.Button(root, text="Generate and Print Report", command=generate_report, width=30).pack(pady=5)
    tk.Button(root, text="Clear All Transactions", command=clear_transactions, width=30).pack(pady=5)
    tk.Button(root, text="Exit", command=root.quit, width=30).pack(pady=5)

    root.mainloop()

if __name__ == "__main__":
    main_gui()

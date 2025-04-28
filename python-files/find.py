import tkinter as tk
from tkinter import filedialog, messagebox
import os
import re
from openpyxl import Workbook

class FolderCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Проверка папок и файлов")
        self.last_check_type = None

        pad_y = 3
        section_pad = 10

        self.folder_path = tk.StringVar()
        self.start_num_str = tk.StringVar()
        self.end_num_str = tk.StringVar()

        self.missing_numbers = []
        self.problem_numbers = []

        # GUI элементы
        tk.Label(root, text="Путь к папке:").grid(row=0, column=0, sticky='w', pady=(section_pad, pad_y))
        tk.Entry(root, textvariable=self.folder_path, width=60).grid(row=0, column=1, sticky='w', pady=(section_pad, pad_y))
        tk.Button(root, text="Обзор...", command=self.browse_folder).grid(row=0, column=2, pady=(section_pad, pad_y))

        # Валидация для цифровых полей
        val_cmd = root.register(self.validate_digits)
        
        tk.Label(root, text="Начало диапазона:").grid(row=1, column=0, sticky='ww', pady=pad_y)
        tk.Entry(root, textvariable=self.start_num_str, 
                validate='key', validatecommand=(val_cmd, '%S')).grid(row=1, column=1, sticky='w', pady=pad_y)

        tk.Label(root, text="Конец диапазона:").grid(row=2, column=0, sticky='ww', pady=pad_y)
        tk.Entry(root, textvariable=self.end_num_str,
                validate='key', validatecommand=(val_cmd, '%S')).grid(row=2, column=1, sticky='w', pady=pad_y)

        tk.Button(root, text="Проверить папки \n (Enter)", command=self.check_folders).grid(row=3, column=0, pady=(section_pad, pad_y), sticky='s')
        tk.Button(root, text="Проверить Справки \n (Пробел)", command=self.check_docx).grid(row=3, column=1, pady=(section_pad, pad_y), sticky='n')

        self.result_text = tk.Text(root, height=10, width=70)
        self.result_text.grid(row=4, column=0, columnspan=3, pady=pad_y, padx=10)

        tk.Button(root, text="Сохранить результаты", command=self.export_results).grid(row=5, column=0, columnspan=2, pady=(section_pad, pad_y), sticky='nsew')

    def validate_digits(self, char):
        """Разрешает ввод только цифр"""
        return char.isdigit()

    def browse_folder(self):
        """Выбор папки через диалоговое окно"""
        folder_selected = filedialog.askdirectory()
        self.folder_path.set(folder_selected)

    def extract_number(self, folder_name):
        """Извлекает числовую часть из имени папки"""
        match = re.match(r'^\d+', folder_name)
        return int(match.group()) if match else None

    def validate_range(self):
        """Проверка и обработка диапазона"""
        start_str = self.start_num_str.get().strip()
        end_str = self.end_num_str.get().strip()

        # Обработка начала диапазона
        if not start_str:
            start = 1
        else:
            start = int(start_str)

        # Проверка конца диапазона
        if not end_str:
            messagebox.showerror("Ошибка", "Укажите конец диапазона")
            return None, None
        try:
            end = int(end_str)
        except ValueError:
            messagebox.showerror("Ошибка", "Конец диапазона должен быть числом")
            return None, None

        if start > end:
            messagebox.showerror("Ошибка", "Начало диапазона не может быть больше конца")
            return None, None

        return start, end

    def check_folders(self):
        """Проверка отсутствующих папок"""
        self.last_check_type = "folders"
        path = self.folder_path.get()
        start, end = self.validate_range()
        if start is None or end is None:
            return

        if not os.path.isdir(path):
            messagebox.showerror("Ошибка", "Папка не существует")
            return

        expected = set(range(start, end + 1))
        found = set()

        for entry in os.listdir(path):
            if os.path.isdir(os.path.join(path, entry)):
                num = self.extract_number(entry)
                if num is not None and num in expected:
                    found.add(num)

        self.missing_numbers = sorted(expected - found)
        self.result_text.delete('1.0', tk.END)
        if self.missing_numbers:
            self.result_text.insert(tk.END, "Отсутствуют папки:\n" + ", ".join(map(str, self.missing_numbers)))
        else:
            self.result_text.insert(tk.END, "Папки — ВСЕ!")

    def check_docx(self):
        """Проверка отсутствия DOCX файлов"""
        self.last_check_type = "docx"
        path = self.folder_path.get()
        start, end = self.validate_range()
        if start is None or end is None:
            return

        if not os.path.isdir(path):
            messagebox.showerror("Ошибка", "Папка не существует")
            return

        self.problem_numbers = []

        for entry in os.listdir(path):
            entry_path = os.path.join(path, entry)
            if os.path.isdir(entry_path):
                num = self.extract_number(entry)
                if num is not None and start <= num <= end:
                    has_docx = False
                    for root_dir, _, files in os.walk(entry_path):
                        if any(f.lower().endswith('.docx') for f in files):
                            has_docx = True
                            break
                    if not has_docx:
                        self.problem_numbers.append(num)

        self.result_text.delete('1.0', tk.END)
        if self.problem_numbers:
            self.result_text.insert(tk.END, "Папки без справок:\n" + ", ".join(map(str, self.problem_numbers)))
        else:
            self.result_text.insert(tk.END, "СПРАВКИ - ВСЕ!")

    def export_results(self):
        """Экспорт результатов в Excel"""
        if not self.last_check_type:
            messagebox.showinfo("Инфо", "Сначала выполните проверку")
            return

        data = self.missing_numbers if self.last_check_type == "folders" else self.problem_numbers
        msg_type = "отсутствующих папок" if self.last_check_type == "folders" else "проблемных папок"

        if not data:
            messagebox.showinfo("Инфо", f"Нет данных {msg_type} для экспорта")
            return

        wb = Workbook()
        ws = wb.active
        for num in data:
            ws.append(['', num])  # Пустой первый столбец

        # Добавляем формулу ВПР в третью колонку
        if ws.max_row >= 1:
            ws['C1'] = 'ВПР(B1; A:B; 2; ЛОЖЬ)'

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title=f"Сохранить результаты {msg_type}"
        )
        
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Успех", "Данные успешно экспортированы")

if __name__ == "__main__":
    root = tk.Tk()
    app = FolderCheckerApp(root)
    
    # Привязка горячих клавиш
    root.bind('<Return>', lambda event: app.check_folders())
    root.bind('<space>', lambda event: app.check_docx())
    
    root.mainloop()
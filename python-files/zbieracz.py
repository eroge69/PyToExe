Python 3.14.0b1 (tags/v3.14.0b1:b092705, May  7 2025, 10:22:31) [MSC v.1943 64 bit (AMD64)] on win32
Enter "help" below or click "Help" above for more information.
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.utils import get_column_letter

class DaneOsoboweApp:
    def __init__(self, master):
        self.master = master
        master.title("Zbieranie Danych Osobowych")

        self.dane = {}
        self.pola = ["Imię", "Nazwisko", "PESEL", "Wykształcenie"]
        self.indeks_pola = 0
        self.entry = None
        self.label = None

        self.create_widgets()
        self.next_field()

    def create_widgets(self):
        self.label = tk.Label(self.master, text="")
        self.label.pack(pady=10)

        self.entry = tk.Entry(self.master)
        self.entry.pack(pady=5)
        self.entry.bind("<Return>", self.zatwierdz_dane) # Zatwierdzenie Enterem

        self.zatwierdz_button = tk.Button(self.master, text="Zatwierdź", command=self.zatwierdz_dane)
        self.zatwierdz_button.pack(pady=10)

        self.popraw_button = tk.Button(self.master, text="Popraw Dane", command=self.popraw_dane, state=tk.DISABLED)
        self.popraw_button.pack(pady=5)

        self.nastepny_button = tk.Button(self.master, text="Zapisz i Nowy", command=self.zapisz_i_nowy, state=tk.DISABLED)
        self.nastepny_button.pack(pady=10)

        self.dane_wyswietlane = tk.Label(self.master, text="")
        self.dane_wyswietlane.pack(pady=10)

        self.koniec_button = tk.Button(self.master, text="Zakończ", command=self.master.quit)
        self.koniec_button.pack(pady=10)

    def next_field(self):
        if self.indeks_pola < len(self.pola):
            self.label.config(text=f"Podaj {self.pola[self.indeks_pola]}:")
            self.entry.delete(0, tk.END)
            self.entry.config(state=tk.NORMAL)
            self.zatwierdz_button.config(state=tk.NORMAL)
            self.popraw_button.config(state=tk.DISABLED)
            self.nastepny_button.config(state=tk.DISABLED)
            self.dane_wyswietlane.config(text="")
        else:
            self.wyswietl_dane_do_potwierdzenia()

    def zatwierdz_dane(self, event=None):
        if self.indeks_pola < len(self.pola):
            pole = self.pola[self.indeks_pola]
            wartosc = self.entry.get()
            if wartosc:
                self.dane[pole] = wartosc
                self.indeks_pola += 1
                self.next_field()
            else:
                messagebox.showerror("Błąd", "Pole nie może być puste.")

    def wyswietl_dane_do_potwierdzenia(self):
        wyswietlany_tekst = "Wprowadzone dane:\n"
        for klucz, wartosc in self.dane.items():
            wyswietlany_tekst += f"{klucz}: {wartosc}\n"
        self.dane_wyswietlane.config(text=wyswietlany_tekst)
...         self.label.config(text="Czy chcesz poprawić dane?")
...         self.entry.config(state=tk.DISABLED)
...         self.zatwierdz_button.config(state=tk.DISABLED)
...         self.popraw_button.config(state=tk.NORMAL)
...         self.nastepny_button.config(state=tk.NORMAL)
... 
...     def popraw_dane(self):
...         self.label.config(text="Wybierz pole do poprawy:")
...         # Można by tu dodać listę rozwijaną z polami do wyboru, ale dla prostoty wracamy do pierwszego pola
...         self.indeks_pola = 0
...         self.next_field()
...         self.dane = {} # Resetujemy dane do ponownego wprowadzenia
...         self.dane_wyswietlane.config(text="")
... 
...     def zapisz_i_nowy(self):
...         self.zapisz_do_excela([self.dane[pole] for pole in self.pola])
...         self.dane = {}
...         self.indeks_pola = 0
...         self.next_field()
... 
...     def zapisz_do_excela(self, dane_do_zapisu):
...         try:
...             workbook = openpyxl.load_workbook('dane_osobowe.xlsx')
...             sheet = workbook.active
...         except FileNotFoundError:
...             workbook = openpyxl.Workbook()
...             sheet = workbook.active
...             # Dodaj nagłówki, jeśli to nowy plik
...             for col_num, pole in enumerate(self.pola, 1):
...                 col_letter = get_column_letter(col_num)
...                 sheet[f"{col_letter}1"] = pole
... 
...         sheet.append(dane_do_zapisu)
...         workbook.save('dane_osobowe.xlsx')
...         messagebox.showinfo("Sukces", "Dane zostały zapisane do pliku Excel.")
... 
... root = tk.Tk()
... app = DaneOsoboweApp(root)

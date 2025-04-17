import os
import re
import tkinter as tk
from tkinter import filedialog
from email import policy
from email.parser import BytesParser
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_info_from_eml(eml_path):
    with open(eml_path, 'rb') as fp:
        msg = BytesParser(policy=policy.default).parse(fp)

    subject = msg['subject']
    sender = msg['from']
    print(f"Θέμα: {subject}")
    print(f"Αποστολέας: {sender}")

    html_body = None
    for part in msg.walk():
        content_type = part.get_content_type()
        if content_type == 'text/html':
            html_body = part.get_content()
            break

    if html_body is None:
        print("Δεν βρέθηκε HTML περιεχόμενο.")
        return None, None, None

    soup = BeautifulSoup(html_body, 'html.parser')
    text = soup.get_text(separator="\n")
    return subject, sender, text

def parse_passenger_list(text):
    flight_number = ''
    flight_time = ''
    lines = text.splitlines()

    for i, line in enumerate(lines):
        if 'Flight info' in line:
            for j in range(i+1, min(i+10, len(lines))):
                flight_match = re.search(r'([A-Z]{2,3}\.\d{3,4})', lines[j])
                time_match = re.search(r'(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})', lines[j])
                if flight_match:
                    flight_number = flight_match.group(1)
                if time_match:
                    flight_time = f"{time_match.group(1)} - {time_match.group(2)}"
            break

    route_match = re.search(r'Route\s*:?\s*([A-Z]{3})\s*-\s*([A-Z]{3})', text)
    time_direction = 'DEPARTURE'
    if route_match:
        origin, destination = route_match.groups()
        if origin == 'KGS':
            time_direction = 'DEPARTURE'
        elif destination == 'KGS':
            time_direction = 'ARRIVAL'

    passenger_lines = []
    read = False
    for line in lines:
        if 'No' in line and 'NAAM' in line:
            read = True
            continue
        if read:
            if line.strip() == "" or line.startswith("==="):
                continue
            passenger_lines.append(line)

    combined_lines = []
    buffer = ""
    for line in passenger_lines:
        if re.match(r'^\s*(INF|CHD|#)', line):
            # Αν είναι "INF" ή "CHD", την κρατάμε ξεχωριστά
            if buffer:
                combined_lines.append(buffer)
            combined_lines.append(line.strip())  # Προσθήκη της γραμμής INF/CHD ξεχωριστά
            buffer = ""  # Κενό για να ξεκινήσει μια νέα γραμμή
        else:
            if buffer:
                combined_lines.append(buffer)
            buffer = line.strip()

    if buffer:
        combined_lines.append(buffer)

    data = []
    for line in combined_lines:
        if 'YES' not in line and 'PRV' not in line:
            continue

        head_match = re.match(
            r'(?P<index>\d+)?\.?\s*'
            r'(?P<name>.+?)\s+'
            r'(?P<age>\d+)\s+'
            r'(?P<return>\d{4}-\d{2}-\d{2})\s+'
            r'(?P<resnr>\d+)\s+'
            r'(?P<rest>.+)',
            line
        )
        if not head_match:
            continue

        gd = head_match.groupdict()
        rest = gd['rest'].rstrip()

        try:
            trf_match = re.search(r'(YES|PRV)\s+[A-Z]\s+[A-Z]$', rest)
            if not trf_match:
                continue

            trf_start = trf_match.start()
            trf = trf_match.group(1)

            agent_raw = rest[trf_start - 10:trf_start].strip()
            hotel = rest[:trf_start - 10].strip()

            info = {
                'index': gd['index'],
                'name': gd['name'].strip(),
                'age': gd['age'],
                'return': gd['return'],
                'resnr': gd['resnr'],
                'hotel': hotel,
                'agent': agent_raw,
                'trf': trf,
                'direction': time_direction
            }
            data.append(info)
        except Exception as e:
            print(f"⚠ Σφάλμα σε γραμμή:\n{line}\n{e}")

    return data, flight_number, flight_time

def export_to_excel(data, eml_filename, flight_number="", flight_time="", save_folder=None):
    df = pd.DataFrame(data)
    base_filename = os.path.splitext(eml_filename)[0] + '.xlsx'
    safe_filename = re.sub(r'[\\/*?:"<>|]', "_", base_filename)

    if save_folder:
        output_path = os.path.join(save_folder, safe_filename)
    else:
        output_path = safe_filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Passengers"

    ws['A1'] = "Flight"
    ws['B1'] = flight_number if flight_number else "N/A"
    ws['A2'] = "Time"
    ws['B2'] = flight_time if flight_time else "N/A"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_path)
    print(f"✅ Το Excel αποθηκεύτηκε: {output_path}")

def main():
    root = tk.Tk()
    root.withdraw()

    eml_path = filedialog.askopenfilename(filetypes=[("EML files", "*.eml")])
    if not eml_path:
        print("Δεν επιλέχθηκε αρχείο.")
        return

    print(f"Επιλέξατε το αρχείο: {eml_path}")
    subject, sender, text = extract_info_from_eml(eml_path)
    if not text:
        print("Δεν βρέθηκε περιεχόμενο προς επεξεργασία.")
        return

    data, flight_number, flight_time = parse_passenger_list(text)
    if not data:
        print("Δεν βρέθηκαν επιβάτες με TRF 'YES' ή 'PRV'.")
        return

    save_folder = filedialog.askdirectory(title="Επιλέξτε φάκελο για αποθήκευση Excel")
    if not save_folder:
        print("Δεν επιλέχθηκε φάκελος.")
        return

    export_to_excel(data, os.path.basename(eml_path), flight_number, flight_time, save_folder)

if __name__ == '__main__':
    main()

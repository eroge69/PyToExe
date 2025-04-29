# 在程式開頭添加隱式依賴聲明
import hiddenimports
hiddenimports = ['pandas._libs.tslibs.np_datetime', 'openpyxl.styles.stylesheet']
import os as os
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from streamlit_calendar import calendar
from openpyxl import load_workbook
import threading

FILE_PATH = os.path.join(os.path.dirname(__file__), "TESTFILE.xlsx")

ACTION_TYPE_COLORS = {
    "CL Brand New": "#FEF4C0",
    "CL Top Up":    "#C8F2DC",
    "CMC":          "#E4D9FE",
}

# 在 ACTION_TYPE_COLORS 定義下方新增這兩個函數
def generate_light_color(action_type):
    """用哈希值生成一致性淺色系，確保相同 action_type 總是返回相同顏色"""
    hue = hash(action_type) % 360  # 將哈希值映射到色相環(0-359度)
    return hsl_to_hex(hue, saturation=25, lightness=90)  # 低飽和度+高亮度=淺色

def hsl_to_hex(h, s, l):
    """將 HSL 顏色轉換為十六進位格式"""
    # 公式轉換參考：https://www.rapidtables.com/convert/color/hsl-to-rgb.html
    h /= 360
    s /= 100
    l /= 100
    
    if s == 0:
        r = g = b = l
    else:
        def hue_to_rgb(p, q, t):
            t += 1 if t < 0 else 0
            t -= 1 if t > 1 else 0
            if t < 1/6: return p + (q - p) * 6 * t
            if t < 1/2: return q
            if t < 2/3: return p + (q - p) * (2/3 - t) * 6
            return p
            
        q = l * (1 + s) if l < 0.5 else l + s - l * s
        p = 2 * l - q
        r = hue_to_rgb(p, q, h + 1/3)
        g = hue_to_rgb(p, q, h)
        b = hue_to_rgb(p, q, h - 1/3)
    
    # 將 RGB 值轉換為 HEX 格式
    r = int(round(r * 255))
    g = int(round(g * 255))
    b = int(round(b * 255))
    return "#{:02x}{:02x}{:02x}".format(r, g, b)

# 修改原有的 action_to_event 函數中的 color 獲取方式
def action_to_event(row):
    start = row["開始日期"]
    end = row["結束日期"] + timedelta(days=1)
    type_ = row["客戶類型"]
    
    # 修改這行代碼 ↓
    color = ACTION_TYPE_COLORS.get(type_, generate_light_color(type_))
    
    title = (
        f'{type_}｜{row["所屬部門"]}｜{row["客戶標簽"]}｜{row["獎賞類型"]}｜{row["折扣幅度"]}｜'
        f'{row["發送方法"]}｜發佈人數:{row["發佈人數"]}'
    )
    return {
        "title": title,
        "start": start.strftime("%Y-%m-%d") if pd.notnull(start) else "",
        "end": end.strftime("%Y-%m-%d") if pd.notnull(end) else "",
        "id": str(row.name),
        "color": color,
        "textColor": "black",
        "extendedProps": {
            k: (v.strftime("%Y-%m-%d") if isinstance(v, pd.Timestamp) else v) for k, v in row.items()
        }
    }

OPTION_TYPES = [
    "所屬部門", "PIC", "客戶類型", "客戶標簽", "發送方法", "獎賞類型", "折扣幅度"
]

# ========== 1. 加速與防呆快取 ==========
@st.cache_data
def load_actions():
    try:
        df = pd.read_excel(FILE_PATH, sheet_name="Actions", engine="openpyxl", dtype=str)
        for col in ["執行日期", "開始日期", "結束日期"]:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        df = df.dropna(subset=["執行日期", "開始日期", "結束日期"])
        return df
    except Exception as e:
        st.error(f"載入 Actions 失敗: {e}")
        return pd.DataFrame(columns=["執行日期", "開始日期", "結束日期"] + OPTION_TYPES + ["發佈人數"])

@st.cache_data
def load_options():
    try:
        opt_df = pd.read_excel(FILE_PATH, sheet_name="Options", engine="openpyxl", dtype=str)
        return {
            t: [x for x in opt_df[t].dropna().unique()] if t in opt_df.columns else []
            for t in OPTION_TYPES
        }
    except Exception:
        return {t: [] for t in OPTION_TYPES}

def save_actions(df):
    # 只覆蓋 Actions 分頁
    book = load_workbook(FILE_PATH)
    if "Actions" in book.sheetnames:
        del book["Actions"]
        book.save(FILE_PATH)
    with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel(writer, sheet_name="Actions", index=False)
    st.cache_data.clear()

def save_options(option_dict):
    # 只覆蓋 Options 分頁
    book = load_workbook(FILE_PATH)
    if "Options" in book.sheetnames:
        del book["Options"]
        book.save(FILE_PATH)
    # 補齊欄，空值填充
    max_len = max(len(v) for v in option_dict.values())
    data = {k: v + [""]*(max_len-len(v)) for k, v in option_dict.items()}
    df_opt = pd.DataFrame(data)
    with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df_opt.to_excel(writer, sheet_name="Options", index=False)
    st.cache_data.clear()

def action_to_event(row):
    start = row["開始日期"]
    end = row["結束日期"] + timedelta(days=1)
    type_ = row["客戶類型"]
    color = ACTION_TYPE_COLORS.get(type_, "#cccccc")
    title = (
        f'{type_}｜{row["所屬部門"]}｜{row["客戶標簽"]}｜{row["獎賞類型"]}｜{row["折扣幅度"]}｜'
        f'{row["發送方法"]}｜發佈人數:{row["發佈人數"]}'
    )
    return {
        "title": title,
        "start": start.strftime("%Y-%m-%d") if pd.notnull(start) else "",
        "end": end.strftime("%Y-%m-%d") if pd.notnull(end) else "",
        "id": str(row.name),
        "color": color,
        "textColor": "black",
        "extendedProps": {
            k: (v.strftime("%Y-%m-%d") if isinstance(v, pd.Timestamp) else v) for k, v in row.items()
        }
    }

def main():
    st.set_page_config("CreFIT P&M Calender", layout="wide")
    st.title("CreFIT P&M Calender")
    df = load_actions()
    option_dict = load_options()
    if not any(option_dict.values()):
        for t in OPTION_TYPES:
            option_dict[t] = list(df[t].dropna().unique())

    # 日曆區
    events = [action_to_event(row) for _, row in df.iterrows()]
    col1, col2, col3 = st.columns([1,4,1])
    with col2:
        cal = calendar(
            events=events,
            options={
                "headerToolbar": {
                    "left": "prev,next today",
                    "center": "title",
                    "right": "dayGridMonth"
                },
                "initialView": "dayGridMonth",
                "locale": "zh-tw",
                "height": 900,
            },
            custom_css="""
                .fc {max-width: 1050px !important; margin: auto;}
                .fc-scroller, .fc-daygrid-body {max-height: 890px !important;}
                .fc-event-main { font-size:14px;}
            """
        )
    st.divider()

    # 新增保存按鈕
    st.header("新增保存按鈕")
    with st.form("save_option_form", clear_on_submit=True):
        cols = st.columns(4)
        new_options = {}
        for idx, ot in enumerate(OPTION_TYPES):
            new_options[ot] = cols[idx%4].text_area(
                f"{ot}（每行一個）",
                value="\n".join(option_dict[ot]),
                height=80,
                key=ot+"_edit"
            )
        do_save = st.form_submit_button("保存所有資料選項")
        if do_save:
            true_opt = {k: [i.strip() for i in v.split('\n') if i.strip()] for k,v in new_options.items()}
            save_options(true_opt)
            st.success("所有選項已成功儲存，下次啟動將自動帶入。")

    # 新增行動
    st.header("新增行動")
    with st.form("add_action_form", clear_on_submit=True):
        date_cols = st.columns(3)
        exec_date = date_cols[0].date_input("執行日期", value=datetime.today())
        start_date = date_cols[1].date_input("開始日期")
        end_date = date_cols[2].date_input("結束日期")
        form_cols = st.columns(4)
        form_values = {}
        idx = 0
        for ot in OPTION_TYPES:
            opt_list = option_dict[ot]
            form_values[ot] = form_cols[idx % 4].selectbox(ot, [""] + opt_list, key="add_"+ot)
            idx += 1
        發佈人數 = st.text_input("發佈人數", value="")
        submitted = st.form_submit_button("確認新增")
        if submitted:
            if not start_date or not end_date or any(not v for v in form_values.values()):
                st.error("請填寫所有必填欄位。")
            elif not 發佈人數.isdigit():
                st.error("發佈人數僅能輸入數字")
            else:
                new_row = {
                    "執行日期": pd.to_datetime(exec_date),
                    "開始日期": pd.to_datetime(start_date),
                    "結束日期": pd.to_datetime(end_date),
                    **{k: v for k, v in form_values.items()},
                    "發佈人數": 發佈人數
                }
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                save_actions(df)
                st.success("新增完成！請刷新日曆。")

    # 編輯現有行動
    st.header("🛠 編輯或刪除現有行動")
    if not df.empty:
        df["簡介"] = df.apply(lambda r: f'{r["執行日期"].strftime("%Y-%m-%d")} | {r["客戶類型"]} | {r["所屬部門"]} | {r["客戶標簽"]} | {r["獎賞類型"]}', axis=1)
        action_idx = st.selectbox("請選擇要編輯的行動", df.index, format_func=lambda i: df.at[i, "簡介"])
        row = df.loc[action_idx]
        with st.form(f"edit_existing_action_{action_idx}"):
            date_cols = st.columns(3)
            exec_date = date_cols[0].date_input("執行日期", value=row["執行日期"])
            start_date = date_cols[1].date_input("開始日期", value=row["開始日期"])
            end_date = date_cols[2].date_input("結束日期", value=row["結束日期"])
            form_cols = st.columns(4)
            edit_vals = {}
            idx2 = 0
            for ot in OPTION_TYPES:
                edit_vals[ot] = form_cols[idx2 % 4].text_input(ot, value=str(row[ot]) if pd.notnull(row[ot]) else "")
                idx2 += 1
            發佈人數 = st.text_input("發佈人數", value=str(row["發佈人數"]) if pd.notnull(row["發佈人數"]) else "")
            c1, c2 = st.columns([1,1])
            update = c1.form_submit_button("儲存修改")
            delete = c2.form_submit_button("刪除此行動")
            if update:
                if not 發佈人數.isdigit():
                    st.error("發佈人數僅能輸入數字")
                else:
                    df.at[action_idx, "執行日期"] = pd.to_datetime(exec_date)
                    df.at[action_idx, "開始日期"] = pd.to_datetime(start_date)
                    df.at[action_idx, "結束日期"] = pd.to_datetime(end_date)
                    for k, v in edit_vals.items():
                        df.at[action_idx, k] = v
                    df.at[action_idx, "發佈人數"] = 發佈人數
                    save_actions(df)
                    st.success("行動已更新！請刷新日曆。")
            if delete:
                df.drop(action_idx, inplace=True)
                save_actions(df)
                st.warning("行動已刪除！請刷新日曆。")
    else:
        st.info("目前尚無行動可編輯。")

if __name__ == "__main__":
    main()

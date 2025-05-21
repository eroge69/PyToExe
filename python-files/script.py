import os
import zipfile
import shutil # For removing temporary directories
import openpyxl # Import the openpyxl library for Excel

def rename_files_in_folder(current_folder_path, folder_name):
    """
    Renames specific files in the current folder by appending the folder's name.
    """
    files_to_rename = {
        "etc_group.txt": f"{folder_name}_group.txt",
        "etc_passwd.txt": f"{folder_name}_passwd.txt",
        "etc_sudoers.txt": f"{folder_name}_sudoers.txt"
    }

    print(f"\n--- Starting File Renaming in {folder_name} ---")
    print(f"Processing in folder: {current_folder_path}")
    print("-" * 30)

    for old_filename, new_filename in files_to_rename.items():
        old_filepath = os.path.join(current_folder_path, old_filename)
        new_filepath = os.path.join(current_folder_path, new_filename)

        if os.path.exists(old_filepath):
            try:
                os.rename(old_filepath, new_filepath)
                print(f"Renamed: '{old_filename}' to '{new_filename}'")
            except OSError as e:
                print(f"Error renaming '{old_filename}': {e}")
        else:
            print(f"File not found for renaming: '{old_filename}'")
    print(f"--- File Renaming in {folder_name} Complete ---")

def _process_files_in_directory(directory_path, file_to_exclude):
    """
    Helper function to read and concatenate content from files in a given directory,
    excluding a specific file.
    Returns the concatenated content.
    """
    collected_content = ""
    print(f"  Inspecting directory for aggregation: {directory_path}")
    for item_name in os.listdir(directory_path):
        item_path = os.path.join(directory_path, item_name)
        if os.path.isfile(item_path) and item_name != file_to_exclude:
            try:
                # Using errors='ignore' for robustness, though UTF-8 is preferred for such files
                with open(item_path, 'r', encoding='utf-8', errors='ignore') as f:
                    print(f"    Reading content from: '{item_name}'")
                    collected_content += f.read() + "\n"  # Add a newline between contents
            except Exception as e:
                print(f"    Error reading file '{item_name}': {e}")
        elif os.path.isfile(item_path) and item_name == file_to_exclude:
            print(f"    Skipping excluded file: '{item_name}'")
    return collected_content

def _find_and_process_leaf_dir_recursively(base_search_path, file_to_exclude):
    """
    Recursively searches for a leaf folder (no subdirectories) within base_search_path.
    Processes files in the first suitable leaf folder found that yields content.
    Returns the collected content.
    """
    for dirpath, dirnames, filenames in os.walk(base_search_path):
        # Check if it's a leaf folder (no subdirectories)
        if not dirnames:  # This is a leaf folder
            print(f"  Found leaf folder: {dirpath}")
            # Check if this leaf folder actually contains any files (other than the excluded one)
            has_relevant_files = any(
                os.path.isfile(os.path.join(dirpath, fname)) and fname != file_to_exclude for fname in filenames
            )
            if has_relevant_files:
                content = _process_files_in_directory(dirpath, file_to_exclude)
                if content:
                    return content  # Return content from the first leaf folder with processable files
            else:
                print(f"  Leaf folder {dirpath} has no relevant files to process.")
    return "" # No suitable leaf folder with content found

def aggregate_sudoers_d_files(current_folder_path, folder_name):
    """
    Aggregates sudoers files. Looks for 'sudoers.d', then a zip, then recursively.
    Returns the path to the output file if created, otherwise None.
    """
    print(f"\n--- Starting sudoers.d Aggregation in {folder_name} ---")
    output_filename = f"{folder_name}_sudoers.d"
    output_filepath = os.path.join(current_folder_path, output_filename)
    file_to_exclude = "998_sudo_access"
    collected_content = ""
    processed_source_path = None # To keep track of where the content came from

    print(f"Output file will be: {output_filepath}")
    print(f"Will exclude file: {file_to_exclude} from aggregation.")
    print("-" * 30)

    # 1. Check for "sudoers.d" directory directly
    sudoers_d_direct_path = os.path.join(current_folder_path, "sudoers.d")
    if os.path.isdir(sudoers_d_direct_path):
        print(f"Found 'sudoers.d' directory directly at: {sudoers_d_direct_path}")
        collected_content = _process_files_in_directory(sudoers_d_direct_path, file_to_exclude)
        if collected_content:
            processed_source_path = sudoers_d_direct_path

    # 2. If not found or empty, look for a ZIP file in the current folder
    if not collected_content:
        print(f"\n'sudoers.d' not found directly or was empty. Looking for a ZIP file in '{current_folder_path}'...")
        zip_file_path = None
        for item_name in sorted(os.listdir(current_folder_path)): # sorted for predictable behavior
            if item_name.lower().endswith(".zip"):
                zip_file_path = os.path.join(current_folder_path, item_name)
                print(f"Found ZIP file: {zip_file_path}")
                break  # Use the first ZIP file found

        if zip_file_path:
            # Define a temporary extraction folder name
            extract_to_folder_name = f"{folder_name}_unzipped_sudoers_temp"
            extract_to_path = os.path.join(current_folder_path, extract_to_folder_name)

            if os.path.exists(extract_to_path):
                print(f"  Removing existing temporary extraction folder: {extract_to_path}")
                shutil.rmtree(extract_to_path)

            try:
                os.makedirs(extract_to_path, exist_ok=True)
                print(f"  Extracting '{zip_file_path}' to '{extract_to_path}'...")
                with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_to_path)
                print("  Extraction complete.")

                # 3. Recursively search for a leaf folder in the extracted content
                print(f"\n  Recursively searching for a leaf folder in extracted content: {extract_to_path}")
                collected_content = _find_and_process_leaf_dir_recursively(extract_to_path, file_to_exclude)
                if collected_content:
                    processed_source_path = extract_to_path + " (recursively searched)"

            except zipfile.BadZipFile:
                print(f"  Error: '{zip_file_path}' is not a valid ZIP file or is corrupted.")
            except Exception as e:
                print(f"  An error occurred during ZIP processing or recursive search: {e}")
            finally:
                # Clean up the temp folder
                if os.path.exists(extract_to_path):
                    print(f"  Cleaning up temporary extraction folder: {extract_to_path}")
                    shutil.rmtree(extract_to_path)

        else:
            print("No ZIP file found.")

    # 4. If still no content, perform a general recursive search starting from the current_folder_path
    if not collected_content:
        print(f"\nNo content from 'sudoers.d' or ZIP file. "
              f"Attempting general recursive search for a leaf folder in: {current_folder_path}")
        collected_content = _find_and_process_leaf_dir_recursively(current_folder_path, file_to_exclude)
        if collected_content:
             processed_source_path = current_folder_path + " (recursively searched)"


    # 5. Write the output file
    if collected_content:
        try:
            with open(output_filepath, 'w', encoding='utf-8') as outfile:
                outfile.write(collected_content)
            print(f"\nSuccessfully created/updated '{output_filename}' using content from '{processed_source_path}'.")
            return output_filepath # Return the path if successful
        except Exception as e:
            print(f"Error writing to output file '{output_filename}': {e}")
            return None # Return None if writing failed
    elif processed_source_path and not collected_content:
           print(f"\nProcessed source ('{processed_source_path}'), but it contained no relevant files to aggregate (or only the excluded file). Output file not created/updated.")
           return None
    else:
        print("\nNo suitable files or folders found to process after all checks. Output file not created.")
        return None

def create_excel_report_for_folder(current_folder_path, folder_name):
    """
    Creates an Excel file with sheets for completeness and processed file contents for a single folder.
    Populates 'Group', 'Passwd', 'Sudoers', and 'Sudoers.d' sheets line by line starting from B2, adding folder_name to column A.
    Adds "Server" to cell A1 and "Content" to B1 in these sheets.
    The Excel filename includes the folder name.
    """
    print(f"\n--- Starting Excel Report Generation in {folder_name} ---")
    excel_filename = f"Final Linux Consolidated Sheet_{folder_name}.xlsx" # Excel filename includes folder name
    excel_filepath = os.path.join(current_folder_path, excel_filename)

    try:
        # Create a new workbook
        wb = openpyxl.Workbook()

        # Remove the default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Define sheet names
        sheet_names = ["Completeness", "Group", "Passwd", "Sudoers", "Sudoers.d"]

        # Create sheets and populate with content
        for sheet_name in sheet_names:
            ws = wb.create_sheet(sheet_name)
            file_path = None

            # Add headers to specific sheets
            if sheet_name in ["Group", "Passwd", "Sudoers", "Sudoers.d"]:
                ws['A1'] = "Server"
                ws['B1'] = "Content" # Added header for content column

            if sheet_name == "Completeness":
                # Add a simple completeness message
                ws['A1'] = "OS Consolidation Report Status"
                ws['A2'] = "File renaming and aggregation tasks attempted."
                # More detailed status could be added here based on task outcomes

            elif sheet_name == "Group":
                file_path = os.path.join(current_folder_path, f"{folder_name}_group.txt")
                if file_path and os.path.exists(file_path):
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            # Read file line by line and write to Excel starting from B2
                            row_num = 2 # Start row at 2
                            for line in f:
                                ws[f'A{row_num}'] = folder_name # Write folder_name to column A
                                ws[f'B{row_num}'] = line.strip() # Write line to column B, remove leading/trailing whitespace
                                row_num += 1
                        print(f"  Populated '{sheet_name}' sheet line by line from '{os.path.basename(file_path)}' starting at B2, with folder name in A.")
                    except Exception as e:
                        print(f"  Error reading file '{os.path.basename(file_path)}' for Excel sheet '{sheet_name}': {e}")
                        ws['B2'] = f"Error reading file: {os.path.basename(file_path)}"
                else:
                     print(f"  File not found for sheet '{sheet_name}': '{os.path.basename(file_path)}'. Sheet will be empty or contain error message.")
                     ws['B2'] = f"Source file not found: {os.path.basename(file_path)}"

            elif sheet_name == "Passwd":
                file_path = os.path.join(current_folder_path, f"{folder_name}_passwd.txt")
                if file_path and os.path.exists(file_path):
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            # Read file line by line and write to Excel starting from B2
                            row_num = 2 # Start row at 2
                            for line in f:
                                ws[f'A{row_num}'] = folder_name # Write folder_name to column A
                                ws[f'B{row_num}'] = line.strip() # Write line to column B, remove leading/trailing whitespace
                                row_num += 1
                        print(f"  Populated '{sheet_name}' sheet line by line from '{os.path.basename(file_path)}' starting at B2, with folder name in A.")
                    except Exception as e:
                        print(f"  Error reading file '{os.path.basename(file_path)}' for Excel sheet '{sheet_name}': {e}")
                        ws['B2'] = f"Error reading file: {os.path.basename(file_path)}"
                else:
                     print(f"  File not found for sheet '{sheet_name}': '{os.path.basename(file_path)}'. Sheet will be empty or contain error message.")
                     ws['B2'] = f"Source file not found: {os.path.basename(file_path)}"


            elif sheet_name == "Sudoers":
                file_path = os.path.join(current_folder_path, f"{folder_name}_sudoers.txt")
                if file_path and os.path.exists(file_path):
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            # Read file line by line and write to Excel starting from B2
                            row_num = 2 # Start row at 2
                            for line in f:
                                ws[f'A{row_num}'] = folder_name # Write folder_name to column A
                                ws[f'B{row_num}'] = line.strip() # Write line to column B, remove leading/trailing whitespace
                                row_num += 1
                        print(f"  Populated '{sheet_name}' sheet line by line from '{os.path.basename(file_path)}' starting at B2, with folder name in A.")
                    except Exception as e:
                        print(f"  Error reading file '{os.path.basename(file_path)}' for Excel sheet '{sheet_name}': {e}")
                        ws['B2'] = f"Error reading file: {os.path.basename(file_path)}"
                else:
                     print(f"  File not found for sheet '{sheet_name}': '{os.path.basename(file_path)}'. Sheet will be empty or contain error message.")
                     ws['B2'] = f"Source file not found: {os.path.basename(file_path)}"

            elif sheet_name == "Sudoers.d":
                file_path = os.path.join(current_folder_path, f"{folder_name}_sudoers.d")
                if file_path and os.path.exists(file_path):
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            # Read file line by line and write to Excel starting from B2
                            row_num = 2 # Start row at 2
                            for line in f:
                                # Keep empty lines as requested
                                ws[f'A{row_num}'] = folder_name # Write folder_name to column A
                                ws[f'B{row_num}'] = line.strip() # Write line to column B, remove leading/trailing whitespace
                                row_num += 1
                        print(f"  Populated '{sheet_name}' sheet line by line from '{os.path.basename(file_path)}' starting at B2, with folder name in A (including empty lines).")
                    except Exception as e:
                        print(f"  Error reading file '{os.path.basename(file_path)}' for Excel sheet '{sheet_name}': {e}")
                        ws['B2'] = f"Error reading file: {os.path.basename(file_path)}"
                else:
                     print(f"  File not found for sheet '{sheet_name}': '{os.path.basename(file_path)}'. Sheet will be empty or contain error message.")
                     ws['B2'] = f"Source file not found: {os.path.basename(file_path)}"


        # Save the workbook
        wb.save(excel_filepath)
        print(f"\nSuccessfully created Excel report: '{excel_filename}'")

    except Exception as e:
        print(f"\nError creating Excel report '{excel_filename}': {e}")

    print(f"--- Excel Report Generation in {folder_name} Complete ---")


def merge_excel_files(parent_directory):
    """
    Merges data from individual Excel files in subdirectories into a single consolidated Excel file.
    """
    print(f"\n--- Starting Excel File Merging ---")
    consolidated_excel_filename = "Final Linux Consolidated Sheet.xlsx"
    consolidated_excel_filepath = os.path.join(parent_directory, consolidated_excel_filename)

    try:
        # Create the consolidated workbook
        consolidated_wb = openpyxl.Workbook()

        # Remove the default sheet
        if 'Sheet' in consolidated_wb.sheetnames:
            consolidated_wb.remove(consolidated_wb['Sheet'])

        # Define sheet names for merging
        merge_sheet_names = ["Group", "Passwd", "Sudoers", "Sudoers.d"]

        # Create sheets in the consolidated workbook and add headers
        for sheet_name in merge_sheet_names:
            ws = consolidated_wb.create_sheet(sheet_name)
            ws['A1'] = "Server"
            ws['B1'] = "Content" # Header for content column

        # Add the Completeness sheet
        completeness_ws = consolidated_wb.create_sheet("Completeness")
        completeness_ws['A1'] = "OS Consolidation Report Status"
        completeness_ws['A2'] = "Merging of subdirectory reports attempted."
        # Could add more detailed status here if needed

        print(f"Created consolidated workbook structure: '{consolidated_excel_filename}'")

        # Iterate through subdirectories to find and merge Excel files
        for item_name in os.listdir(parent_directory):
            item_path = os.path.join(parent_directory, item_name)

            if os.path.isdir(item_path):
                subdirectory_path = item_path
                subdirectory_name = item_name
                individual_excel_filename = f"Final Linux Consolidated Sheet_{subdirectory_name}.xlsx"
                individual_excel_filepath = os.path.join(subdirectory_path, individual_excel_filename)

                if os.path.exists(individual_excel_filepath):
                    print(f"  Merging data from: '{individual_excel_filepath}'")
                    try:
                        individual_wb = openpyxl.load_workbook(individual_excel_filepath)

                        for sheet_name in merge_sheet_names:
                            if sheet_name in individual_wb.sheetnames:
                                individual_ws = individual_wb[sheet_name]
                                consolidated_ws = consolidated_wb[sheet_name]

                                # Append rows from the individual sheet (starting from row 2)
                                # to the consolidated sheet
                                for row in individual_ws.iter_rows(min_row=2): # Start from the second row
                                    row_data = [cell.value for cell in row]
                                    consolidated_ws.append(row_data) # Append the row data

                                print(f"    Merged '{sheet_name}' sheet.")

                    except Exception as e:
                        print(f"  Error merging data from '{individual_excel_filename}': {e}")
                else:
                    print(f"  Individual Excel file not found in '{subdirectory_name}': '{individual_excel_filename}'")

        # Save the consolidated workbook
        consolidated_wb.save(consolidated_excel_filepath)
        print(f"\nSuccessfully created consolidated Excel report: '{consolidated_excel_filename}'")

    except Exception as e:
        print(f"\nError creating consolidated Excel report '{consolidated_excel_filename}': {e}")

    print("--- Excel File Merging Complete ---")


if __name__ == "__main__":
    parent_directory = os.getcwd()
    print(f"Script running from parent directory: {parent_directory}")

    # --- Step 1 & 2 & 3: Process each subdirectory ---
    for item_name in os.listdir(parent_directory):
        item_path = os.path.join(parent_directory, item_name)

        # Check if the item is a directory
        if os.path.isdir(item_path):
            subdirectory_path = item_path
            subdirectory_name = item_name

            print(f"\n=== Processing Subdirectory: {subdirectory_name} ===")

            try:
                # Change the current working directory to the subdirectory
                os.chdir(subdirectory_path)

                # The folder_name for processing is simply the subdirectory name
                folder_name_for_processing = subdirectory_name

                # Step 1: Rename files in the current subdirectory
                rename_files_in_folder(subdirectory_path, folder_name_for_processing)

                # Step 2: Aggregate sudoers.d files in the current subdirectory
                aggregated_sudoers_d_path = aggregate_sudoers_d_files(subdirectory_path, folder_name_for_processing)

                # Step 3: Create Excel report in the current subdirectory
                # This now creates an Excel file named Final_Linux_Consolidated_Sheet_<folder_name>.xlsx
                create_excel_report_for_folder(subdirectory_path, folder_name_for_processing)

            except Exception as e:
                print(f"An error occurred while processing subdirectory '{subdirectory_name}': {e}")
            finally:
                # Change back to the parent directory before processing the next subdirectory
                os.chdir(parent_directory)
                print(f"=== Finished Processing Subdirectory: {subdirectory_name} ===")

    # --- Step 4: Merge the individual Excel files ---
    merge_excel_files(parent_directory)

    print("\nAll tasks finished.")


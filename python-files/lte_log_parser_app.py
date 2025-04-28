import re
import os
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def convert_timestamp_to_12hour(timestamp_24h):
    try:
        # Parse the 24-hour format time
        time_obj = datetime.strptime(timestamp_24h, "%H:%M:%S.%f")
        # Convert to 12-hour format
        return time_obj.strftime("%I:%M:%S %p").lstrip("0")
    except:
        return timestamp_24h  # Return original if conversion fails

def parse_log_file(file_path, target_messages):
    try:
        with open(file_path, "r") as file:
            log_content = file.read()
    except FileNotFoundError:
        messagebox.showerror("Error", f"{file_path} file not found!")
        return None
    except Exception as e:
        messagebox.showerror("Error", f"Error reading {file_path}: {e}")
        return None
    
    # Split the log content by message types
    # Look for message headers which typically start with a date and have a message type
    message_pattern = r'(\d{4}\s\w{3}\s\d{2}\s+\d{2}:\d{2}:\d{2}\.\d{3}\s+\[\w+\]\s+\w+\s+(.+?)(?=\d{4}\s\w{3}\s\d{2}\s+\d{2}:\d{2}:\d{2}\.\d{3}|\Z))'
    messages = re.findall(message_pattern, log_content, re.DOTALL)
    
    if not messages:
        messagebox.showwarning("Warning", "No valid messages found in the log file.")
        return None
    
    grouped_messages = {}
    
    for full_message, message_type in messages:
        # Extract message type name (clean up the string)
        message_name = message_type.strip().split('\n')[0]
        
        # Skip messages that are not in the target list
        if message_name not in target_messages:
            continue
        
        # Extract timestamp
        timestamp_match = re.search(r'(\d{2}:\d{2}:\d{2}\.\d{3})', full_message)
        timestamp_24h = timestamp_match.group(1) if timestamp_match else "Unknown"
        timestamp_12h = convert_timestamp_to_12hour(timestamp_24h)
        
        # Parse table data
        table_data = parse_table(full_message, timestamp_12h, message_name)
        
        if table_data:
            # Initialize message type entry if not exists
            if message_name not in grouped_messages:
                grouped_messages[message_name] = {
                    'headers': table_data['headers'],
                    'data': []
                }
            
            # Append data rows to the existing message type
            grouped_messages[message_name]['data'].extend(table_data['data'])
    
    return grouped_messages

def parse_table(message_content, timestamp, message_name):
    lines = [line.strip() for line in message_content.split('\n') if line.strip()]
    
    # For PUSCH Tx Report, we need special processing
    if message_name == "LTE LL1 PUSCH Tx Report":
        return parse_pusch_tx_report(lines, timestamp)
    
    # Default table parsing for other message types
    header_rows = []
    data_rows = []
    separator_count = 0
    in_table = False
    
    for line in lines:
        # Check for separator lines
        if line.startswith('-'):
            separator_count += 1
            if separator_count == 1:
                in_table = True
            continue
        
        # Only process table data
        if in_table:
            # Header rows
            if separator_count == 1 and line.startswith('|'):
                cells = [cell.strip() for cell in line.split('|') if cell]
                header_rows.append(cells)
            # Data rows
            elif separator_count >= 2 and line.startswith('|'):
                cells = [cell.strip() for cell in line.split('|') if cell]
                # Insert timestamp as first column in data rows
                data_rows.append([timestamp] + cells)
    
    if not header_rows or not data_rows:
        return None
    
    # Determine the number of columns needed for headers
    max_header_cols = max(len(row) for row in header_rows) if header_rows else 0
    
    # Initialize the combined headers
    combined_headers = [""] * max_header_cols
    
    # Process each header row and combine vertically
    for row in header_rows:
        for i, cell in enumerate(row):
            if i < max_header_cols:
                if combined_headers[i]:
                    combined_headers[i] += " " + cell
                else:
                    combined_headers[i] = cell
    
    # Create final headers with timestamp as first column
    final_headers = ["Timestamp"] + combined_headers
    
    return {
        'headers': final_headers,
        'data': data_rows
    }

def parse_pusch_tx_report(lines, timestamp):
    """Parse PUSCH Tx Report with special handling for payload sections."""
    header_rows = []
    data_rows = []
    separator_count = 0
    in_table = False
    
    # First pass: extract table structure normally
    for line in lines:
        # Check for separator lines
        if line.startswith('-'):
            separator_count += 1
            if separator_count == 1:
                in_table = True
            continue
        
        # Only process table data
        if in_table:
            # Header rows
            if separator_count == 1 and line.startswith('|'):
                cells = [cell.strip() for cell in line.split('|') if cell]
                header_rows.append(cells)
            # Data rows
            elif separator_count >= 2 and line.startswith('|'):
                cells = [cell.strip() for cell in line.split('|') if cell]
                data_rows.append(cells)  # Don't add timestamp yet
    
    if not header_rows or not data_rows:
        return None
    
    # Determine the number of columns needed for headers
    max_header_cols = max(len(row) for row in header_rows) if header_rows else 0
    
    # Initialize the combined headers
    combined_headers = [""] * max_header_cols
    
    # Process each header row and combine vertically
    for row in header_rows:
        for i, cell in enumerate(row):
            if i < max_header_cols:
                if combined_headers[i]:
                    combined_headers[i] += " " + cell
                else:
                    combined_headers[i] = cell
    
    # Find indices of ACK and CQI payload columns
    ack_index = -1
    cqi_index = -1
    for i, header in enumerate(combined_headers):
        if "ACK Payload" in header:
            ack_index = i
        elif "CQI Payload" in header:
            cqi_index = i
    
    # Process data rows for payload merging
    merged_rows = []
    i = 0
    while i < len(data_rows):
        current_row = data_rows[i]
        
        # Create a row template with timestamp
        merged_row = [timestamp] + [""] * len(combined_headers)
        
        # Copy non-payload data
        for j, cell in enumerate(current_row):
            if j < len(combined_headers):
                merged_row[j+1] = cell  # +1 because we added timestamp
        
        # For ACK Payload - Find and collect all payload data
        if ack_index >= 0:
            ack_payload_data = []
            next_idx = i + 1
            
            # Look for ACK Payload header in current row
            if ack_index < len(current_row) and "ACK Payload" in current_row[ack_index]:
                # Skip the header row and find all payload data rows
                while next_idx < len(data_rows):
                    next_row = data_rows[next_idx]
                    
                    # Check if this is a payload data row (not a header or next main row)
                    if any(h in next_row[0] if len(next_row) > 0 else False for h in ["Payload", "Config"]):
                        break
                        
                    # Collect all values in this row as part of the payload
                    payload_values = " ".join(cell for cell in next_row if cell.strip())
                    if payload_values.strip():
                        ack_payload_data.append(payload_values)
                    
                    next_idx += 1
            
            # Combine all payload data into a single cell
            if ack_payload_data:
                merged_row[ack_index+1] = " ".join(ack_payload_data)
        
        # For CQI Payload - Similar logic
        if cqi_index >= 0:
            cqi_payload_data = []
            next_idx = i + 1
            
            # Find CQI Payload section
            found_cqi_header = False
            while next_idx < len(data_rows):
                next_row = data_rows[next_idx]
                
                # Check if we've found the CQI Payload header
                if len(next_row) > cqi_index and "CQI Payload" in next_row[cqi_index]:
                    found_cqi_header = True
                    next_idx += 1  # Skip the header row
                    continue
                
                # If we've found the header, collect payload data
                if found_cqi_header:
                    # Stop if we hit another header or main section
                    if any(h in next_row[0] if len(next_row) > 0 else False for h in ["Payload", "Config"]):
                        break
                    
                    # Collect payload data
                    payload_values = " ".join(cell for cell in next_row if cell.strip())
                    if payload_values.strip():
                        cqi_payload_data.append(payload_values)
                
                next_idx += 1
                
                # If we haven't found a CQI header yet and we hit another main section, stop looking
                if not found_cqi_header and next_idx < len(data_rows) and len(next_row) > 0:
                    if "Config" in next_row[0] or (len(next_row) > ack_index and "ACK Payload" in next_row[ack_index]):
                        break
            
            # Combine all CQI payload data into a single cell
            if cqi_payload_data:
                merged_row[cqi_index+1] = " ".join(cqi_payload_data)
        
        merged_rows.append(merged_row)
        
        # Skip to the next main row
        i += 1
        while i < len(data_rows) and (
            (ack_index >= 0 and i < len(data_rows) and ack_index < len(data_rows[i]) and 
             "ACK Payload" in data_rows[i][ack_index]) or
            (cqi_index >= 0 and i < len(data_rows) and 
             any("CQI Payload" in row[cqi_index] if cqi_index < len(row) else False 
                 for row in data_rows[i:i+3]))
        ):
            i += 1
    
    # Create final headers with timestamp as first column
    final_headers = ["Timestamp"] + combined_headers
    
    return {
        'headers': final_headers,
        'data': merged_rows
    }

def export_to_excel(grouped_messages, output_file):
    try:
        # Create a new Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for message_name, table_data in grouped_messages.items():
                # Convert data to pandas DataFrame
                df = pd.DataFrame(table_data['data'], columns=table_data['headers'])
                
                # Create safe sheet name (Excel limits sheet names to 31 chars)
                sheet_name = message_name
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:28] + "..."
                
                # Write DataFrame to Excel
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get workbook and sheet
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply autofilter to the header row
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Format header row
                header_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
                header_font = Font(bold=True)
                
                # Calculate column width and format headers
                for col_idx, column in enumerate(df.columns, 1):
                    column_letter = get_column_letter(col_idx)
                    
                    # Format header cell
                    cell = worksheet[f"{column_letter}1"]
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    
                    # Calculate width based on content
                    max_length = 0
                    
                    # Check header length
                    header_length = len(str(column))
                    if header_length > max_length:
                        max_length = header_length
                    
                    # Check content length
                    for content in df[column]:
                        content_length = len(str(content)) if content else 0
                        if content_length > max_length:
                            max_length = min(content_length, 50)  # Cap at 50 to avoid extremely wide columns
                    
                    # Set column width with some padding (1.2 multiplier)
                    adjusted_width = max_length * 1.2
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
        return True
    except Exception as e:
        messagebox.showerror("Error", f"Error creating Excel file: {e}")
        return False

class LTELogProcessor(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("LTE Log Processor")
        self.geometry("600x500")
        
        # List of message types to extract
        self.target_messages = [
            "LTE LL1 PUSCH Tx Report",
            "LTE LL1 PCFICH Decoding Results",
            "LTE RLC DL Statistics",
            "LTE LL1 PDSCH Demapper Configuration",
            "LTE DCI Information Report",
            "LTE LL1 PHICH Decoding Results",
            "LTE PDSCH Stat Indication",
            "LTE PUSCH Power Control",
            "LTE PUCCH Power Control"
        ]
        
        self.create_widgets()
    
    def create_widgets(self):
        # Input frame
        input_frame = ttk.LabelFrame(self, text="Input File")
        input_frame.pack(fill="x", padx=10, pady=10)
        
        self.file_path_var = tk.StringVar()
        self.file_path_entry = ttk.Entry(input_frame, textvariable=self.file_path_var, width=50)
        self.file_path_entry.pack(side="left", padx=5, pady=10, fill="x", expand=True)
        
        browse_btn = ttk.Button(input_frame, text="Browse", command=self.browse_file)
        browse_btn.pack(side="right", padx=5, pady=10)
        
        # Output frame
        output_frame = ttk.LabelFrame(self, text="Output File")
        output_frame.pack(fill="x", padx=10, pady=10)
        
        self.output_path_var = tk.StringVar()
        self.output_path_entry = ttk.Entry(output_frame, textvariable=self.output_path_var, width=50)
        self.output_path_entry.pack(side="left", padx=5, pady=10, fill="x", expand=True)
        
        save_btn = ttk.Button(output_frame, text="Browse", command=self.browse_save_location)
        save_btn.pack(side="right", padx=5, pady=10)
        
        # Message types frame
        message_frame = ttk.LabelFrame(self, text="Message Types")
        message_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Create a scrollable frame for message checkboxes
        canvas = tk.Canvas(message_frame)
        scrollbar = ttk.Scrollbar(message_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Add message type checkboxes
        self.message_vars = {}
        for message in self.target_messages:
            var = tk.BooleanVar(value=True)
            self.message_vars[message] = var
            cb = ttk.Checkbutton(scrollable_frame, text=message, variable=var)
            cb.pack(anchor="w", padx=5, pady=2)
        
        # Select/Deselect All buttons
        select_frame = ttk.Frame(self)
        select_frame.pack(fill="x", padx=10, pady=5)
        
        select_all_btn = ttk.Button(select_frame, text="Select All", command=self.select_all)
        select_all_btn.pack(side="left", padx=5)
        
        deselect_all_btn = ttk.Button(select_frame, text="Deselect All", command=self.deselect_all)
        deselect_all_btn.pack(side="left", padx=5)
        
        # Process button
        process_btn = ttk.Button(self, text="Process Log", command=self.process_log)
        process_btn.pack(pady=10)
        
        # Status bar
        self.status_var = tk.StringVar()
        status_bar = ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.status_var.set("Ready")
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Log File",
            filetypes=[("Text Files", "*.txt"), ("Log Files", "*.log"), ("All Files", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            # Set default output filename based on input
            base_name = os.path.basename(file_path)
            name_without_ext = os.path.splitext(base_name)[0]
            default_output = os.path.join(os.path.dirname(file_path), f"{name_without_ext}_output.xlsx")
            self.output_path_var.set(default_output)
    
    def browse_save_location(self):
        file_path = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if file_path:
            self.output_path_var.set(file_path)
    
    def select_all(self):
        for var in self.message_vars.values():
            var.set(True)
    
    def deselect_all(self):
        for var in self.message_vars.values():
            var.set(False)
    
    def process_log(self):
        input_file = self.file_path_var.get()
        output_file = self.output_path_var.get()
        
        if not input_file:
            messagebox.showerror("Error", "Please select an input log file.")
            return
        
        if not output_file:
            messagebox.showerror("Error", "Please specify an output Excel file.")
            return
        
        # Get selected message types
        selected_messages = [msg for msg, var in self.message_vars.items() if var.get()]
        
        if not selected_messages:
            messagebox.showerror("Error", "Please select at least one message type.")
            return
        
        self.status_var.set("Processing log file...")
        self.update_idletasks()
        
        # Parse log file
        grouped_messages = parse_log_file(input_file, selected_messages)
        
        if not grouped_messages:
            self.status_var.set("No valid data found in log file.")
            return
        
        # Export to Excel
        self.status_var.set("Exporting to Excel...")
        self.update_idletasks()
        
        success = export_to_excel(grouped_messages, output_file)
        
        if success:
            self.status_var.set(f"Successfully exported {len(grouped_messages)} message types to Excel.")
            messagebox.showinfo("Success", f"Data exported to {output_file}")
        else:
            self.status_var.set("Failed to export data.")

def main():
    app = LTELogProcessor()
    app.mainloop()

if __name__ == "__main__":
    main()
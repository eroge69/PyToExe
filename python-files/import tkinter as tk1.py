import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
 
def save_data():
    data = {
        "Creation Date": creation_date.get(),
        "Employee Name": employee_name.get(),
        "Vendor Name": vendor_name.get(),
        "agent aame": agent_name.get(),
        "Mobile Number": mobile_number.get(),
        "Width": width.get(),
        "Quality": quality.get(),
        "Rate": rate.get(),
        "Remark": remark.get(),
        
    }
 
    file_path = "C:/Users/Kapil Jodhani/Downloads/11/sampling_data.xlsx"
 
    # Check if file exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Add header row
        ws.append(list(data.keys()))
 
    # Append data row
    ws.append(list(data.values()))
    wb.save(file_path)
 
    messagebox.showinfo("Saved", f"Data saved to {file_path} successfully!")
 
# Create window
root = tk.Tk()
root.title("Sampling Master Form")
root.geometry("400x500")
 
# Variables
creation_date = tk.StringVar(value=datetime.today().strftime('%Y-%m-%d'))
employee_name = tk.StringVar()
vendor_name = tk.StringVar()
agent_name = tk.StringVar()
mobile_number = tk.StringVar()
width = tk.StringVar()
quality = tk.StringVar()
rate = tk.StringVar()
remark = tk.StringVar()

 
# Input validation for integer fields
def only_int(char):
    return char.isdigit()
 
vcmd = (root.register(only_int), '%S')
 
# Form fields
fields = [
    ("Creation Date", creation_date),
    ("Employee Name", employee_name),
    ("Vendor Name", vendor_name),
    ("agent name", agent_name),
    ("Mobile Number", mobile_number),
    ("Width", width),
    ("Quality", quality),
    ("Rate", rate),
    ("Remark", remark),
]
 
for idx, (label, var) in enumerate(fields):
    tk.Label(root, text=label).grid(row=idx, column=0, pady=5, padx=10, sticky="w")
    if label in ["Mobile Number", "Width", "Rate"]:  # Apply number-only validation
        tk.Entry(root, textvariable=var, validate='key', validatecommand=vcmd, width=30).grid(row=idx, column=1, pady=5)
    else:
        tk.Entry(root, textvariable=var, width=30).grid(row=idx, column=1, pady=5)
 
# Dropdown option
# tk.Label(root, text="Option").grid(row=len(fields), column=0, pady=5, padx=10, sticky="w")
# ttk.Combobox(root, textvariable=option, values=["Option 1", "Option 2", "Option 3"], width=28).grid(row=len(fields), column=1, pady=5)
 
# Save button
tk.Button(root, text="Save", command=save_data, bg="green", fg="white", width=15).grid(row=len(fields)+1, column=0, columnspan=2, pady=20)
 
root.mainloop()
 
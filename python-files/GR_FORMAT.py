# Re-execute the GR FORMAT processing after environment reset

import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define paths
source_file = "/mnt/data/MANIFEST SKG 111-01018220  KARAYOLU  22 ADM 172.xlsx"
output_e_arsiv = "/mnt/data/MANIFEST SKG 111-01018220 e-arsiv.xlsx"
output_gr_format = "/mnt/data/MANIFEST SKG 111-01018220 GR FORMAT.xlsx"
output_awb_list = "/mnt/data/atr_awb_list.txt"

# Load workbook
wb = load_workbook(source_file)
ws = wb.active

# Extract headers and their column indexes
headers = [cell.value for cell in ws[1]]
def get_col_idx(name): return headers.index(name) + 1 if name in headers else None
ref2_idx = get_col_idx("Ref.2")
customs_value_idx = get_col_idx("Customs Value")
awb_no_idx = get_col_idx("Awb No")
hs_code_idx = get_col_idx("HS CODE")

# Define color fills
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# List to collect ATR-requiring AWBs
atr_required_awbs = []

# Process rows
for row in range(2, ws.max_row + 1):
    if ref2_idx:
        ws.cell(row=row, column=ref2_idx).fill = green_fill
    if customs_value_idx:
        cell = ws.cell(row=row, column=customs_value_idx)
        cell.fill = yellow_fill
        try:
            val = float(str(cell.value).replace(",", "."))
            if val >= 150:
                awb = ws.cell(row=row, column=awb_no_idx).value
                if awb:
                    atr_required_awbs.append(str(awb))
        except:
            pass

    hs_cell = ws.cell(row=row, column=hs_code_idx)
    if hs_cell.value in [None, ""]:
        hs_cell.value = "620342110000"

# Add IOSS at the end
last_col = ws.max_column + 1
ws.cell(row=1, column=last_col).value = "IOSS"
for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=last_col).value = "IM5286025168"

# Remove unnecessary columns
cols_to_remove = [
    'Gumruk Beyan Tutar', 'Hacim Kg', 'Shipper EORI', 'Latitude', 'Cari Kod', 'Yükseklik',
    'Yabanci Fatura No', 'ATR No', 'Fatura', 'Servis', 'NOT', 'Longitude', 'Shipper VAT ID',
    'Platform', 'Boy', 'Mikro İhracat', 'En', 'Yabanci Fatura Tarihi', 'ATR Tarihi'
]
remove_indexes = [i + 1 for i, h in enumerate(headers) if h in cols_to_remove]
for idx in sorted(remove_indexes, reverse=True):
    ws.delete_cols(idx)

# Save outputs
wb.save(output_gr_format)
shutil.copy(source_file, output_e_arsiv)
with open(output_awb_list, "w") as f:
    for awb in sorted(set(atr_required_awbs)):
        f.write(f"{awb}\n")

output_gr_format, output_e_arsiv, output_awb_list

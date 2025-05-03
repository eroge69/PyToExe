#!/usr/bin/env python
# coding: utf-8

# In[1]:


import tkinter as tk
from tkinter import Listbox, Scrollbar, ttk,font, messagebox
from openpyxl import load_workbook, Workbook
import tkinter.font as font
import os


# In[2]:


import tkinter as tk
from tkinter import ttk, font, messagebox
from openpyxl import load_workbook, Workbook
import os

# --- Functions ---

def load_column_data(col_index):
    workbook = load_workbook("exhibit.xlsx")
    sheet = workbook.active
    values = [str(row[0]) for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index, values_only=True) if row[0]]
    return list(set(values))

def load_last_year_data(filter_name):
    wb = load_workbook("exhibit.xlsx")
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        company_exhibit = str(row[1]).strip()
        name = row[4]
        position = row[5]
        product_name = row[10]
        
        if filter_name.lower() in company_exhibit.lower():
            data.append((company_exhibit, name, position, product_name))
    return data

def create_empty_table():
    if hasattr(root, 'last_year_frame'):
        root.last_year_frame.destroy()

    root.last_year_frame = tk.Frame(root)
    root.last_year_frame.pack(side="right", anchor="n", padx=10, pady=10)

    tk.Label(root.last_year_frame, text="Last Year Data", font=bold_font).pack(pady=(0, 5))

    tree = ttk.Treeview(root.last_year_frame, columns=("Company Exhibit", "Name", "Position", "Product Name"), show="headings")
    tree.pack(fill="both", expand=True)

    for col in ("Company Exhibit", "Name", "Position", "Product Name"):
        tree.heading(col, text=col)
        tree.column(col, anchor="center")

    return tree

def show_last_year_table():
    filter_text = entry1.get().strip()
    data = load_last_year_data(filter_text)
    for item in tree.get_children():
        tree.delete(item)
    for row in data:
        tree.insert("", "end", values=row)

class SuggestionBox:
    def __init__(self, entry_widget, data_source):
        self.entry = entry_widget
        self.data = data_source

        self.popup = tk.Toplevel(root)
        self.popup.overrideredirect(True)
        self.popup.withdraw()

        self.listbox = tk.Listbox(self.popup, width=50)
        self.listbox.pack()
        self.listbox.bind("<<ListboxSelect>>", self.select_suggestion)
        self.listbox.bind("<Return>", self.select_suggestion)
        self.listbox.bind("<Up>", self.on_listbox_key)
        self.listbox.bind("<Down>", lambda e: "break")

        self.entry.bind("<KeyRelease>", self.on_key_release)
        self.entry.bind("<Down>", self.on_arrow_key)

    def on_key_release(self, event):
        text = self.entry.get().lower()
        if not text:
            self.hide()
            return
        matches = [item for item in self.data if text in item.lower()]
        if matches:
            self.show(matches)
        else:
            self.hide()

    def show(self, matches):
        self.popup.deiconify()
        self.popup.geometry(f"+{self.entry.winfo_rootx()}+{self.entry.winfo_rooty() + self.entry.winfo_height()}")
        self.listbox.delete(0, tk.END)
        for item in matches:
            self.listbox.insert(tk.END, item)

    def hide(self):
        self.popup.withdraw()

    def select_suggestion(self, event=None):
        try:
            selected = self.listbox.get(self.listbox.curselection())
            self.entry.delete(0, tk.END)
            self.entry.insert(0, selected)
            self.hide()
        except:
            pass

    def on_arrow_key(self, event):
        if event.keysym == "Down":
            self.listbox.focus_set()
            self.listbox.selection_set(0)
            return "break"

    def on_listbox_key(self, event):
        if event.keysym == "Return":
            self.select_suggestion()
            self.entry.focus_set()
        elif event.keysym == "Up" and self.listbox.curselection()[0] == 0:
            self.entry.focus_set()
            self.listbox.selection_clear(0)

# --- GUI Setup ---
root = tk.Tk()
root.title("Product Entry Form with Autocomplete")
bold_font = font.Font(size=12, weight="bold")

# --- Top Search Field and Button ---
search_frame = tk.Frame(root)
search_frame.pack(anchor="w", padx=10, pady=(10, 0))

tk.Label(search_frame, text="Company Name").pack(side="left", padx=(0, 5))
entry1 = tk.Entry(search_frame, width=50)
entry1.pack(side="left", padx=(0, 10))

search_button = tk.Button(search_frame, text="Search Last Year", command=show_last_year_table)
search_button.pack(side="left")

# Load column data for company names
company_names_from_excel = load_column_data(2)
box1 = SuggestionBox(entry1, company_names_from_excel)

tree = create_empty_table()

# --- Contact Person Section ---
contact_frame = tk.LabelFrame(root, text="Contact Person", padx=10, pady=10, bd=1, relief="solid", font=bold_font)
contact_frame.pack(padx=10, pady=5)

positions = ["Management", "Production", "Purchase", "R&D", "Unknown"]
contact_entries = []

for i in range(4):
    tk.Label(contact_frame, text=f"Person {i + 1}:").grid(row=i, column=0, sticky="w", padx=5, pady=5)
    person_entry = tk.Entry(contact_frame, width=25)
    person_entry.grid(row=i, column=1, padx=5, pady=5)
    position_var = tk.StringVar()
    position_var.set(positions[4])
    position_dropdown = tk.OptionMenu(contact_frame, position_var, *positions)
    position_dropdown.grid(row=i, column=2, padx=5, pady=5)
    contact_entries.append((person_entry, position_var))

# --- Product Frame ---
product_frame = tk.LabelFrame(root, text="Products", padx=10, pady=10, bd=1, relief="solid", font=bold_font)
product_frame.pack(padx=10, pady=5)

product_data = load_column_data(11)
suggestion_boxes = []

industry_options = ["Bakery", "Beverage", "Confectionary", "Dairy", "Infant", "Instant", "Meat",
                    "Oil", "Pharma", "Sauce", "Supplement", "Other"]
industry_vars = []
product_widgets = []

def on_industry_change(index):
    def callback(*args):
        if not root.winfo_exists():
            return
        try:
            new_value = industry_vars[index].get()
            for j in range(index + 1, len(industry_vars)):
                industry_vars[j].set(new_value)
        except tk.TclError:
            pass
    return callback

for i in range(10):
    tk.Label(product_frame, text=f"Product {i + 1}:").grid(row=i, column=0, sticky="w")

    product_entry = tk.Entry(product_frame, width=25)
    product_entry.grid(row=i, column=1, padx=(5, 10))

    industry_var = tk.StringVar()
    industry_var.set(industry_options[-1])
    industry_menu = tk.OptionMenu(product_frame, industry_var, *industry_options)
    industry_menu.grid(row=i, column=2, padx=(5, 10))
    industry_vars.append(industry_var)
    industry_var.trace_add("write", on_industry_change(i))

    tk.Label(product_frame, text="------>").grid(row=i, column=8, sticky="w")

    tk.Label(product_frame, text="Notes:").grid(row=i, column=10, sticky="w")
    notes_entry = tk.Entry(product_frame, width=25)
    notes_entry.grid(row=i, column=11, padx=(5, 15))

    sample_var = tk.BooleanVar()
    solution_var = tk.BooleanVar()
    quotation_var = tk.BooleanVar()

    tk.Checkbutton(product_frame, text="Sample", variable=sample_var).grid(row=i, column=4, padx=2)
    tk.Checkbutton(product_frame, text="Solution", variable=solution_var).grid(row=i, column=5, padx=2)
    tk.Checkbutton(product_frame, text="Quotation", variable=quotation_var).grid(row=i, column=6, padx=2)

    suggestion_boxes.append(SuggestionBox(product_entry, product_data))
    product_widgets.append((product_entry, industry_var, notes_entry, sample_var, solution_var, quotation_var))

# --- Save Function ---
def save_entry():
    company = entry1.get().strip()
    if not company:
        messagebox.showwarning("Missing Info", "Company name is required.")
        return

    # Check if company is new
    if company not in company_names_from_excel:
        result = messagebox.askyesno("New Company?", "Are you sure the Company Name is new?")
        if not result:
            return  # Don't save

    contact_data = []
    for name_entry, pos_var in contact_entries:
        contact_data.append(name_entry.get().strip())
        contact_data.append(pos_var.get())

    product_rows = []
    for widgets in product_widgets:
        name, industry, notes, sample, solution, quotation = widgets
        if name.get().strip():
            product_rows.append([
                name.get().strip(),
                industry.get(),
                notes.get().strip(),
                sample.get(),
                solution.get(),
                quotation.get()
            ])

    if not product_rows:
        messagebox.showwarning("No Products", "Please enter at least one product.")
        return

    filename = "exhibit_saved.xlsx"
    if os.path.exists(filename):
        wb = load_workbook(filename)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        headers = ["Company", "Contact1", "Position1", "Contact2", "Position2",
                   "Contact3", "Position3", "Contact4", "Position4",
                   "Product", "Industry", "Notes", "Sample", "Solution", "Quotation"]
        sheet.append(headers)

    for row in product_rows:
        sheet.append([company] + contact_data + row)

    wb.save(filename)
    messagebox.showinfo("Saved", "Data saved successfully to exhibit_saved.xlsx")
    clear_form()

def clear_form():
    entry1.delete(0, tk.END)
    for name_entry, pos_var in contact_entries:
        name_entry.delete(0, tk.END)
        pos_var.set(positions[4])

    for widgets in product_widgets:
        name, industry, notes, sample, solution, quotation = widgets
        name.delete(0, tk.END)
        industry.set(industry_options[-1])
        notes.delete(0, tk.END)
        sample.set(False)
        solution.set(False)
        quotation.set(False)

# --- Save and Clear Buttons ---
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

save_button = tk.Button(button_frame, text="Save Entry", command=save_entry, bg="lightgreen", font=bold_font)
save_button.pack(side="left", padx=5)

clear_button = tk.Button(button_frame, text="Clear Form", command=clear_form, bg="red", fg="white", font=bold_font)
clear_button.pack(side="left", padx=5)


root.mainloop()


# In[3]:


pip install pyinstaller


# In[5]:


import os

script_name = "exhibition_form.ipynb"  # Change this to your actual filename

os.system(f"pyinstaller --onefile --windowed {script_name}")


# In[ ]:




